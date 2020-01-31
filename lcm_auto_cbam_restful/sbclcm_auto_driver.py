# -*- coding:UTF-8 -*-

import requests
from requests.adapters import HTTPAdapter
import urllib3
import json
import yaml
import time
import os
import sys
import platform
import subprocess
import random
import paramiko
import traceback
import logging
import shutil
import zipfile
import openpyxl
from openpyxl import load_workbook
from paramiko_expect import SSHClientInteraction

'''
pip install requrired 3rd party modules:
pip3 install requests
pip3 install paramiko
pip3 install openpyxl
pip3 install pexpect
Note: pexpect doesn't work for windows, use winpexpect instead
pip3 install paramiko-expect
'''

'''
Globals
'''
global logger
logger = logging.getLogger()

working_dir = r'D:\Programs\JetBrains\PycharmProjects\py37projects\lcm_auto_cbam_restful'
sig_data_dir = working_dir + r'\data\sig-plane'
media_data_dir = working_dir + r'\data\media-plane'
sig_arts_dir = working_dir + r'\data\sig-plane-arts'
media_arts_dir = working_dir + r'\data\media-plane-arts'
log_file = working_dir + r'\sbclcm_auto.log'

sig_vnfdId = ''
media_vnfdId = ''
sig_vnfdId_SU = ''
media_vnfdId_SU = ''

config_file = 'config.yaml'

def setup_globals():
    global cbam_url, client_id, client_secret, gui_client_id, gui_client_passwd, proxies
    global vnflcm_base_path, vnfpkgm_base_path
    global operationState_list, operation_list
    global sigVersion, sigVersion_SU, mediaVersion, mediaVersion_SU
    global requests_types, requests_retcode
    global os_username, os_passwd
    global vnf_type_list
    global sig_vnfpkg_name_baseload, sig_vnfpkg_name_drbaseload, \
        sig_vnfpkg_name_sutoload, sig_vnfpkg_name_cssutoload, sig_vnfpkg_supported_type
    global media_vnfpkg_name_baseload, media_vnfpkg_name_sutoload, media_vnfpkg_supported_type
    global wkk, ssh_type
    global backup_server_ip, backup_server_login, backup_server_passwd, backup_server_dir, \
        backup_server1, backup_server2, backup_server_creds, backup_server_login_pubkey, \
        local_private_key, backup_server_login_passwd, backup_server_passwd_passwd

    global sig_vnf_name, sig_oam_vip, sig_oama_ip, sig_oamb_ip, sig_oam_login, sig_oam_passwd
    global sig_default_rel, sig_toload_rel, sig_arts_type
    global sig_yact_server_ip, sig_yact_user, sig_yact_passwd, sig_yact_dir
    global sig_dif_name, sig_dr_dif_name, sig_su_dif_name, sig_cssu_dif_name
    global sig_vnfpkg_name, sig_dr_vnfpkg_name, sig_su_vnfpkg_name, sig_cssu_vnfpkg_name
    global sig_server_type_list
    global sig_instantiation_json_file, sig_dr_instantiation_json_file, sig_cssu_instantiation_json_file
    global sig_bulk_conf_url
    global sig_backup_file1_http, sig_backup_file2_http, sig_backup_file1_creds, sig_backup_file2_creds
    global sig_upgrade_file_http, sig_upgrade_file_creds
    global sig_fixed_scm_ip, sig_restore_media_plane
    global sig_backup_zip, sig_cssu_zip, sig_backup_file_name, sig_local_backup_dir
    global sig_su_deft_url, sig_su_deft_key, sig_su_to_image, sig_su_to_version

    global media_vnf_name, media_scma_ip, media_scmb_ip, media_scm_vip
    global media_dif_name, media_su_dif_name
    global media_arts_type
    global media_yact_server_ip, media_yact_user, media_yact_passwd, media_yact_dir, media_openrc_file
    global media_inject_well_known_temp_key, media_script_url
    global media_vnfpkg_name, media_su_vnfpkg_name
    global media_extension_json_file, media_dr_extension_json_file, \
        media_instantiation_json_file, media_dr_instantiation_json_file
    global media_local_backup_dir, media_backup_zip
    global media_lcm_user, media_appl_user, media_appl_passwd
    global media_su_to_image, media_backup_url

    global system
    system = platform.system()
    if system == 'Windows':
        log('Windows system.')
    elif system == 'Linux':
        log('Linux system.')
    else:
        log('Only Windows and Linux system supported. Error out.')
        exit(1)

    try:
        os.chdir(working_dir)
        data = load_yaml_data(file=config_file)

        data_com = data['common']
        data_sig = data['sig_plane']
        data_med = data['media_plane']

        cbam_url            = data_com['cbam_url']
        client_id           = data_com['client_id']
        client_secret       = data_com['client_secret']
        gui_client_id       = data_com['gui_client_id']
        gui_client_passwd   = data_com['gui_client_passwd']
        proxies             = data_com['proxies']

        vnflcm_base_path    = data_com['vnflcm_base_path']
        vnfpkgm_base_path   = data_com['vnfpkgm_base_path']

        operationState_list = data_com['operationState_list']
        operation_list      = data_com['operation_list']

        sigVersion          = data_com['sigVersion']
        sigVersion_SU       = data_com['sigVersion_SU']
        mediaVersion        = data_com['mediaVersion']
        mediaVersion_SU     = data_com['mediaVersion_SU']

        requests_types      = data_com['requests_types']
        requests_retcode    = data_com['requests_retcode']

        os_username         = data_com['os_username']
        os_passwd           = data_com['os_passwd']

        vnf_type_list       = data_com['vnf_type_list']

        sig_vnfpkg_name_baseload    = data_com['sig_vnfpkg_name_baseload']
        sig_vnfpkg_name_drbaseload  = data_com['sig_vnfpkg_name_drbaseload']
        sig_vnfpkg_name_sutoload    = data_com['sig_vnfpkg_name_sutoload']
        sig_vnfpkg_name_cssutoload  = data_com['sig_vnfpkg_name_cssutoload']
        sig_vnfpkg_supported_type   = data_com['sig_vnfpkg_supported_type']

        media_vnfpkg_name_baseload  = data_com['media_vnfpkg_name_baseload']
        media_vnfpkg_name_sutoload  = data_com['media_vnfpkg_name_sutoload']
        media_vnfpkg_supported_type = data_com['media_vnfpkg_supported_type']

        wkk                         = data_com['wkk']
        ssh_type                    = data_com['ssh_type']

        backup_server_ip            = data_com['backup_server_ip']
        backup_server_login         = data_com['backup_server_login']
        backup_server_passwd        = data_com['backup_server_passwd']
        backup_server_dir           = data_com['backup_server_dir']
        backup_server1              = data_com['backup_server1']
        backup_server2              = data_com['backup_server2']
        backup_server_creds         = data_com['backup_server_creds']
        backup_server_login_pubkey  = data_com['backup_server_login_pubkey']
        local_private_key           = data_com['local_private_key']
        backup_server_login_passwd  = data_com['backup_server_login_passwd']
        backup_server_passwd_passwd = data_com['backup_server_passwd_passwd']

        sig_vnf_name        = data_sig['sig_vnf_name']
        sig_oam_vip         = data_sig['sig_oam_vip']
        sig_oama_ip         = data_sig['sig_oama_ip']
        sig_oamb_ip         = data_sig['sig_oamb_ip']
        sig_oam_login       = data_sig['sig_oam_login']
        sig_oam_passwd      = data_sig['sig_oam_passwd']

        sig_default_rel     = data_sig['sig_default_rel']
        sig_toload_rel      = data_sig['sig_toload_rel']
        sig_arts_type       = data_sig['sig_arts_type']

        sig_yact_server_ip  = data_sig['sig_yact_server_ip']
        sig_yact_user       = data_sig['sig_yact_user']
        sig_yact_passwd     = data_sig['sig_yact_passwd']
        sig_yact_dir        = data_sig['sig_yact_dir']

        sig_dif_name        = data_sig['sig_dif_name']
        sig_dr_dif_name     = data_sig['sig_dr_dif_name']
        sig_su_dif_name     = data_sig['sig_su_dif_name']
        sig_cssu_dif_name   = data_sig['sig_cssu_dif_name']

        sig_vnfpkg_name     = data_sig['sig_vnfpkg_name']
        sig_dr_vnfpkg_name  = data_sig['sig_dr_vnfpkg_name']
        sig_su_vnfpkg_name  = data_sig['sig_su_vnfpkg_name']
        sig_cssu_vnfpkg_name    = data_sig['sig_cssu_vnfpkg_name']
        sig_server_type_list    = data_sig['sig_server_type_list']

        sig_instantiation_json_file         = data_sig['sig_instantiation_json_file']
        sig_dr_instantiation_json_file      = data_sig['sig_dr_instantiation_json_file']
        sig_cssu_instantiation_json_file    = data_sig['sig_cssu_instantiation_json_file']

        sig_bulk_conf_url       = data_sig['sig_bulk_conf_url']
        sig_backup_file1_http   = data_sig['sig_backup_file1_http']
        sig_backup_file2_http   = data_sig['sig_backup_file2_http']
        sig_backup_file1_creds  = data_sig['sig_backup_file1_creds']
        sig_backup_file2_creds  = data_sig['sig_backup_file2_creds']
        sig_upgrade_file_http   = data_sig['sig_upgrade_file_http']
        sig_upgrade_file_creds  = data_sig['sig_upgrade_file_creds']

        sig_fixed_scm_ip        = data_sig['sig_fixed_scm_ip']
        sig_restore_media_plane = data_sig['sig_restore_media_plane']

        sig_backup_zip          = data_sig['sig_backup_zip']
        sig_cssu_zip            = data_sig['sig_cssu_zip']
        sig_backup_file_name    = data_sig['sig_backup_file_name']
        sig_local_backup_dir    = data_sig['sig_local_backup_dir']

        sig_su_deft_url         = data_sig['sig_su_deft_url']
        sig_su_deft_key         = data_sig['sig_su_deft_key']
        sig_su_to_image         = data_sig['sig_su_to_image']
        sig_su_to_version       = data_sig['sig_su_to_version']

        media_vnf_name          = data_med['media_vnf_name']
        media_scma_ip           = data_med['media_scma_ip']
        media_scmb_ip           = data_med['media_scmb_ip']
        media_scm_vip           = data_med['media_scm_vip']

        media_dif_name          = data_med['media_dif_name']
        media_su_dif_name       = data_med['media_su_dif_name']
        media_arts_type         = data_med['media_arts_type']

        media_yact_server_ip    = data_med['media_yact_server_ip']
        media_yact_user         = data_med['media_yact_user']
        media_yact_passwd       = data_med['media_yact_passwd']
        media_yact_dir          = data_med['media_yact_dir']
        media_openrc_file       = data_med['media_openrc_file']

        media_inject_well_known_temp_key    = data_med['media_inject_well_known_temp_key']
        media_script_url                    = data_med['media_script_url']

        media_vnfpkg_name       = data_med['media_vnfpkg_name']
        media_su_vnfpkg_name    = data_med['media_su_vnfpkg_name']

        media_extension_json_file           = data_med['media_extension_json_file']
        media_dr_extension_json_file        = data_med['media_dr_extension_json_file']
        media_instantiation_json_file       = data_med['media_instantiation_json_file']
        media_dr_instantiation_json_file    = data_med['media_dr_instantiation_json_file']

        media_local_backup_dir  = data_med['media_local_backup_dir']
        media_backup_zip        = data_med['media_backup_zip']

        media_lcm_user          = data_med['media_lcm_user']
        media_appl_user         = data_med['media_appl_user']
        media_appl_passwd       = data_med['media_appl_passwd']

        media_su_to_image       = data_med['media_su_to_image']
        media_backup_url        = data_med['media_backup_url']

    except Exception:
        traceback.print_exc()
    finally:
        dump_globals()
        log('setup_globals completed.')

def dump_globals():
    log('Start of dumping globals:')
    log('common data:')
    log('cbam_url: ' + cbam_url)
    log('client_id: ' + client_id)
    log('client_secret: ' + client_secret)
    log('gui_client_id: ' + gui_client_id)
    log('gui_client_passwd: ' + gui_client_passwd)
    log('proxies: ' + str(proxies))
    log('vnflcm_base_path: ' + vnflcm_base_path)
    log('vnfpkgm_base_path: ' + vnfpkgm_base_path)
    log('operationState_list: ' + str(operationState_list))
    log('operation_list: ' + str(operation_list))
    log('sigVersion: ' + sigVersion)
    log('sigVersion_SU:' + sigVersion_SU)
    log('mediaVersion: ' + mediaVersion)
    log('mediaVersion_SU: ' + mediaVersion_SU)
    log('requests_types: ' + str(requests_types))
    log('requests_retcode:' + str(requests_retcode))
    log('os_username: ' + os_username)
    log('os_passwd: ' + os_passwd)
    log('vnf_type_list: ' + str(vnf_type_list))
    log('sig_vnfpkg_name_baseload: ' + sig_vnfpkg_name_baseload)
    log('sig_vnfpkg_name_drbaseload: ' + sig_vnfpkg_name_drbaseload)
    log('sig_vnfpkg_name_sutoload: ' + sig_vnfpkg_name_sutoload)
    log('sig_vnfpkg_name_cssutoload: ' + sig_vnfpkg_name_cssutoload)
    log('sig_vnfpkg_supported_type; ' + str(sig_vnfpkg_supported_type))
    log('media_vnfpkg_name_baseload: ' + media_vnfpkg_name_baseload)
    log('media_vnfpkg_name_sutoload: ' + media_vnfpkg_name_sutoload)
    log('media_vnfpkg_supported_type: ' + str(media_vnfpkg_supported_type))
    log('wkk: ' + wkk)
    log('ssh_type: ' + ssh_type)
    log('backup_server_ip: ' + backup_server_ip)
    log('backup_server_login: ' + backup_server_login)
    log('backup_server_passwd: ' + backup_server_passwd)
    log('backup_server_dir: ' + backup_server_dir)
    log('backup_server1: ' + backup_server1)
    log('backup_server2: ' + backup_server2)
    log('backup_server_creds: ' + backup_server_creds)
    log('backup_server_login_pubkey: ' + backup_server_login_pubkey)
    log('local_private_key: ' + local_private_key)
    log('backup_server_login_passwd: ' + backup_server_login_passwd)
    log('backup_server_passwd_passwd: ' + backup_server_passwd_passwd)
    log('sig_plane data:')
    log('sig_vnf_name: ' + sig_vnf_name)
    log('sig_oam_vip: ' + sig_oam_vip)
    log('sig_oama_ip: ' + sig_oama_ip)
    log('sig_oamb_ip: ' + sig_oamb_ip)
    log('sig_oam_login: ' + sig_oam_login)
    log('sig_oam_passwd: ' + sig_oam_passwd)
    log('sig_default_rel: ' + sig_default_rel)
    log('sig_toload_rel: ' + sig_toload_rel)
    log('sig_arts_type: ' + str(sig_arts_type))
    log('sig_yact_server_ip: ' + sig_yact_server_ip)
    log('sig_yact_user: ' + sig_yact_user)
    log('sig_yact_passwd: ' + str(sig_yact_passwd))
    log('sig_yact_dir: ' + sig_yact_dir)
    log('sig_dif_name: ' + sig_dif_name)
    log('sig_dr_dif_name: ' + sig_dr_dif_name)
    log('sig_su_dif_name: ' + sig_su_dif_name)
    log('sig_cssu_dif_name: ' + sig_cssu_dif_name)
    log('sig_vnfpkg_name: ' + sig_vnfpkg_name)
    log('sig_dr_vnfpkg_name: ' + sig_dr_vnfpkg_name)
    log('sig_su_vnfpkg_name: ' + sig_su_vnfpkg_name)
    log('sig_cssu_vnfpkg_name: ' + sig_cssu_vnfpkg_name)
    log('sig_server_type_list: ' + str(sig_server_type_list))
    log('sig_instantiation_json_file: ' + sig_instantiation_json_file)
    log('sig_dr_instantiation_json_file: ' + sig_dr_instantiation_json_file)
    log('sig_cssu_instantiation_json_file: ' + sig_cssu_instantiation_json_file)
    log('sig_bulk_conf_url: ' + sig_bulk_conf_url)
    log('sig_backup_file1_http: ' + sig_backup_file1_http)
    log('sig_backup_file2_http: ' + sig_backup_file2_http)
    log('sig_backup_file1_creds: ' + sig_backup_file1_creds)
    log('sig_backup_file2_creds: ' + sig_backup_file2_creds)
    log('sig_upgrade_file_http: ' + sig_upgrade_file_http)
    log('sig_upgrade_file_creds: ' + sig_upgrade_file_creds)
    log('sig_fixed_scm_ip: ' + sig_fixed_scm_ip)
    log('sig_restore_media_plane: ' + sig_restore_media_plane)
    log('sig_backup_zip: ' + sig_backup_zip)
    log('sig_cssu_zip: ' + sig_cssu_zip)
    log('sig_backup_file_name: ' + sig_backup_file_name)
    log('sig_local_backup_dir: ' + sig_local_backup_dir)
    log('sig_su_deft_url: ' + sig_su_deft_url)
    log('sig_su_deft_key: ' + str(sig_su_deft_key))
    log('sig_su_to_image: ' + sig_su_to_image)
    log('sig_su_to_version: ' + sig_su_to_version)
    log('media_plane data:')
    log('media_vnf_name: ' + media_vnf_name)
    log('media_scma_ip: ' + media_scma_ip)
    log('media_scmb_ip: ' + media_scmb_ip)
    log('media_scm_vip: ' + media_scm_vip)
    log('media_dif_name: ' + media_dif_name)
    log('media_su_dif_name: ' + media_su_dif_name)
    log('media_arts_type: ' + str(media_arts_type))
    log('media_yact_server_ip: ' + media_yact_server_ip)
    log('media_yact_user: ' + media_yact_user)
    log('media_yact_passwd: ' + media_yact_passwd)
    log('media_yact_dir:' + media_yact_dir)
    log('media_openrc_file: ' + media_openrc_file)
    log('media_inject_well_known_temp_key: ' + media_inject_well_known_temp_key)
    log('media_script_url: ' + media_script_url)
    log('media_vnfpkg_name: ' + media_vnfpkg_name)
    log('media_su_vnfpkg_name: ' + media_su_vnfpkg_name)
    log('media_extension_json_file: ' + media_extension_json_file)
    log('media_dr_extension_json_file: ' + media_dr_extension_json_file)
    log('media_instantiation_json_file: ' + media_instantiation_json_file)
    log('media_dr_instantiation_json_file: ' + media_dr_instantiation_json_file)
    log('media_local_backup_dir: ' + media_local_backup_dir)
    log('media_backup_zip: ' + media_backup_zip)
    log('media_lcm_user: ' + media_lcm_user)
    log('media_appl_user: ' + media_appl_user)
    log('media_appl_passwd: ' + media_appl_passwd)
    log('media_su_to_image: ' + media_su_to_image)
    log('media_backup_url: ' + media_backup_url)
    log('End of dumping globals.')


'''
Utils Functions
'''
def log(msg, level=logging.INFO):
    if level == logging.INFO:
        logger.info(msg)

def log_print(msg):
    print(msg)

def print_response(response):
    print('response: ', response)
    print('response.headers: ', response.headers)
    print('response.text: ', response.text)

def print_response_details(response):
    print('response: ', response)
    if response.headers:
        print('resposne.headers:')
        for i, j in response.headers.items():
            print(i, j)
    if response.request.headers:
        print('response.request.headers:')
        for i, j in response.request.headers.items():
            print(i, j)
    if response.text:
        print('resposne.text:\n', response.text)
        # for i, j in response.text.json():
        #     print(i, j)

def logger_response(response):
    log(response)
    log('response.headers: ' + response.headers)
    log('response.text: ' + response.text)

def dump_response_data(response, funcname):
    log('Now at function:' + funcname)
    log(response.text)

def load_yaml_data(file=None):
    if file is None:
        log('Please specify the yaml file.')
        exit(1)
    with open(file, 'r') as f:
        data = f.read()
    data = yaml.load(data)
    return data

def dunp_yaml_data(data='', file=''):
    if data and file:
        with open(file, 'w') as f:
            yaml.dump(data, f)
    else:
        log('No data to dump. Error out.')
        exit(1)

def ssh_command(cmd, ip, login, pass_key, type):
    log('In Func: ' + sys._getframe().f_code.co_name)
    log('cmd: ' + cmd)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    if type == 'passwd':
        ssh.connect(hostname=ip, port=22, username=login, password=pass_key)
    elif type == 'pubkey':
        private_key = paramiko.RSAKey.from_private_key_file(pass_key)
        ssh.connect(hostname=ip, port=22, username=login, pkey=private_key)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    res, err = stdout.read(), stderr.read()
    result = res if res else err
    ssh.close()
    return result.decode()

def ssh_command_passwd(cmd, ip, login, passwd):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    res, err = stdout.read(), stderr.read()
    result = res if res else err
    ssh.close()
    return result.decode()

def ssh_command_pubkey(cmd, ip, login, privkey):
    private_key = paramiko.RSAKey.from_private_key_file(privkey)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, pkey=private_key)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    res, err = stdout.read(), stderr.read()
    result = res if res else err
    ssh.close()
    return result.decode()

def ssh_scp_put(ip, login, passwd, local_file, remote_file):
    log('In Func: ' + sys._getframe().f_code.co_name)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.put(local_file, remote_file)
    ssh.close()

def ssh_scp_get(ip, login, passwd, remote_file, local_file):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.get(remote_file, local_file)
    ssh.close()

def ssh_paramiko_pexpect(cmd, ip, login, passwd):
    """
    https://github.com/fgimian/paramiko-expect/blob/master/examples/paramiko_expect-demo.py
    :param cmd: not used in the func
    :param ip: act scm vip
    :param login: appl user
    :param passwd: appl user passwd
    :return: the output of vi node
    # hostname = '10.75.44.12'
    # hostname = '100.69.127.152'
    # username = 'diag'
    # password = '-assured'
    # PROMPT = r'diag:(main|diag)(:vMG.*?)?#'
    """
    PROMPT = '.*ACT-SCM.*'
    hostname = ip
    username = login
    password = passwd
    try:
        client = paramiko.SSHClient()
        # Set SSH key parameters to auto accept unknown hosts
        client.load_system_host_keys()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        # Connect to the host
        client.connect(hostname=hostname, username=username, password=password)
        # Create a client interaction class which will interact with the host
        # with SSHClientInteraction(client, timeout=10, display=True) as interact:
        with SSHClientInteraction(client, timeout=10, buffer_size=10240, display=True) as interact:
            interact.expect(PROMPT)
            # Run the first command and capture the cleaned output, if you want
            # the output without cleaning, simply grab current_output instead.
            cmd = 'vi node'
            interact.send(cmd)
            interact.expect(PROMPT, timeout=10)
            cmd_output_vi_node = interact.current_output_clean
            cmd = 'logout'
            interact.send(cmd)
            interact.expect()
    except Exception:
        traceback.print_exc()
    finally:
        try:
            client.close()
        except Exception:
            pass
    return cmd_output_vi_node

def get_token():
    response = None
    headers = {'Accept': '*/*', 'Content-Type': 'application/x-www-form-urlencoded'}
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret
    }
    s = requests.Session()
    s.mount('http://', HTTPAdapter(max_retries=10))
    s.mount('https://', HTTPAdapter(max_retries=10))
    address = cbam_url + '/auth/realms/cbam/protocol/openid-connect/token'
    try:
        response = s.post(address, data=data, headers=headers, timeout=240, verify=False, proxies=proxies)
    except requests.exceptions.RequestException as e:
        print(e)
    # check response code
    if response:
        if response.status_code != 200:
            log('Func get_token failed. Response status_code: ' + str(response.status_code))
            log('Response text: ' + response.text)
            exit(1)
        return 'bearer ' + response.json()["access_token"]
    else:
        log('Func get_token failed. Response: None.')
        return None

# requests_types = ['GET', 'POST', 'DELETE', 'PATCH']
def send_request(type='GET', url='', data=None, expect_code=200, timeout=180, retries=10, func=''):
    log('In Func:' + sys._getframe().f_code.co_name)
    response = None
    # check params
    if data is None:
        data = {}
    if type not in requests_types:
        log('Unsupported request type.')
        exit(1)
    log('type: ' + type)
    log('url: ' + url)
    log('data: ' + str(data))
    log('expect_code: ' + str(expect_code))
    # form request then send the request, get response
    token_bear = get_token()
    s = requests.Session()
    s.mount('http://', HTTPAdapter(max_retries=retries))
    s.mount('https://', HTTPAdapter(max_retries=retries))
    try:
        if type == 'GET':
            headers = {"Authorization": token_bear}
            response = s.get(url, headers=headers, timeout=timeout, verify=False, proxies=proxies)
        if type == 'POST':
            headers = {"Authorization": token_bear, "Content-Type": "application/json"}
            response = s.post(url, headers=headers, json=data, timeout=timeout, verify=False, proxies=proxies)
        if type == 'DELETE':
            headers = {"Authorization": token_bear}
            response = s.delete(url, headers=headers, timeout=timeout, verify=False, proxies=proxies)
        if type == 'PATCH':
            headers = {"Authorization": token_bear, "Content-Type": "application/json"}
            response = s.patch(url, headers=headers, json=data, timeout=timeout, verify=False, proxies=proxies)
    except requests.exceptions.RequestException as e:
        print(e)
    # check response code
    if response:
        if response.status_code != expect_code:
            log('Response failed at: ' + func)
            log('Response status_code: ' + str(response.status_code))
            log('Response text: ' + response.text)
            exit(1)
        return response
    else:
        log('Func send_request failed. Response: None.')
        return None

def get_vnfs():
    token_bear = get_token()
    headers = {'Authorization': token_bear}
    address = cbam_url + '/vnflcm/v1' + '/vnf_instances'
    response = requests.get(address, headers=headers, verify=False, proxies=proxies)
    data = json.loads(response.text)
    for vnf in data:
        for i, j in vnf.items():
            print(i, ':', j)

def get_vnf_packages():
    token_bear = get_token()
    headers = {'Authorization': token_bear}
    address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages'
    response = requests.get(address, headers=headers, verify=False, proxies=proxies)
    dump_response_data(response, 'get_vnf_packages')


'''
Utils Classes
'''
class SetUps(object):
    pass

class CleanUps(object):
    pass

class LcmUtils(object):
    # LcmLcpUtils:
    # read config data from local yaml file
    # oam swithover
    # init services
    # reboot hosts
    # run specific sripts
    # ...
    def __init__(self):
        self.sig_oam_vip    = sig_oam_vip
        self.sig_oama_ip    = sig_oama_ip
        self.sig_oamb_ip    = sig_oamb_ip
        self.sig_oam_login  = sig_oam_login
        self.sig_oam_passwd = sig_oam_passwd
        self.sig_host_dict  = {}

    def prep_local_known_hosts(self, ip=''):
        """
        # Need to rm the entry in known_hosts file on local pc.
        :param ip: oam IP or scm IP
        """
        log('In Func:' + sys._getframe().f_code.co_name)
        log('Rmove the existing entry of IP: {0} in /c/Users/shawnx/.ssh/known_hosts.'.format(ip))
        known_hosts_file = r'C:\Users\shawnx\.ssh\known_hosts'
        tmp_list = []
        try:
            with open(known_hosts_file, 'r') as f:
                tmp_list = f.readlines()
            for val in tmp_list:
                if ip in val:
                    tmp_list.remove(val)
            with open(known_hosts_file, 'w') as f:
                f.writelines(tmp_list)
        except Exception:
            traceback.print_exc()

    def get_sig_host_dict(self):
        """
        Get a dict of flexible name to host name map
        Set sig_host_dict
        Example:
            {'sbclcm03-oam-a': 'sbclcm03-s00c01h0',
            'sbclcm03-sc1-a': 'sbclcm03-s00c02h0',
            'sbclcm03-cfed-a': 'sbclcm03-s00c03h0',
            'sbclcm03-bgc1-a': 'sbclcm03-s00c04h0',
            'sbclcm03-fwa-a': 'sbclcm03-s00c05h0',
            'sbclcm03-ccf-a': 'sbclcm03-s00c06h0',
            'sbclcm03-fwp-a': 'sbclcm03-s00c07h0',
            'sbclcm03-dfed-a': 'sbclcm03-s00c08h0',
            'sbclcm03-oam-b': 'sbclcm03-s01c01h0',
            'sbclcm03-sc1-b': 'sbclcm03-s01c02h0',
            'sbclcm03-cfed-b': 'sbclcm03-s01c03h0',
            'sbclcm03-bgc1-b': 'sbclcm03-s01c04h0',
            'sbclcm03-fwa-b': 'sbclcm03-s01c05h0',
            'sbclcm03-ccf-b': 'sbclcm03-s01c06h0',
            'sbclcm03-fwp-b': 'sbclcm03-s01c07h0',
            'sbclcm03-dfed-b': 'sbclcm03-s01c08h0'}
        """
        self.sig_host_dict = {}
        mtce_host_file = '/var/opt/lib/mtce/data/mtce_host.data'
        ip = self.sig_oama_ip
        login = self.sig_oam_login
        passwd = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = 'cat ' + mtce_host_file
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))
        if 'APPLICATION: sbc' not in result:
            log('Failed to get data in {0}.'.format(mtce_host_file))
            exit(1)
        data = result.split('\n')
        for line in data:
            if '_GUEST' in line:
                tmp_list = line.split(';')
                val = tmp_list[0]
                key = tmp_list[12]
                self.sig_host_dict[key] = val
        if self.sig_host_dict == {}:
            log('Failed to get sig_host_dict. Error out.')
            exit(1)
        log(str(self.sig_host_dict))

    def run_remcli_cmd(self, remcli_cmd='su 0 0 0'):
        """
        Get REMcli output per cmd
        :param remcli_cmd: the cmd feed to REMcli
        Example:
        # 4143 or AC:  query All Cars
        # 4443 or DC:  query All Degraded Cars
        # 6163 or ac:  query All Cars on a service member
        # 6173 or as:  status of All Service members
        # 6463 or dc:  query Degraded Cars on a service member
        # 6466 or df:  status of DiskFull service members
        # 7063 or pc:  Print Counters for inits and switchovers
        # 7263 or rc:  Reset Counters for inits and switchovers
        # 7068 or ph:  Print Heartbeat loss counters
        # 7273 or rs:  REMc Switchover. Sends cmd to ACT REMc
        # 7375 or su:  status of diskless service members, used by SU
        # 7376 or sv:  Software Version of all service members (or use 8386)
        """
        ip      = self.sig_oam_vip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/bin/REMcli ' + remcli_cmd
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))

    def oam_switchover(self):
        """
        For oam switchover, need to use oam external fixed IP,
        as there would be 'client_loop: send disconnect: Connection reset by peer'
        if using external floating IP
        """
        ip      = self.sig_oama_ip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/MIcmd state vc'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))
        cmd = '/opt/LSS/sbin/MIcmd switch'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))
        str_success = 'MI virtual cluster switched over successfully'
        if str_success in result:
            log(str_success)
        else:
            # For now, just ignore the switchover failure
            # TODO: check failure str and retry switchover again
            log('MIcmd switchover failed.')

    def oam_failover(self):
        """
        1. check which oam side is A
        2. reboot oam A
        """
        ip      = self.sig_oama_ip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/MIcmd state vc'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))
        cmd = '/opt/LSS/bin/MIvmstate'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))
        if 'A' in result:
            log('OAM-A state: A')
        elif 'S' in result:
            log('OAM-A state: S')
            ip = self.sig_oamb_ip
        else:
            log('OAM-A state not A or S. Need to check in system. Error out.')
            exit(1)
        cmd = 'reboot'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log('Wait 300s for system to be stablized.')
        time.sleep(300)
        cmd = '/opt/LSS/sbin/MIcmd state vc'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))

    def oam_reboot(self, oam_side='A'):
        """
        Reboot oam, side A or B: oam-a or oam-b
        :param oam_side: 'A' or 'B
        """
        if oam_side == 'A':
            ip = self.sig_oama_ip
        elif oam_side == 'B':
            ip = self.sig_oamb_ip
        else:
            log('oam_side only supports A or B. Error out.')
            exit(1)
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/MIcmd state vc'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))
        cmd = 'reboot'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log('Wait 300s for system to be stablized.')
        time.sleep(300)
        cmd = '/opt/LSS/sbin/MIcmd state vc'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))

    def vm_reboot(self, hostname='oam-a'):
        """
        Reboot specific vm
        First login oam
        Then ssh hostbname reboot
        :param hostname: flexible host name
        """
        ip = self.sig_oam_vip
        login = self.sig_oam_login
        passwd = self.sig_oam_passwd
        sshtype = 'passwd'
        self.get_sig_host_dict()
        # get full flexible host name
        hname = None
        for name in self.sig_host_dict.keys():
            if hostname in name:
                hname = name
        if hname:
            cmd = 'ssh ' + hname + ' reboot'
            result = ssh_command(cmd, ip, login, passwd, sshtype)
            # output is like:
            # cmd: ssh sbclcm03-sc1-a reboot
            # Connection to sbclcm03-sc1-a closed by remote host.
            log(str(result))
        else:
            log('Failed to get hostname of the VM to be rebooted. Error out.')
            exit(1)

    def health_check(self):
        """
        Run health on oam
        """
        ip      = self.sig_oam_vip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/health --test all'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))

    def sbc_health_check(self):
        """
        Run health on oam
        """
        ip      = self.sig_oam_vip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/sbc_health --test all'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))

    def sig_health_check(self):
        """
        Run health on oam
        """
        ip      = self.sig_oam_vip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/sbc_health --test sp'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))

    def media_health_check(self):
        """
        Run health on oam
        """
        ip      = self.sig_oam_vip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/sbc_health --test mp'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))

    def sig_quick_health_check(self):
        """
        Run health on oam
        """
        ip      = self.sig_oam_vip
        login   = self.sig_oam_login
        passwd  = self.sig_oam_passwd
        sshtype = 'passwd'
        cmd = '/opt/LSS/sbin/sbc_health --test sp quick'
        result = ssh_command(cmd, ip, login, passwd, sshtype)
        log(str(result))


'''
VNF Packages
'''
class VnfPkgs(object):
    def __init__(self):
        self.sig_vnfpkg_baseload    = sig_data_dir + '\\' + sig_vnfpkg_name_baseload
        self.sig_vnfpkg_drbaseload  = sig_data_dir + '\\' + sig_vnfpkg_name_drbaseload
        self.sig_vnfpkg_sutoload    = sig_data_dir + '\\' + sig_vnfpkg_name_sutoload
        self.sig_vnfpkg_cssutoload  = sig_data_dir + '\\' + sig_vnfpkg_name_cssutoload
        self.media_vnfpkg_baseload  = media_data_dir + '\\' + media_vnfpkg_name_baseload
        self.media_vnfpkg_sutoload  = media_data_dir + '\\' + media_vnfpkg_name_sutoload
        self.vnfpkgs_list = []

    def get_vnfpkgs(self):
        self.vnfpkgs_list = []
        token_bear = get_token()
        headers = {'Authorization': token_bear}
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages'
        response = requests.get(address, headers=headers, verify=False, proxies=proxies)
        data = json.loads(response.text)
        if len(data) != 0:
            for vp in data:
                vnfpkg = VnfPkg()
                vnfpkg.id = vp['id']
                vnfpkg.vnfdId = vp['vnfdId']
                vnfpkg.vnfProvider = vp['vnfProvider']
                vnfpkg.vnfProductName = vp['vnfProductName']
                vnfpkg.vnfSoftwareVersion = vp['vnfSoftwareVersion']
                vnfpkg.vnfdVersion = vp['vnfdVersion']
                vnfpkg.onboardingState = vp['onboardingState']
                vnfpkg.operationalState = vp['operationalState']
                vnfpkg.userDefinedData = vp['userDefinedData']
                vnfpkg._links = vp['_links']
                self.vnfpkgs_list.append(vnfpkg)

    def dump_vnfpkgs(self):
        if len(self.vnfpkgs_list) != 0:
            for vp in self.vnfpkgs_list:
                vp.dump_vnfpkg()

    def print_vnfpkgs(self):
        if len(self.vnfpkgs_list) != 0:
            for vp in self.vnfpkgs_list:
                vp.print_vnfpkg()

    def upload_vnfpkg(self, sig_media='sig', version = '', type='instantiation'):
        vnfpkg = ''
        id = ''
        if sig_media not in ['sig', 'media']:
            log('upload_vnfpkg: need to be either sig or media.')
            exit(1)
        if sig_media == 'sig':
            if type not in sig_vnfpkg_supported_type:
                log('vnfpkg type not supported. Need to be one of ' +
                            str(sig_vnfpkg_supported_type))
            if version == sigVersion:
                if type == 'instantiation':
                    vnfpkg = self.sig_vnfpkg_baseload
                elif type == 'dr':
                    vnfpkg = self.sig_vnfpkg_drbaseload
            elif version == sigVersion_SU:
                if type == 'su':
                    vnfpkg = self.sig_vnfpkg_sutoload
                elif type == 'cssu':
                    vnfpkg = self.sig_vnfpkg_cssutoload
        if sig_media == 'media':
            if type not in media_vnfpkg_supported_type:
                log('vnfpkg type not supported. Need to be one of ' +
                            str(media_vnfpkg_supported_type))
            if version == mediaVersion:
                if type == 'instantiation':
                    vnfpkg = self.media_vnfpkg_baseload
            elif version == mediaVersion_SU:
                if type == 'su':
                    vnfpkg = self.media_vnfpkg_sutoload
        log('vnfpkg: ' + vnfpkg)
        # To-do: add check if the vnfpkg has already been uploaded
        id = self.get_vnfpkg_by_swversion(version)
        if id:
            # delete the vnf pkg
            log('Delete existing vnf pkg of same version, id: ' + id)
            self.delete_vnfpkg_by_id(id)
        # Now upload the vnd pkg
        token_bear = get_token()
        # First POST vnf_packages to generate one new vnf pkg
        headers = {'Authorization': token_bear}
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages'
        response = requests.post(address, headers=headers, verify=False, proxies=proxies)
        if response.status_code != 201:
            log('vnfpkg creation failed.')
            exit(1)
        data = json.loads(response.text)
        id = data['id']
        log('New vnfpkg id: ' + id)
        log('New vnfpkg info: ' + response.text)
        #Then PUT content to content of new vnf pkg
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages/' + id + '/package_content'
        headers = {'Authorization': token_bear, 'Content-Type': 'application/octet-stream'}
        response = requests.put(address, data=open(vnfpkg, 'rb'), headers=headers, verify=False, proxies=proxies)
        if response.status_code != 202:
            log('vnfpkg content upload failed.')
            # Here need to delete the newly created vnf pkg
            headers = {'Authorization': token_bear}
            address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages/' + id
            requests.delete(address, headers=headers, verify=False, proxies=proxies)
            exit(1)
        log('VNF PKG uploaded successfully.')
        time.sleep(10)

    def get_vnfpkg_by_swversion(self, swversion):
        self.get_vnfpkgs()
        for vp in self.vnfpkgs_list:
            if vp.vnfSoftwareVersion.endswith(swversion):
                log('id in get_vnfpkg_by_swversion: ' + vp.id)
                return vp.id

    def delete_vnfpkg_by_swversion(self, swversion):
        id = self.get_vnfpkg_by_swversion(swversion)
        self.delete_vnfpkg_by_id(id)

    def delete_vnfpkg_by_id(self, id):
        token_bear = get_token()
        headers = {'Authorization': token_bear}
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages/' + id
        response = requests.get(address, headers=headers, verify=False, proxies=proxies)
        log('The vnfpkg to be deleted: ' + response.text)
        requests.delete(address, headers=headers, verify=False, proxies=proxies)


'''
VNF Package
'''
class VnfPkg(object):
    def __init__(self):
        self.id = ''
        self.vnfdId = ''
        self.vnfProvider = ''
        self.vnfProductName = ''
        self.vnfSoftwareVersion = ''
        self.vnfdVersion = ''
        self.onboardingState = ''
        self.operationalState = ''
        self.userDefinedData = ''
        self._links = ''

    def get_id(self):
        return self.id

    def get_vnfdId(self):
        return self.vnfdId

    def get_vnfProductName(self):
        return self.vnfProductName

    def get_vnfSoftwareVersion(self):
        return self.vnfSoftwareVersion

    def get_vnfdVersion(self):
        return self.vnfdVersion

    def dump_vnfpkg(self):
        log('id: ' + self.id)
        log('vnfdId: ' + self.vnfdId)
        log('vnfProvider: ' + self.vnfProvider)
        log('vnfProductName: ' + self.vnfProductName)
        log('vnfSoftwareVersion: ' + self.vnfSoftwareVersion)
        log('vnfdVersion: ' + self.vnfdVersion)
        log('onboardingState: ' + self.onboardingState)
        log('operationalState: ' + self.operationalState)
        log('userDefinedData: ' + str(self.userDefinedData))
        log('_links: ' + str(self._links))

    def print_vnfpkg(self):
        print('id: ', self.id)
        print('vnfdId: ', self.vnfdId)
        print('vnfProvider: ', self.vnfProvider)
        print('vnfProductName: ', self.vnfProductName)
        print('vnfSoftwareVersion: ', self.vnfSoftwareVersion)
        print('vnfdVersion: ', self.vnfdVersion)
        print('onboardingState: ', self.onboardingState)
        print('operationalState: ', self.operationalState)
        print('userDefinedData: ', self.userDefinedData)
        print('_links: ', self._links)
        print('\n')

'''
VNF Package Operations
'''
def sigvnf_UploadVnfpkg(swVersion, type):
    vnfpkgs = VnfPkgs()
    # vnfpkgs.get_vnfpkgs()
    vnfpkgs.upload_vnfpkg(sig_media='sig', version=swVersion, type=type)
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()

def sigvnf_DeleteVnfpkg(swVersion):
    vnfpkgs = VnfPkgs()
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()
    vnfpkgs.delete_vnfpkg_by_swversion(swVersion)

# Media VNF Package Operations
def mediavnf_UploadVnfpkg(swVersion, type):
    vnfpkgs = VnfPkgs()
    # vnfpkgs.get_vnfpkgs()
    vnfpkgs.upload_vnfpkg(sig_media='media', version=swVersion, type=type)
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()

def mediavnf_DeleteVnfpkg(swVersion):
    vnfpkgs = VnfPkgs()
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()
    vnfpkgs.delete_vnfpkg_by_swversion(swVersion)


'''
SBC VNF
'''
class SBCVnf(object):
    def __init__(self, vnfdId, apiVersion, name, description):
        self.name = name
        self.description = description
        self.vnfdId = vnfdId
        self.apiVersion = apiVersion
        self.id = ''
        self.vnf_type = ''
        self.opStatus = ''
        self.aspectId = ''
        self.scaleLevel = 1
        self.maxScaleLevel = 5
        self.operationState = ''
        self.vnfcInstanceId_List = []
        self.instantiate_file = ''
        self.extension_file = ''
        self.dr_extension_file = ''
        self.cssu_instantiate_file = ''
        self.dr_instantiate_file = ''
        self.vnflcm_base_path = cbam_url + '/vnflcm/v1'
        self.vnflcm_instances = self.vnflcm_base_path + '/vnf_instances/'
        self.vnfpkgm_base_path = cbam_url + '/vnfpkgm/v1'
        self.vnflcm_create = vnflcm_base_path + '/vnf_instances'

    def set_vnflcm_apis(self):
        self.operationState             = vnflcm_base_path + '/vnf_lcm_op_occs'
        self.vnflcm_create              = vnflcm_base_path + '/vnf_instances'
        self.vnflcm_instantiate         = self.vnflcm_instances + self.id + '/instantiate'
        self.vnflcm_terminate           = self.vnflcm_instances + self.id + '/terminate'
        self.vnflcm_delete              = self.vnflcm_instances + self.id
        self.vnflcm_get                 = self.vnflcm_instances + self.id
        self.vnflcm_modify              = self.vnflcm_instances + self.id
        self.vnflcm_scale               = self.vnflcm_instances + self.id + '/scale'
        self.vnflcm_heal                = self.vnflcm_instances + self.id + '/heal'
        self.vnflcm_upgrade             = self.vnflcm_instances + self.id + '/upgrade'

    def write_status(self):
        pass

    def get_vnfdId(self):
        return self.vnfdId

    def get_id(self):
        return self.id

    def set_id(self, id):
        self.id = id

    def create_vnf(self):
        data = {
            "vnfdId": self.vnfdId,
            "vnfInstanceName": self.name,
            "vnfInstanceDescription": self.description
        }
        response = send_request(type='POST', url=self.vnflcm_create, data=data,
                                expect_code=201, func=sys._getframe().f_code.co_name)
        dump_response_data(response, 'create_vnf')
        data = json.loads(response.text)
        self.id = data['id']
        log('create_vnf:id: '+ self.get_id())
        # self.set_vnflcm_apis()

    def modify_extension_vnf(self):
        data = json.load(open(self.extension_file, "rb"))
        response = send_request(type='PATCH', url=self.vnflcm_modify, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def dr_modify_extension_vnf(self):
        data = json.load(open(self.dr_extension_file, "rb"))
        response = send_request(type='PATCH', url=self.vnflcm_modify, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def instantiate_vnf(self):
        data = json.load(open(self.instantiate_file, "rb"))
        response = send_request(type='POST', url=self.vnflcm_instantiate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def dr_instantiate_vnf(self):
        data = json.load(open(self.dr_instantiate_file, "rb"))
        response = send_request(type='POST', url=self.vnflcm_instantiate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    #MODIFY use PATCH
    def modify_auto_backup_vnf(self, periodInSeconds=86220, enabled=False):
        data = {
            "vnfConfigurableProperties": {
                "operation_triggers": {
                    "auto_backup": {
                        "periodInSeconds": periodInSeconds,
                        "enabled": enabled
                    }
                }
            }
        }
        response = send_request(type='PATCH', url=self.vnflcm_modify, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def modify_auto_scale_vnf(self, periodInSeconds=600, enabled=False):
        data = {
            "vnfConfigurableProperties": {
                "operation_triggers": {
                    "auto_scale": {
                        "periodInSeconds": periodInSeconds,
                        "enabled": enabled
                    }
                }
            }
        }
        response = send_request(type='PATCH', url=self.vnflcm_modify, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def terminate_vnf(self):
        data = {
            "terminationType": "FORCEFUL"
        }
        response = send_request(type='POST', url=self.vnflcm_terminate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def delete_vnf(self):
        response = send_request(type='DELETE', url=self.vnflcm_delete, data=None,
                                expect_code=204, func=sys._getframe().f_code.co_name)

    def upgrade_vnf(self, vnfdId):
        data = {
            "vnfdId": vnfdId,
            "apiVersion": "4.0"
        }
        response = send_request(type='POST', url=self.vnflcm_upgrade, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def get_vnf(self):
        response = send_request(type='GET', url=self.vnflcm_get, data=None,
                                expect_code=200, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        return data

    def scale_vnf(self, scaleType='SCALE_OUT', step=1, aspectId="sc_Aspect"):
        data = {
            "type": scaleType,
            "aspectId": aspectId,
            "numberOfSteps": step
        }
        response = send_request(type='POST', url=self.vnflcm_scale, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def get_scaleStatus(self, aspectId="sc_Aspect"):
        data = self.get_vnf()
        scaleStatus = data['instantiatedVnfInfo']['scaleStatus']
        log('scaleStatus: ' + str(scaleStatus))
        for ss in scaleStatus:
            if ss['aspectId'] == aspectId:
                self.scaleLevel=ss['scaleLevel']
                self.maxScaleLevel=ss['maxScaleLevel']
                self.aspectId = ss['aspectId']
                log('scaleLevel: ' + str(self.scaleLevel))
                log('maxScaleLevel: ' + str(self.maxScaleLevel))
                log('aspectId: ' + str(self.aspectId))

    def get_vnfcInstanceId_List(self):
        """
        This is for api v4, used by signaling plane
        """
        data = self.get_vnf()
        for vnfcinfo in data['instantiatedVnfInfo']['vnfcResourceInfo']:
            self.vnfcInstanceId_List.append(vnfcinfo['id'])
        log('vnfcInstanceId_List: '+ str(self.vnfcInstanceId_List))

    def heal_vnf(self, vnfcInstanceId=None, cause=''):
        if vnfcInstanceId is None:
            vnfcInstanceId = ['OAM.NOKIA-LCP-VMA']
        data = {
            "cause": cause,
            "vnfcInstanceIds": vnfcInstanceId,
            "additionalParams": {"monitorRetries": 90, "monitorDelay": 20}
        }
        response = send_request(type='POST', url=self.vnflcm_heal, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def get_latestOperation(self):
        response = send_request(type='GET', url=self.operationState, data=None,
                                expect_code=200, timeout=10, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        for item in data:
            if item['vnfInstanceId'] == self.id:
                return item
        return None

    def get_latestOperationState(self):
        response = send_request(type='GET', url=self.operationState, data=None,
                                expect_code=200, timeout=10, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        for item in data:
            if item['vnfInstanceId'] == self.id:
                return item['operationState']
        return None

    def get_operationState(self, operation, operationName):
        self.opStatus = ''
        response = send_request(type='GET', url=self.operationState, data=None,
                                expect_code=200, timeout=60, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        for item in data:
            # Need to find a way to use operationName
            if (item['vnfInstanceId'] == self.id) and (item['operation'] == operation):
                self.opStatus = item['operationState']
                break
        return self.opStatus

'''
Sig VNF
'''
class SigVnf(SBCVnf):
    def __init__(self, vnfdId, apiVersion='4', name='sbclcm-sig-plane',
                 description='SBC LCM Sig-Plane VNF for Auto Test via CBAM REST APIs'):
        super().__init__(vnfdId, apiVersion, name, description)
        self.name = name
        self.description = description
        self.vnfdId = vnfdId
        self.id = ''
        self.vnf_type = 'sig'
        self.instantiate_file       = sig_data_dir + '\\' + sig_instantiation_json_file
        self.dr_instantiate_file    = sig_data_dir + '\\' + sig_dr_instantiation_json_file
        self.cssu_instantiate_file  = sig_data_dir + '\\' + sig_cssu_instantiation_json_file

    def set_vnflcm_apis(self):
        super().set_vnflcm_apis()
        self.vnflcm_custom_backup               = self.vnflcm_instances + self.id + '/custom/backup'
        self.vnflcm_custom_connect              = self.vnflcm_instances + self.id + '/custom/connect_SBC_Media_VNF'
        self.vnflcm_custom_disconnect           = self.vnflcm_instances + self.id + '/custom/disconnect_SBC_Media_VNF'
        self.vnflcm_custom_dbrestore            = self.vnflcm_instances + self.id + '/custom/DB_Restore'
        self.vnflcm_custom_upgrade_precheck     = self.vnflcm_instances + self.id + '/custom/upgrade_precheck'
        self.vnflcm_custom_upgrade_1_apply      = self.vnflcm_instances + self.id + '/custom/upgrade_1_apply'
        self.vnflcm_custom_upgrade_2_activate   = self.vnflcm_instances + self.id + '/custom/upgrade_2_activate'
        self.vnflcm_custom_upgrade_3_commit     = self.vnflcm_instances + self.id + '/custom/upgrade_3_commit'
        self.vnflcm_custom_upgrade_archive      = self.vnflcm_instances + self.id + '/custom/upgrade_archive'

    def create_vnf(self):
        super().create_vnf()
        self.set_vnflcm_apis()

    def cssu_instantiate_vnf(self):
        data = json.load(open(self.cssu_instantiate_file, "rb"))
        response = send_request(type='POST', url=self.vnflcm_instantiate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def backup_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_backup, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def connect_sbc_media_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_connect, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def disconnect_sbc_media_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_disconnect, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def dbrestore_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_dbrestore, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_precheck(self):
        data = {}
        # Need to carry empty data in requests.post, otherwise will fail
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_precheck, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_1_apply(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_1_apply, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_2_activate(self):
        data = {}
        # Need to carry empty data in requests.post, otherwise will fail
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_2_activate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_3_commit(self):
        data = {}
        # Need to carry empty data in requests.post, otherwise will fail
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_3_commit, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_archive(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_archive, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

'''
Media VNF
'''
class MediaVnf(SBCVnf):
    def __init__(self, vnfdId, apiVersion='3', name='sbclcm-media-plane',
                 description='SBC LCM Media-Plane VNF for Auto Test via CBAM REST APIs'):
        super().__init__(vnfdId, apiVersion, name, description)
        self.name = name
        self.description = description
        self.vnfdId = vnfdId
        self.apiVersion= apiVersion
        self.id = ''
        self.vnf_type = 'media'
        self.vmName_List = []
        self.vmHeal_List = []
        self.name2slot_Map = {}
        self.extension_file = media_data_dir + '\\' + media_extension_json_file
        self.dr_extension_file = media_data_dir + '\\' + media_dr_extension_json_file
        self.instantiate_file = media_data_dir + '\\' + media_instantiation_json_file
        self.dr_instantiate_file = media_data_dir + '\\' + media_dr_instantiation_json_file

    def set_vnflcm_apis(self):
        super().set_vnflcm_apis()
        if self.apiVersion == '4':
            self.vnflcm_instances       = cbam_url + '/vnflcm/v1' + '/vnf_instances/'
            self.vnflcm_create          = cbam_url + '/vnflcm/v1' + '/vnf_instances'
        if self.apiVersion == '3':
            self.vnflcm_instances       = cbam_url + '/vnfm/lcm/v3' + '/vnfs/'
            self.vnflcm_create          = cbam_url + '/vnfm/lcm/v3' + '/vnfs'
            self.vnflcm_modify          = self.vnflcm_instances + self.id
            self.vnflcm_instantiate     = self.vnflcm_instances + self.id + '/instantiate'
            self.vnflcm_upgrade         = self.vnflcm_instances + self.id + '/upgrade'
        self.operationState             = cbam_url + '/vnflcm/v1' + '/vnf_lcm_op_occs'
        self.vnflcm_custom_backup       = self.vnflcm_instances + self.id + '/custom/backup'
        self.vnflcm_custom_issu         = self.vnflcm_instances + self.id + '/custom/issu'
        self.vnflcm_custom_nssu         = self.vnflcm_instances + self.id + '/custom/nssu'
        self.vnflcm_custom_rollback     = self.vnflcm_instances + self.id + '/custom/rollback'
        self.vnflcm_custom_backout      = self.vnflcm_instances + self.id + '/custom/backout'
        self.vnflcm_custom_restore      = self.vnflcm_instances + self.id + '/custom/restore'
        self.vnflcm_custom_post_dr      = self.vnflcm_instances + self.id + '/custom/Post_Disaster_Recovery'
        self.vnflcm_custom_update_lcm_admin_user    = self.vnflcm_instances + self.id + '/custom/update_lcm_admin_user'
        self.vnflcm_custom_register_lcm_user        = self.vnflcm_instances + self.id + '/custom/register_lcm_user'
        self.vnflcm_custom_unregister_lcm_user      = self.vnflcm_instances + self.id + '/custom/unregister_lcm_user'

    def create_vnf(self):
        super().create_vnf()
        self.set_vnflcm_apis()

    # overide upgrade_vnf in SBCVnf
    def upgrade_vnf(self, vnfdId):
        if self.apiVersion == '3':
            data = {
                "vnfdId": vnfdId
            }
        if self.apiVersion == '4':
            data = {
                "vnfdId": vnfdId,
                "apiVersion": "4.0"
            }
        response = send_request(type='POST', url=self.vnflcm_upgrade, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def get_vmName_List(self):
        data = self.get_vnf()
        for vnfcinfo in data['instantiatedVnfInfo']['vnfcResourceInfo']:
            self.vmName_List.append(vnfcinfo['computeResource']['additionalData']['name'])
        log('vmName_List: '+ str(self.vmName_List))

    def get_name2slot_Map(self):
        """
        {
            'sbgw01vm020': '20',
            'sbgw01vm017': '17',
            'sbgw01vm002': '2',
            'sbgw01vm003': '3',
            'sbgw01vm001': '1',
            'sbgw01vm019': '19',
            'sbgw01vm018': '18',
            'sbgw01vm004': '4',
            'sbgw01vm005': '5',
            'sbgw01vm015': '15',
            'sbgw01vm014': '14',
            'sbgw01vm008': '8',
            'sbgw01vm016': '16',
            'sbgw01vm011': '11',
            'sbgw01vm010': '10',
            'sbgw01vm013': '13',
            'sbgw01vm007': '7'
        }
        """
        data = self.get_vnf()
        self.name2slot_Map = data['extensions']['media_params']['name_map_slot']
        log('name2slot_Map: '+ str(self.name2slot_Map))

    def heal_vnf(self, vmHealList='', cause='', lcm_user='cloud-user'):
        # vmHealList is actually one string with comma seperated vm name
        data = {
            "cause": cause,
            "vnfcInstanceIds": [],
            "additionalParams": {"vmHealList": vmHealList, "lcm_user": lcm_user}
        }
        response = send_request(type='POST', url=self.vnflcm_heal, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def backup_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_backup, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def restore_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_restore, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def post_dr(self):
        data = {}
        # Need to carry empty data in requests.post, otherwise will fail
        response = send_request(type='POST', url=self.vnflcm_custom_post_dr, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def issu_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_issu, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def nssu_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_nssu, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def rollback_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_rollback, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def backout_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_backout, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)


'''
SBC VNF LCM Tests
'''
class SBCVnfLcmTestDriver(object):
    def __init__(self, sbcvnf):
        self.getvnfs = cbam_url + '/vnflcm/v1' + '/vnf_instances/'
        self.sbcvnf = sbcvnf
        self.vnf_id = ''
        self.vnf_name = ''
        # vnf_type is 'sig' or 'media'
        self.vnf_type = self.sbcvnf.vnf_type
        self.wkk = wkk
        # centos is for pubkey login
        self.backup_server_login_pubkey = backup_server_login_pubkey
        self.local_private_key = local_private_key
        # root is for passwd login
        self.backup_server_login_passwd = backup_server_login_passwd
        self.backup_server_passwd_passwd = backup_server_passwd_passwd
        self.ssh_type = ssh_type
        self.backup_server_ip = backup_server_ip
        self.backup_server_dir = backup_server_dir
        self.backup_server1 = backup_server1
        self.backup_server2 = backup_server2
        self.backup_server_creds1 = backup_server_creds
        self.backup_server_creds2 = self.backup_server_creds1

    def restore_vnf(self):
        """
        Need to restore vnf each time after vnf has been created
        1. get vnf id
        2. set id for vnf
        3. restore the vnf's apis
        """
        self.vnf_id = self.get_vnf_id()
        self.sbcvnf.set_id(self.vnf_id)
        self.sbcvnf.set_vnflcm_apis()

    def get_vnf_id(self):
        token_bear = get_token()
        headers = {'Authorization': token_bear}
        response = requests.get(self.getvnfs, headers=headers, verify=False, proxies=proxies)
        if response.status_code == 200:
            data = json.loads(response.text)
            for vnf in data:
                # Fow now, use vnfProductName to determine vnf id
                if vnf['vnfProductName'] == 'SBC' and self.vnf_type == 'sig':
                    log('get_vnf_id: sig vnf id: ' + vnf['id'])
                    return vnf['id']
                if vnf['vnfProductName'] == 'SBC-media' and self.vnf_type == 'media':
                    log('get_vnf_id: media vnf id: ' + vnf['id'])
                    return vnf['id']
        else:
            log('Failed to get_vnf_id')
            exit(1)

    def create_instantiate(self):
        self.sbcvnf.create_vnf()
        if self.vnf_type == 'media':
            self.sbcvnf.modify_extension_vnf()
        self.sbcvnf.instantiate_vnf()
        self.sbcvnf.get_vnf()

    def dr_create_instantiate(self):
        self.sbcvnf.create_vnf()
        if self.vnf_type == 'media':
            self.sbcvnf.dr_modify_extension_vnf()
        self.sbcvnf.dr_instantiate_vnf()
        self.sbcvnf.get_vnf()

    def terminate(self):
        self.sbcvnf.terminate_vnf()

    def delete(self):
        self.sbcvnf.delete_vnf()

    def change_package_version(self, vnfdId):
        self.sbcvnf.upgrade_vnf(vnfdId)

    def get_scale_status(self, aspectId="sc_Aspect"):
        self.sbcvnf.get_scaleStatus(aspectId=aspectId)
        return self.sbcvnf.scaleLevel, self.sbcvnf.maxScaleLevel

    def scale(self, scaleType='SCALE_OUT', step=1, aspectId="sc_Aspect"):
        self.sbcvnf.get_scaleStatus(aspectId=aspectId)
        scaleLevel = self.sbcvnf.scaleLevel
        maxScaleLevel = self.sbcvnf.maxScaleLevel
        if scaleType == 'SCALE_OUT':
            if scaleLevel == maxScaleLevel:
                log('SBC VNF maxScaleLevel has been reached, no action.')
                return
            elif scaleLevel < maxScaleLevel:
                self.sbcvnf.scale_vnf(scaleType='SCALE_OUT', step=step, aspectId=aspectId)
        elif scaleType == 'SCALE_IN':
            if scaleLevel == 1:
                log('SBC min SC number has been reached, no action.')
                return
            elif scaleLevel <= maxScaleLevel:
                self.sbcvnf.scale_vnf(scaleType='SCALE_IN', step=step, aspectId=aspectId)

    def get_vnfcInstanceId_List_4Heal(self):
        self.sbcvnf.get_vnfcInstanceId_List()
        return self.sbcvnf.vnfcInstanceId_List

    def heal(self, vnfcInstanceId=None, cause=''):
        if vnfcInstanceId is None:
            vnfcInstanceId = ['OAM.NOKIA-LCP-VMA']
        self.sbcvnf.heal_vnf(vnfcInstanceId, cause)

    def modify_auto_scale(self, periodInSeconds=600, enabled=False):
        self.sbcvnf.modify_auto_scale_vnf(periodInSeconds, enabled)

    def modify_auto_backup(self, periodInSeconds=86220, enabled=False):
        self.sbcvnf.modify_auto_backup_vnf(periodInSeconds, enabled)

    def check_latestOpState(self, timeout=30):
        status = self.sbcvnf.get_latestOperationState()
        if status is not None:
            return status

    def wait_processing(self, timeout=30):
        log('wait_processing')
        wait = 0
        while wait < timeout:
            status = self.check_latestOpState()
            if status in ['PROCESSING', 'STARTING']:
                time.sleep(60)
                wait = wait + 1
            else:
                break

    def check_opstatus(self, operation='INSTANTIATE', operationName='', timeout=30):
        wait = 0
        while wait < timeout:
            status = self.sbcvnf.get_operationState(operation, operationName)
            if status == 'PROCESSING':
                log('SBC VNF ' + operation + ' ' + operationName + ' PROCESSING')
                log('Wait for 60 secs...')
            elif status == 'FAILED':
                log('SBC VNF ' + operation + ' ' + operationName + ' FAILED')
                # log the detailed error info
                data = self.sbcvnf.get_latestOperation()
                log(str(data))
                exit(1)
            elif status == 'ROLLED_BACK ':
                log('SBC VNF ' + operation + ' ' + operationName + ' ROLLED_BACK')
                exit(2)
            elif status == 'COMPLETED':
                log('SBC VNF ' + operation + ' ' + operationName + ' COMPLETED')
                return 0
            else:
                log('SBC VNF ' + operation + ' ' + operationName + ' status: ' + status)
            # Check status every 60 secs
            time.sleep(60)
            wait = wait + 1

'''
Sig VNF LCM Tests
'''
class SigVnfLcmTestDriver(SBCVnfLcmTestDriver):
    def __init__(self, sbcvnf):
        super().__init__(sbcvnf)
        # self.vnf_type = 'sig'
        self.vnf_name = sig_vnf_name
        self.sig_oama_ip = sig_oama_ip
        self.sig_oam_login = sig_oam_login
        self.sig_oam_passwd = sig_oam_passwd
        self.fixed_scm_ip = sig_fixed_scm_ip
        self.restore_media_plane = sig_restore_media_plane
        self.cssu_zip = sig_cssu_zip
        self.backup_zip = sig_backup_zip
        self.su_deft_url = sig_su_deft_url
        self.su_deft_key = sig_su_deft_key
        self.su_to_image = sig_su_to_image
        self.su_to_version = sig_su_to_version
        # Following are additionalParams for various operations
        self.additinalParams_Backup_Local = {
            'backupServer1': '',
            'backup_server_credentials1': '',
            'backupServer2': '',
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_1 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': '',
            'backupServer2': '',
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_2 = {
            'backupServer1': '',
            'backup_server_credentials1': '',
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_12 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': '',
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_Creds_1 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': self.backup_server_creds1,
            'backupServer2': '',
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_Creds_2 = {
            'backupServer1': '',
            'backup_server_credentials1': '',
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': self.backup_server_creds2
        }
        self.additinalParams_Backup_Remote_Creds_12 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': self.backup_server_creds1,
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': self.backup_server_creds2
        }
        self.additinalParams_Connection = {
            'FixedScmIpAddress': self.fixed_scm_ip,
            'LcmUser': media_lcm_user,
            'ApplUser': media_appl_user,
            'ApplUserPw': media_appl_passwd
        }
        self.additinalParams_Disconnection = {
            'FixedScmIpAddress': self.fixed_scm_ip,
            'LcmUser': 'cloud-user'
        }
        self.additinalParams_DBRestore = {
            'backupInfo': self.backup_server1 + self.backup_zip,
            'restore_media_plane': self.restore_media_plane
        }
        self.additinalParams_Upgrade1Apply = {
            'upgradeImageName': self.su_to_image,
            'upgradeVersion': self.su_to_version,
            'upgradeDeftKey': self.su_deft_key,
            'upgradeDeftUrl': self.su_deft_url
        }
        self.additinalParams_UpgradeArchive = {
            'upgradeDeftKey': self.su_deft_key,
            'upgradeDeftUrl': self.su_deft_url,
            'upgradeVersion': self.su_to_version,
            'upgradeServer': self.backup_server1,
            'upgradeServerCredentials': ''
        }
        self.additinalParams_UpgradeArchive_Creds = {
            'upgradeDeftKey': self.su_deft_key,
            'upgradeDeftUrl': self.su_deft_url,
            'upgradeVersion': self.su_to_version,
            'upgradeServer': self.backup_server1,
            'upgradeServerCredentials': self.backup_server_creds1
        }

    def custom_backup(self, additionalParam):
        self.sbcvnf.backup_vnf(additionalParam)

    def custom_connect_SBC_Media_VNF(self, additionalParam):
        self.sbcvnf.connect_sbc_media_vnf(additionalParam)

    def custom_disconnect_SBC_Media_VNF(self, additionalParam):
        self.sbcvnf.disconnect_sbc_media_vnf(additionalParam)

    def custom_dbrestore(self, additionalParam):
        self.sbcvnf.dbrestore_vnf(additionalParam)

    def custom_upgrade_precheck(self):
        self.sbcvnf.upgrade_precheck()

    def custom_upgrade_1_apply(self, additionalParam):
        self.sbcvnf.upgrade_1_apply(additionalParam)

    def custom_upgrade_2_activate(self):
        self.sbcvnf.upgrade_2_activate()

    def custom_upgrade_3_commit(self):
        self.sbcvnf.upgrade_3_commit()

    def custom_upgrade_archive(self, additionalParam):
        self.sbcvnf.upgrade_archive(additionalParam)

    def cssu_create_instantiate(self):
        self.sbcvnf.create_vnf()
        self.sbcvnf.cssu_instantiate_vnf()
        self.sbcvnf.get_vnf()

    # def dr_create_instantiate(self):
    #     self.sbcvnf.create_vnf()
    #     self.sbcvnf.dr_instantiate_vnf()
    #     self.sbcvnf.get_vnf()

    def prep_bkserver_pubkey(self):
        """
        # Need to setup pubkey authentication after instantiation
        """
        log('In Func:' + sys._getframe().f_code.co_name)
        pubkey = ''
        # Delete the pubkey in backup server authorized_keys
        # Example: sed - i '/sbclcm03/d' /home/centos/.ssh/authorized_keys
        sshtype = 'passwd'
        cmd = "sed -i " + "'/" + self.vnf_name + "/d' " + "/home/centos/.ssh/authorized_keys"
        ssh_command(cmd, self.backup_server_ip, self.backup_server_login_passwd,
                             self.backup_server_passwd_passwd, sshtype)
        # Get oam ssh pubkey of root
        cmd = 'cat /root/.ssh/id_rsa.pub'
        pubkey = ssh_command(cmd, self.sig_oama_ip, self.sig_oam_login, self.sig_oam_passwd, sshtype)
        os.chdir(working_dir)
        with open('tmp_key', 'w') as f:
            f.write(pubkey)
        # Ship the oam ssh pubkey to /tmp on backup server
        ssh_scp_put(self.backup_server_ip, self.backup_server_login_passwd,
                    self.backup_server_passwd_passwd, 'tmp_key', '/tmp/tmp_key')
        # echo the pubkey to authorized_keys
        cmd = 'cat /tmp/tmp_key >> /home/centos/.ssh/authorized_keys'
        result = ssh_command(cmd, self.backup_server_ip, self.backup_server_login_passwd,
                             self.backup_server_passwd_passwd, sshtype)

    def rm_existing_raw_zip(self, sshtype='passwd'):
        """
        This is to remove existing raw backup/archive zip before backup/ua
        :param sshtype: passwd or pubkey
        """
        log('In Func:' + sys._getframe().f_code.co_name)
        match_str = ['_LCP', 'su_archive_R']
        ip = self.backup_server_ip
        if sshtype == 'passwd':
            login = self.backup_server_login_passwd
            passwd_key = self.backup_server_passwd_passwd
        elif sshtype == 'pubkey':
            login = self.backup_server_login_pubkey
            passwd_key = self.local_private_key
        else:
            log('Only type passwd and pubkey supprted for sshtype. Error out.')
            exit(1)
        cmd = 'ls ' + self.backup_server_dir
        result = ssh_command(cmd, ip, login, passwd_key, sshtype)
        for mstr in match_str:
            if mstr in result:
                rlist = result.split()
                for r in rlist:
                    if mstr in r:
                        cmd = 'rm -f ' + self.backup_server_dir + r
                        log(cmd)
                        ssh_command(cmd, ip, login, passwd_key, sshtype)

    def prep_backup_cssu_zip(self, sshtype='passwd', ziptype='backup'):
        """
        This is combined version to rename and chmod zip afer backup or cssu
        :param sshtype: passwd or pubkey
        :param ziptype: backup or cssu
        """
        log('In Func:' + sys._getframe().f_code.co_name)
        match_str = ''
        newzip = ''
        ip = self.backup_server_ip
        if ziptype == 'backup':
            match_str = '_LCP'
            newzip = self.backup_zip
        elif ziptype == 'cssu':
            match_str = 'su_archive_R'
            newzip = self.cssu_zip
        else:
            log('Only type backup and cssu supprted for ziptype. Error out.')
            exit(1)
        if sshtype == 'passwd':
            login = self.backup_server_login_passwd
            passwd_key = self.backup_server_passwd_passwd
        elif sshtype == 'pubkey':
            login = self.backup_server_login_pubkey
            passwd_key = self.local_private_key
        else:
            log('Only type passwd and pubkey supprted for sshtype. Error out.')
            exit(1)
        cmd = 'ls ' + self.backup_server_dir
        result = ssh_command(cmd, ip, login, passwd_key, sshtype)
        if match_str in result:
            rlist = result.split()
            for r in rlist:
                if match_str in r:
                    cmd = 'mv ' + self.backup_server_dir + r + ' ' + self.backup_server_dir + newzip
                    log(cmd)
                    ssh_command(cmd, ip, login, passwd_key, sshtype)
                    cmd = 'chmod 777 ' + self.backup_server_dir + newzip
                    log(cmd)
                    ssh_command(cmd, ip, login, passwd_key, sshtype)

'''
Media VNF LCM Tests
'''
class MediaVnfLcmTestDriver(SBCVnfLcmTestDriver):
    def __init__(self, sbcvnf):
        super().__init__(sbcvnf)
        self.sbcvnf = sbcvnf
        self.vnf_id = ''
        self.vnf_name = media_vnf_name
        self.scma_ip = media_scma_ip
        self.scmb_ip = media_scmb_ip
        self.scm_vip = media_scm_vip
        self.backup_zip = media_backup_zip
        self.lcm_user = media_lcm_user
        self.appl_user = media_appl_user
        self.appl_passwd = media_appl_passwd
        self.vm_slot_dict = {'ACTIVE-MCM': [], 'STANDBY-MCM': [],
                  'ACTIVE-PIM': [], 'STANDBY-PIM': [],
                  'ACTIVE-SCM': [], 'STANDBY-SCM': []}
        self.stbyPIM_list = []
        self.stbyMCM_list = []
        self.stbySCM_list = []
        self.stbyALL_list = []
        self.su_to_image = media_su_to_image
        self.backup_url = media_backup_url
        # Following are additionalParams for various operations
        self.additinalParams_Backup_Local = {
            'backup_server1': '',
            'backup_server_credentials1': '',
            'backup_server2': '',
            'backup_server_credentials2': '',
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Backup_Remote_1 = {
            'backup_server1': self.backup_server1,
            'backup_server_credentials1': '',
            'backup_server2': '',
            'backup_server_credentials2': '',
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Backup_Remote_2 = {
            'backup_server1': '',
            'backup_server_credentials1': '',
            'backup_server2': self.backup_server2,
            'backup_server_credentials2': '',
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Backup_Remote_12 = {
            'backup_server1': self.backup_server1,
            'backup_server_credentials1': '',
            'backup_server2': self.backup_server2,
            'backup_server_credentials2': '',
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Backup_Remote_Creds_1 = {
            'backup_server1': self.backup_server1,
            'backup_server_credentials1': self.backup_server_creds1,
            'backup_server2': '',
            'backup_server_credentials2': '',
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Backup_Remote_Creds_2 = {
            'backup_server1': '',
            'backup_server_credentials1': '',
            'backup_server2': self.backup_server2,
            'backup_server_credentials2': self.backup_server_creds2,
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Backup_Remote_Creds_12 = {
            'backup_server1': self.backup_server1,
            'backup_server_credentials1': self.backup_server_creds1,
            'backup_server2': self.backup_server2,
            'backup_server_credentials2': self.backup_server_creds2,
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Restore = {
            'BackupUrl': self.backup_url,
            'lcm_user': self.lcm_user
        }
        self.additinalParams_ISSU = {
            'image': self.su_to_image,
            'lcm_user': self.lcm_user
        }
        self.additinalParams_NSSU = {
            'image': self.su_to_image,
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Rollback = {
            'BackupUrl': self.backup_url,
            'lcm_user': self.lcm_user
        }
        self.additinalParams_Backout = {
            'BackupUrl': self.backup_url,
            'lcm_user': self.lcm_user
        }

# ssh diag@100.69.127.152 -assured
# #sbgw01:ACT-SCM:1.11(r0)>=1:diag:main:vMGx# vi node
# =====================================================================
#   Appl-Id  |  Function       |  LBI   |            Appl-state
#            |                 |        |  Oper         |  Admin
# =====================================================================
#   1           ACTIVE-MCM        mcm3     UP              ENABLED
#   2           ACTIVE-MCM        mcm4     UP              ENABLED
#   3          STANDBY-MCM                 UP              ENABLED
#   4           ACTIVE-MCM        mcm2     UP              ENABLED
#   5           ACTIVE-MCM        mcm1     UP              ENABLED
#   7          STANDBY-PIM                 UP              ENABLED
#   8           ACTIVE-PIM        pim5     UP              ENABLED
#   10         STANDBY-SCM                 UP              ENABLED
#   11          ACTIVE-SCM      * scm      UP              ENABLED
#   13         STANDBY-PIM                 UP              ENABLED
#   14          ACTIVE-PIM        pim4     UP              ENABLED
#   15          ACTIVE-PIM        pim3     UP              ENABLED
#   16         STANDBY-PIM                 UP              ENABLED
#   17         STANDBY-PIM                 UP              ENABLED
#   18          ACTIVE-PIM        pim2     UP              ENABLED
#   19          ACTIVE-PIM        pim1     UP              ENABLED
#   20         STANDBY-PIM                 UP              ENABLED
# =====================================================================
#     {
#         'sbgw01vm020': '20',
#         'sbgw01vm017': '17',
#         'sbgw01vm002': '2',
#         'sbgw01vm003': '3',
#         'sbgw01vm001': '1',
#         'sbgw01vm019': '19',
#         'sbgw01vm018': '18',
#         'sbgw01vm004': '4',
#         'sbgw01vm005': '5',
#         'sbgw01vm015': '15',
#         'sbgw01vm014': '14',
#         'sbgw01vm008': '8',
#         'sbgw01vm016': '16',
#         'sbgw01vm011': '11',
#         'sbgw01vm010': '10',
#         'sbgw01vm013': '13',
#         'sbgw01vm007': '7'
#     }

    def heal(self, vmHealList='', cause=''):
        self.sbcvnf.heal_vnf(vmHealList, cause, self.lcm_user)

    def custom_backup(self, additionalParam):
        self.sbcvnf.backup_vnf(additionalParam)

    def custom_restore(self, additionalParam):
        self.sbcvnf.restore_vnf(additionalParam)

    def custom_post_dr(self):
        self.sbcvnf.post_dr()

    def custom_issu(self):
        self.sbcvnf.issu_vnf(self.additinalParams_ISSU)

    def custom_nssu(self):
        self.sbcvnf.nssu_vnf(self.additinalParams_NSSU)

    def custom_rollback(self):
        self.sbcvnf.rollback_vnf(self.additinalParams_Rollback)

    def custom_backout(self):
        self.sbcvnf.backout_vnf(self.additinalParams_Backout)

    def gen_stby_vm_list(self):
        log('In Func:' + sys._getframe().f_code.co_name)
        self.prep_local_known_hosts()
        cmd = 'vi node'
        data = ssh_paramiko_pexpect(cmd, self.scm_vip, self.appl_user, self.appl_passwd)
        log('Output of vi node: ' + str(data))
        data = data.split('\n')
        vmlist = []
        vmdict = {'ACTIVE-MCM': [], 'STANDBY-MCM': [],
                  'ACTIVE-PIM': [], 'STANDBY-PIM': [],
                  'ACTIVE-SCM': [], 'STANDBY-SCM': []}
        for i in data:
            if 'ENABLED' in i:
                vmlist.append(i)
        for i in vmdict.keys():
            for j in vmlist:
                if i in j:
                    vmdict[i].append(j.split(' ')[2])
        # >>> vmdict
        # {'ACTIVE-MCM': ['1', '2', '4', '5'], 'STANDBY-MCM': ['3'], 'ACTIVE-PIM': ['8', '14', '15', '18', '19'],
        #  'STANDBY-PIM': ['7', '13', '16', '17', '20'], 'ACTIVE-SCM': ['11'], 'STANDBY-SCM': ['10']}
        self.vm_slot_dict = vmdict
        self.sbcvnf.get_name2slot_Map()
        log('Media vnf name2slot_Map: ' + str(self.sbcvnf.name2slot_Map))
        # gen standy pim, mcm, scm vm list
        self.stbyPIM_list = []
        self.stbyMCM_list = []
        self.stbySCM_list = []
        for slot in vmdict['STANDBY-PIM']:
            for key, val in self.sbcvnf.name2slot_Map.items():
                if slot == val:
                    self.stbyPIM_list.append(key)
        for slot in vmdict['STANDBY-MCM']:
            for key, val in self.sbcvnf.name2slot_Map.items():
                if slot == val:
                    self.stbyMCM_list.append(key)
        for slot in vmdict['STANDBY-SCM']:
            for key, val in self.sbcvnf.name2slot_Map.items():
                if slot == val:
                    self.stbySCM_list.append(key)
        self.stbyALL_list = self.stbyPIM_list + self.stbyMCM_list + self.stbySCM_list
        log('vm_slot_dict: ' + str(self.vm_slot_dict))
        log('stbyPIM_list: ' + str(self.stbyPIM_list))
        log('stbyMCM_list: ' + str(self.stbyMCM_list))
        log('stbySCM_list: ' + str(self.stbySCM_list))
        log('stbyALL_list: ' + str(self.stbyALL_list))

    def prep_bkserver_pubkey(self):
        log('In Func:' + sys._getframe().f_code.co_name)
        pubkey = ''
        # Delete the pubkey in backup server authorized_keys
        # Example: sed - i '/A7510/d' /home/centos/.ssh/authorized_keys
        sshtype = 'passwd'
        cmd = "sed -i " + "'/A7510/d' " + "/home/centos/.ssh/authorized_keys"
        ssh_command(cmd, self.backup_server_ip, self.backup_server_login_passwd,
                             self.backup_server_passwd_passwd, sshtype)
        # Get scm ssh pubkey of active scm
        sshtype = 'pubkey'
        cmd = 'cat /opt/v7510/data/HOST_RSA.PUB'
        pubkey = ssh_command(cmd, self.scm_vip, self.lcm_user, self.wkk, sshtype)
        os.chdir(working_dir)
        with open('tmp_key', 'w') as f:
            f.write(pubkey)
        # Ship the scm ssh pubkey to /tmp on backup server
        ssh_scp_put(self.backup_server_ip, self.backup_server_login_passwd,
                    self.backup_server_passwd_passwd, 'tmp_key', '/tmp/tmp_key')
        # echo the pubkey to authorized_keys
        sshtype = 'passwd'
        cmd = 'cat /tmp/tmp_key >> /home/centos/.ssh/authorized_keys'
        result = ssh_command(cmd, self.backup_server_ip, self.backup_server_login_passwd,
                             self.backup_server_passwd_passwd, sshtype)

    def rm_backup_zip(self):
        log('In Func:' + sys._getframe().f_code.co_name)
        """
        Remove the BACKUP_C.ZIP on backup server
        """
        log('Rmove the existing BACKUP_C.ZIP backup file on backup server.')
        bkup_file = self.backup_server_dir + self.backup_zip
        cmd = 'rm -f ' + bkup_file
        ssh_command(cmd, self.backup_server_ip, self.backup_server_login_passwd,
                    self.backup_server_passwd_passwd, self.ssh_type)

    def prep_local_known_hosts(self):
        """
        # Need to rm the entry in known_hosts file on local pc otherwise there would be error
        """
        log('In Func:' + sys._getframe().f_code.co_name)
        log('Rmove the existing entry of SCM VIP in /c/Users/shawnx/.ssh/known_hosts.')
        known_hosts_file = r'C:\Users\shawnx\.ssh\known_hosts'
        tmp_list = []
        try:
            with open(known_hosts_file, 'r') as f:
                tmp_list = f.readlines()
            for val in tmp_list:
                if self.scm_vip in val:
                    tmp_list.remove(val)
            with open(known_hosts_file, 'w') as f:
                f.writelines(tmp_list)
        except Exception:
            traceback.print_exc()

'''
Sig VNF Artifact Generator
'''
class SigVnfArtsGenerator(object):
    def __init__(self, type='instantiation', rel='', server_type=None):
        '''server_type is required for DR, CSSU'''
        self.type = type
        self.rel = rel
        self.server_type = server_type
        self.arts_type = sig_arts_type
        self.yact_server_ip = sig_yact_server_ip
        self.yact_user = sig_yact_user
        self.yact_passwd = sig_yact_passwd
        self.yact_dir = sig_yact_dir
        self.sig_dif_name = sig_dif_name
        self.sig_dr_dif_name = sig_dr_dif_name
        self.sig_su_dif_name = sig_su_dif_name
        self.sig_cssu_dif_name = sig_cssu_dif_name
        self.sig_dif_dict = {
            'base': self.sig_dif_name,
            'cssu': self.sig_cssu_dif_name,
            'dr': self.sig_dr_dif_name,
            'su': self.sig_su_dif_name
        }
        self.sig_ne_name = sig_vnf_name
        self.sig_zip = self.sig_ne_name + '.zip'
        self.sig_dif_file = sig_arts_dir + '\\' + self.sig_dif_name
        self.sig_vnfpkg_name = sig_vnfpkg_name
        self.sig_dr_vnfpkg_name = sig_dr_vnfpkg_name
        self.sig_su_vnfpkg_name = sig_su_vnfpkg_name
        self.sig_cssu_vnfpkg_name = sig_cssu_vnfpkg_name
        self.instantiation_json_file = sig_instantiation_json_file
        self.dr_instantiation_json_file = sig_dr_instantiation_json_file
        self.cssu_instantiation_json_file = sig_cssu_instantiation_json_file
        self.os_passwd = os_passwd
        self.backup_server_ip = backup_server_ip
        self.backup_server_login = backup_server_login
        self.backup_server_passwd = backup_server_passwd
        self.backup_server_dir  = backup_server_dir
        self.bulk_conf_url      = sig_bulk_conf_url
        self.backup_file1_http  = sig_backup_file1_http
        self.backup_file2_http  = sig_backup_file2_http
        self.backup_file1_creds = sig_backup_file1_creds
        self.backup_file2_creds = sig_backup_file2_creds
        self.upgrade_file_http  = sig_upgrade_file_http
        self.upgrade_file_creds = sig_upgrade_file_creds
        self.backup_file_name = sig_backup_file_name
        self.ssh_type = ssh_type
        self.backup_file = self.backup_server_dir + self.backup_file_name
        self.local_backup_dir = sig_local_backup_dir
        self.ap_instantiation = {
            "backup_file1": "",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": self.bulk_conf_url,
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation = {
            "backup_file1": self.backup_file1_http,
            "backup_file2": self.backup_file2_http,
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_http1 = {
            "backup_file1": self.backup_file1_http,
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_http2 = {
            "backup_file1": "",
            "backup_file2": self.backup_file2_http,
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_http12 = {
            "backup_file1": self.backup_file1_http,
            "backup_file2": self.backup_file2_http,
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_creds1 = {
            "backup_file1": self.backup_file1_creds,
            "backup_file2": "",
            "backup_server_credentials1": backup_server_creds,
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_creds2 = {
            "backup_file1": "",
            "backup_file2": self.backup_file2_creds,
            "backup_server_credentials1": "",
            "backup_server_credentials2": backup_server_creds,
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_creds12 = {
            "backup_file1": self.backup_file1_creds,
            "backup_file2": self.backup_file2_creds,
            "backup_server_credentials1": backup_server_creds,
            "backup_server_credentials2": backup_server_creds,
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_cssu_instantiation_http1 = {
            "backup_file1": "",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": self.upgrade_file_http,
            "upgrade_server_credentials": ""
        }
        self.ap_cssu_instantiation_creds1 = {
            "backup_file1": "",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": self.upgrade_file_creds,
            "upgrade_server_credentials": backup_server_creds
        }

    def gen_arts_sig(self):
        log('In Func:' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)

        if self.type not in self.arts_type:
            log('Only ' + str(self.arts_type) + ' Supported.')
            exit(1)

        ts = int(time.time())
        log('Current time is: ' + str(ts))

        sig_yact_dir = self.yact_dir + str(ts) + '/'
        cmd = 'mkdir -p ' + sig_yact_dir
        ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)

        if self.type == 'instantiation':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_dif_name
        elif self.type == 'dr':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_dr_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_dr_dif_name
        elif self.type == 'su':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_su_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_su_dif_name
        elif self.type == 'cssu':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_cssu_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_cssu_dif_name

        log('sig_dif_file: ' + sig_dif_file)
        log('sig_rmt_dif_file: ' + sig_rmt_dif_file)
        ssh_scp_put(self.yact_server_ip, self.yact_user, self.yact_passwd, sig_dif_file, sig_rmt_dif_file)

        yact_cmd = '/home/yact/YACT/yact.sh gen-by-dif ' + sig_rmt_dif_file + ' SBC-signaling ' + self.rel
        log('yact_cmd: ' + yact_cmd)
        cmd = 'cd ' + sig_yact_dir + ';' + yact_cmd
        result = ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)
        log('result of yact_cmd: '+ result)

        # Handle SU case
        if self.type == 'su':
            pkg_tools_dir = sig_yact_dir + self.sig_ne_name + '/package/pkg_tools/'
            log('su pkg_tools_dir: ' + pkg_tools_dir)

            # ship from load vnf pkg to pkg_tools_dir and rename it to vnfpkg-from-load.zip
            vnfpkg_from_load = sig_data_dir + '\\' + self.sig_vnfpkg_name
            rmt_vnfpkg_from_load = pkg_tools_dir + 'vnfpkg-from-load.zip'
            ssh_scp_put(self.yact_server_ip, self.yact_user, self.yact_passwd, vnfpkg_from_load, rmt_vnfpkg_from_load)

            # run vnf_upgrade_pkg_gen
            cmd = 'cd ' + pkg_tools_dir + ';' + './vnf_upgrade_pkg_gen -n ../Nokia_sig_SBC-VNF_Package.zip ' \
                                                '-c ./vnfpkg-from-load.zip ' + self.sig_ne_name
            log('vnf_upgrade_pkg_gen cmd: ' + cmd)
            ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)

            # now ship the upgrade vnf pkg to local
            # example: sbclcm03_Nokia_sig_SBC_upgrade-VNF_Package.zip
            vnfpkg_upgrade = pkg_tools_dir + self.sig_ne_name + '_Nokia_sig_SBC_upgrade-VNF_Package.zip'
            log('upgrade package: ' + vnfpkg_upgrade)
            vnfpkg_upgrade_local = sig_data_dir + '\\' + self.sig_su_vnfpkg_name
            ssh_scp_get(self.yact_server_ip, self.yact_user, self.yact_passwd, vnfpkg_upgrade, vnfpkg_upgrade_local)
        # handle other cases
        else:
            # zip remote sig_art_files
            cmd = 'cd ' + sig_yact_dir + ';' + 'zip -r ' + self.sig_ne_name + '.zip ' + self.sig_ne_name
            result = ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)
            log('result of zip sig: ' + result)

            # remove sig zip if exists
            if os.path.exists(self.sig_zip):
                os.remove(self.sig_zip)

            # remote the unziped dir if exists
            if os.path.exists(self.sig_ne_name):
                shutil.rmtree(self.sig_ne_name)

            rmt_zip_file = sig_yact_dir + self.sig_zip
            ssh_scp_get(self.yact_server_ip, self.yact_user, self.yact_passwd, rmt_zip_file, self.sig_zip)

            log('unzip '+ self.sig_zip)
            zip_file = zipfile.ZipFile(self.sig_zip)
            # os.mkdir(sig_ne_name)
            for names in zip_file.namelist():
                zip_file.extract(names, '.')
            zip_file.close()

    def ship_bulkconf_zip(self):
        # scp the bulkconf_artifacts.zip to remote backup server
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            log('artifacts dir not exist. Exit.')
            exit(1)

        # ship the bulkconf_artifacts.zip
        local_bulkconf_file = sig_arts_dir + '\\' + self.sig_ne_name + r'\bulk_netconf\scale\bulkconf_artifacts.zip'
        rmt_bulkconf_file = self.backup_server_dir + 'bulkconf_artifacts.zip'
        log('local_bulkconf_file: ' + local_bulkconf_file)
        log('rmt_bulkconf_file: ' + rmt_bulkconf_file)
        log('ship bulkconf_artifacts.zip to remote backup server.')
        ssh_scp_put(self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                    local_bulkconf_file, rmt_bulkconf_file)

        # chmod
        cmd = 'cd ' + self.backup_server_dir + '; chmod 777 bulkconf_artifacts.zip'
        result = ssh_command(cmd, self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                             self.ssh_type)
        log('result of chmod bulkconf_artifacts.zip: ' + result)

    def cp_vnf_pkg(self):
        # cp the Nokia_sig_SBC-VNF_Package.zip to data\sig-plane
        # to correct name
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            log('artifacts dir not exist. Exit.')
            exit(1)

        srcfile = self.sig_ne_name + r'\package\Nokia_sig_SBC-VNF_Package.zip'

        if self.type == 'instantiation':
            dstfile = sig_data_dir + '\\' + self.sig_vnfpkg_name
        elif self.type == 'dr':
            dstfile = sig_data_dir + '\\' + self.sig_dr_vnfpkg_name
        elif self.type == 'su':
            dstfile = sig_data_dir + '\\' + self.sig_su_vnfpkg_name
        elif self.type == 'cssu':
            dstfile = sig_data_dir + '\\' + self.sig_cssu_vnfpkg_name

        log('Copy ' + srcfile + ' to ' + dstfile)
        shutil.copyfile(srcfile, dstfile)

    def prep_instantiation_json(self):
        """
        # json load, dump
        # with open(file_name, 'r') as f:
        #     data = json.load(f)
        # with open('output.json', 'w') as f:
        #     json.dump(data, f)
        """
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            log('artifacts dir not exist. Exit.')
            exit(1)

        json_file = self.sig_ne_name + r'\package\cbam_json' + '\\' + self.instantiation_json_file

        # dr and cssu need to populate backup server by server_type
        if self.type in ['dr', 'cssu']:
            if self.server_type not in sig_server_type_list:
                log('server_type not populated for ' + self.type)
                exit(1)

        if self.type == 'dr':
            if self.server_type == 'httpserver1':
                self.ap_dr_instantiation = self.ap_dr_instantiation_http1
            if self.server_type == 'httpserver2':
                self.ap_dr_instantiation = self.ap_dr_instantiation_http2
            if self.server_type == 'httpserver12':
                self.ap_dr_instantiation = self.ap_dr_instantiation_http12
            if self.server_type == 'bkupserver1':
                self.ap_dr_instantiation = self.ap_dr_instantiation_creds1
            if self.server_type == 'bkupserver2':
                self.ap_dr_instantiation = self.ap_dr_instantiation_creds2
            if self.server_type == 'bkupserver12':
                self.ap_dr_instantiation = self.ap_dr_instantiation_creds12
            log('server_type : ' + self.server_type)
            log('ap_dr_instantiation : ' + str(self.ap_dr_instantiation))

        if self.type == 'cssu':
            if self.server_type == 'httpserver1':
                self.ap_cssu_instantiation = self.ap_cssu_instantiation_http1
            if self.server_type == 'bkupserver1':
                self.ap_cssu_instantiation = self.ap_cssu_instantiation_creds1
            log('server_type : ' + self.server_type)
            log('ap_cssu_instantiation : ' + str(self.ap_cssu_instantiation))

        # In case of dr, the instantiation json needs to be extracted from backup.zip
        # In case of su, instantiation json is not needed
        if self.type == 'dr':
            if os.path.exists(self.backup_file_name):
                os.remove(self.backup_file_name)

            if os.path.exists(self.local_backup_dir):
                shutil.rmtree(self.local_backup_dir)

            ssh_scp_get(self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                        self.backup_file, self.backup_file_name)

            log('unzip ' + self.backup_file_name)
            os.mkdir(self.local_backup_dir)
            zip_file = zipfile.ZipFile(self.backup_file_name)
            for names in zip_file.namelist():
                zip_file.extract(names, self.local_backup_dir)
            zip_file.close()

            json_file = self.local_backup_dir + '\\' + self.instantiation_json_file

        # load json data
        try:
            with open(json_file, 'r') as jFile:
                data = json.load(jFile)
        except Exception as exc:
            print('ERROR: Failed to load data from {0}.\n[{1}]'.format(
                json_file, str(exc)))

        # change data
        data['vimConnectionInfo'][0]['accessInfo']['password'] = self.os_passwd

        if self.type == 'instantiation':
            data['additionalParams'] = self.ap_instantiation
            output_json = sig_data_dir + '\\' + self.instantiation_json_file
        elif self.type == 'dr':
            data['additionalParams'] = self.ap_dr_instantiation
            output_json = sig_data_dir + '\\' + self.dr_instantiation_json_file
        elif self.type == 'cssu':
            data['additionalParams'] = self.ap_cssu_instantiation
            output_json = sig_data_dir + '\\' + self.cssu_instantiation_json_file

        # dump modified data
        try:
            with open(output_json, 'w') as jFile:
                json.dump(data, jFile, indent=2)
        except Exception as exc:
            print('ERROR: Failed to write data to {0}.\n[{1}]'.format(
                output_json, str(exc)))

        log('instantiation json for ' + self.type + ' created completed.')

    def get_sc_count(self, workbook):
        """
        get the sc count of provided workbook
        :param workbook: workbook to to used
        :return: the row of sc vm_count, the vm_count value
        """
        log('In Func: ' + sys._getframe().f_code.co_name)
        wb = load_workbook(workbook, keep_vba=True)
        sh = wb['DP']
        colB = sh['B']

        row_list = []
        for i in colB:
            if i.value == 'group_tag':
                row_list.append(i.row)

        row_sc = None
        for r in row_list:
            if sh.cell(row=r, column=26).value == 'SC':
                row_sc = r
        if row_sc is not None:
            log('row_sc: ' + str(row_sc))

        row_vm_count = None
        sc_count = None
        for r in range(row_sc, row_sc + 8):
            if sh.cell(row=r, column=2).value == 'vm_count':
                row_vm_count = r
        if row_vm_count is not None:
            log('row_vm_count: ' + str(row_vm_count))

        sc_count = sh.cell(row=row_vm_count, column=26).value
        log('current SC vaule of vm_count: ' + str(sc_count))

        return (row_vm_count, sc_count)

    def set_sc_count(self, workbook, sc_count=5):
        """
        Set the sc count of provided workbook
        :param workbook: workbook to be modified
        :param sc_count: SC VM pair count, default to 5
        """
        log('In Func: ' + sys._getframe().f_code.co_name)
        row_vm_count, initial_sc_count = self.get_sc_count(workbook)

        wb = load_workbook(workbook, keep_vba=True)
        sh = wb['DP']

        # set vm_count value to updated value
        sh.cell(row=row_vm_count, column=26).value = sc_count

        # wb.save(filename='output.xlsm')
        wb.save(filename=workbook)

        log('Successfully set sc count to: ' + str(sc_count))

    def get_initial_sc_value(self):
        """
        get initial sc value
        :return: initial_sc_count
        """
        log('In Func: ' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_dif_name):
            log('Original DIF does not exist. Exit.')
            exit(1)

        workbook = self.sig_dif_name

        row_vm_count, initial_sc_count = self.get_sc_count(workbook)

        log('initial_sc_count: ' + str(initial_sc_count))

        return int(initial_sc_count)

    def gen_dif_updated_sc_value(self, sc_count=5):
        """
        update DIF using updated sc value
        """
        log('In Func: ' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)

        if self.type == 'su':
            workbook = self.sig_su_dif_name
        elif self.type == 'cssu':
            workbook = self.sig_cssu_dif_name
        elif self.type == 'dr':
            workbook = self.sig_dr_dif_name

        if not os.path.exists(workbook):
            log('workbook does not exist. Exit.')
            exit(1)

        self.set_sc_count(workbook, sc_count)

    def prep_arts(self):
        """
        prep arts for instantiation, dr, cssu, su
        su doesn't need json file
        """
        log('In Func: ' + sys._getframe().f_code.co_name)
        # su doesn't need json file, only vnf pkg needed
        if self.type == 'su':
            self.gen_arts_sig()
        else:
            self.gen_arts_sig()
            self.prep_instantiation_json()
            self.cp_vnf_pkg()
        # needs to ship bulkconf artifacts if instantiation
        if self.type == 'instantiation':
            self.ship_bulkconf_zip()

def sigvnf_GenArtsInstantiation():
    sigArts = SigVnfArtsGenerator(type='instantiation', rel=sig_default_rel)
    sigArts.prep_arts()

def sigvnf_GenArtsDR():
    sigArts = SigVnfArtsGenerator(type='dr', rel=sig_default_rel, server_type='httpserver12')
    sigArts.prep_arts()

def sigvnf_GenArtsSU():
    sigArts = SigVnfArtsGenerator(type='su', rel=sig_toload_rel)
    sigArts.prep_arts()

def sigvnf_GenArtsCSSU():
    sigArts = SigVnfArtsGenerator(type='cssu', rel=sig_toload_rel)
    sigArts.prep_arts()

def sigvnf_GenArts():
    sigvnf_GenArtsInstantiation()
    sigvnf_GenArtsDR()
    sigvnf_GenArtsSU()
    sigvnf_GenArtsCSSU()

'''
Media VNF Artifact Generator
'''
# def run_mmyact():
#     try:
#         client = paramiko.SSHClient()
#         client.load_system_host_keys()
#         client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
#         client.connect(hostname='135.251.49.19', username='test', password='initial')
#         with SSHClientInteraction(client, timeout=10, buffer_size=10240, display=True) as interact:
#             ## login yact server
#             PROMPT = '.*[test@ngnsvr10 ~].*'
#             interact.expect(PROMPT)
#             ## cd to yact dir
#             cmd = 'cd ' + '/home/test/shawnx/20.2/'
#             interact.send(cmd)
#             PROMPT = '.*test@ngnsvr10.*'
#             interact.expect(PROMPT)
#             ## run mmyact
#             cmd = '/home/test/bin/mmyact -u user-input-openstack-R20.2-sbgw01.xlsx'
#             interact.send(cmd)
#             PROMPT = '.*#?.*'
#             interact.expect(PROMPT)
#             ## send 1
#             cmd = '1'
#             interact.send(cmd)
#             PROMPT = '.*choice a new path.*'
#             interact.expect(PROMPT)
#             ## choice a new path
#             cmd = 'y'
#             interact.send(cmd)
#             PROMPT = '.*#?.*'
#             interact.expect(PROMPT)
#             ## choose 1 Public Official Versions
#             cmd = '1'
#             interact.send(cmd)
#             PROMPT = '.*#?.*'
#             interact.expect(PROMPT)
#             ## parser out version
#             cmd_output = interact.current_output_clean
#             tmp_list = cmd_output.split()
#             print(tmp_list)
#             data = None
#             version = 'ap100012'
#             for ldata in tmp_list:
#                 if version in ldata:
#                     data = ldata
#             if data == None:
#                 print('Not find ' + version)
#                 return
#             idx = tmp_list.index(data)
#             cmd = tmp_list[idx - 1].split(')')[0]
#             interact.send(cmd)
#             PROMPT = '.*choice a new dif template.*'
#             interact.expect(PROMPT)
#             ##
#             cmd = 'y'
#             interact.send(cmd)
#             PROMPT = '.*#?.*'
#             interact.expect(PROMPT, timeout=120)
#             ##
#             cmd = '2'
#             interact.send(cmd)
#             PROMPT = '.*#?.*'
#             interact.expect(PROMPT)
#             ##
#             # cmd = self.openrc_file
#             cmd = '/home/test/shawnx/20.2/cb0078sa.v2'
#             interact.send(cmd)
#             PROMPT = '.*specific VNFD name.*'
#             interact.expect(PROMPT)
#             ##
#             cmd = 'n'
#             interact.send(cmd)
#             PROMPT = '.*input the specific name.*'
#             interact.expect(PROMPT, timeout=10)
#             ##
#             # cmd = self.media_ne_name
#             cmd = 'sbgw01'
#             interact.send(cmd)
#             PROMPT = '.*test@ngnsvr10.*'
#             interact.expect(PROMPT, timeout=10)
#             cmd_output = interact.current_output_clean
#             print(cmd_output)
#             ##
#             cmd = 'exit'
#             interact.send(cmd)
#             interact.expect()
#     except Exception:
#         traceback.print_exc()
#     finally:
#         try:
#             client.close()
#         except Exception:
#             pass

class MediaVnfArtsGenerator(object):
    def __init__(self, type='instantiation', version=''):
        self.type = type
        self.version = version
        self.arts_type = media_arts_type
        self.yact_server_ip = media_yact_server_ip
        self.yact_user = media_yact_user
        self.yact_passwd = media_yact_passwd
        self.yact_dir = media_yact_dir
        self.media_dif_name = media_dif_name
        self.media_su_dif_name = media_su_dif_name
        self.media_ne_name = media_vnf_name
        self.mmyact_cmd = '/home/test/bin/mmyact -u ' + media_dif_name
        self.media_vnfpkg_name = media_vnfpkg_name
        self.media_su_vnfpkg_name = media_su_vnfpkg_name
        self.extension_json_file = media_extension_json_file
        self.dr_extension_json_file = media_dr_extension_json_file
        self.instantiation_json_file = media_instantiation_json_file
        self.dr_instantiation_json_file = media_dr_instantiation_json_file
        self.os_passwd = os_passwd
        self.openrc_file = media_openrc_file
        self.backup_server_ip = backup_server_ip
        self.backup_server_login = backup_server_login
        self.backup_server_passwd = backup_server_passwd
        self.backup_server_dir = cbackup_server_dir
        self.backup_file_name = media_backup_file_name
        self.ssh_type = ssh_type
        self.backup_file = self.backup_server_dir + self.backup_file_name
        self.local_backup_dir = media_local_backup_dir
        self.ap_instantiation = {
            "inject_well_known_temp_key": media_inject_well_known_temp_key,
            "Scripturl": media_script_url
        }
        self.ap_dr_instantiation = {
            "inject_well_known_temp_key": media_inject_well_known_temp_key,
            "Scripturl": media_backup_url
        }

    def ship_uif(self):
        pass

    def run_interact_cmd(self, interact, cmd=None, prompt='', timeout=10):
        if cmd is None:
            interact.expect(prompt)
        else:
            interact.send(cmd)
            interact.expect(prompt)

    def run_mmyact(self):
        try:
            client = paramiko.SSHClient()
            client.load_system_host_keys()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(hostname=self.yact_server_ip, username=self.yact_user, password=self.yact_passwd)
            with SSHClientInteraction(client, timeout=10, buffer_size=10240, display=True) as interact:
                ## login yact server
                PROMPT = '.*[test@ngnsvr10 ~].*'
                interact.expect(PROMPT)
                ## cd to yact dir
                PROMPT = '.*[test@ngnsvr10 ~].*'
                cmd = 'cd ' + self.yact_dir
                interact.send(cmd)
                interact.expect(PROMPT)
                ## run mmyact
                PROMPT = '.*#?.*'
                # cmd = '/home/test/bin/mmyact -u user-input-openstack-R20.2-sbgw01.xlsx'
                cmd = self.mmyact_cmd
                interact.send(cmd)
                interact.expect(PROMPT)
                ## send 1
                PROMPT = '.*#?.*'
                cmd = '1'
                interact.send(cmd)
                interact.expect(PROMPT)
                ## choice a new path
                PROMPT = '.*choice a new path.*'
                cmd = 'y'
                interact.send(cmd)
                interact.expect(PROMPT)
                ## choose 1 Public Official Versions
                PROMPT = '.*#?.*'
                cmd = '1'
                interact.send(cmd)
                interact.expect(PROMPT)
                ## parser out version
                cmd_output = interact.current_output_clean
                tmp_list = cmd_output.split()
                data = None
                for ldata in tmp_list:
                    if self.version in ldata:
                        data = ldata
                if data == None:
                    log('Not able to find Public Official Versions of ' + self.version)
                    exit(1)
                idx = tmp_list.index(data)
                cmd = tmp_list[idx-1].split(')')[0]
                PROMPT = '.*choice a new dif template.*'
                interact.send(cmd)
                interact.expect(PROMPT, timeout=10)
                ##
                cmd = 'y'
                PROMPT = '.*#?.*'
                interact.send(cmd)
                interact.expect(PROMPT, timeout=60)
                ##
                cmd = '2'
                PROMPT = '.*#?.*'
                interact.send(cmd)
                interact.expect(PROMPT, timeout=10)
                ##
                cmd = self.openrc_file
                PROMPT = '.*specific VNFD name.*'
                interact.send(cmd)
                interact.expect(PROMPT, timeout=10)
                ##
                cmd = 'n'
                PROMPT = '.*input the specific name.*'
                interact.send(cmd)
                interact.expect(PROMPT, timeout=10)
                ##
                cmd = self.media_ne_name
                PROMPT = '.*test@ngnsvr10.*'
                interact.send(cmd)
                interact.expect(PROMPT, timeout=10)
                cmd_output = interact.current_output_clean
                ##
                cmd = 'exit'
                interact.send(cmd)
                interact.expect()
        except Exception:
            traceback.print_exc()
        finally:
            try:
                client.close()
            except Exception:
                pass

    def gen_arts_media(self):
        """
        Procedure:
        1. run mmyact on yact server
        2. ship the artifacts to local media-plane-arts
        3. generate json files
        """
        log('In Func:' + sys._getframe().f_code.co_name)
        os.chdir(media_arts_dir)

        if self.type not in self.arts_type:
            log('Only ' + str(self.arts_type) + ' Supported.')
            exit(1)

        ts = int(time.time())
        log('Current time is: ' + str(ts))

    def gen_extension_json(self):
        pass

    def gen_instantiate_json(self):
        pass


'''
LCM Test Driver
This is the upper caller
It use vnfdId to identify which vnf to operation on
It's stateless by its best
'''
class LcmTestDriver(object):
    def __init__(self, vnfdId):
        self.vnfdId = vnfdId

    def setup_sigDriver(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.restore_vnf()
        self.sigDriver.wait_processing()

    def setup_medDriver(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.medDriver = MediaVnfLcmTestDriver(MediaVnf(vnfdId=self.vnfdId, apiVersion='3'))
        self.medDriver.restore_vnf()
        self.medDriver.wait_processing()

    # In case of vnfdId changed (for example, DR)
    # recreate sigDriver with updated vnfdId
    def refresh_sigDriver(self, vnfdId):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.vnfdId = vnfdId
        # self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=vnfdId))

    def refresh_medDriver(self, vnfdId):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.vnfdId = vnfdId

    '''
    Sig Plane
    '''
    def sigvnf_CreateInstantiate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        # For CreateInstantiate, cann't use setup_sigDriver as vnf is not created yet
        # use SigVnfLcmTestDriver directly
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.create_instantiate()
        self.sigDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)
        # self.sigDriver.prep_bkserver_pubkey()

    def sigvnf_Terminate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.terminate()
        self.sigDriver.check_opstatus(operation='TERMINATE', operationName='', timeout=10)

    def sigvnf_Delete(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.delete()

    def sigvnf_Modify_DisableAutoScale(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_scale(enabled=False)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_Modify_EnableAutoScale(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_scale(enabled=True)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_Modify_DisableAutoBackup(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_backup(enabled=False)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_Modify_EnableAutoBackup(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_backup(enabled=True)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_ScaleOut(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.scale(scaleType='SCALE_OUT', step=1, aspectId="sc_Aspect")
        self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)

    def sigvnf_ScaleIn(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.scale(scaleType='SCALE_IN', step=1, aspectId="sc_Aspect")
        self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)

    def sigvnf_Scale_OutIn(self):
        # First to scale out SC to max number: 5
        # then scale in SC to initial count: depends on initial deployment
        # The system initial count at: /storage/auto_scale/initial_sys_config.json
        # To-do: add func to parser the file and get initial SC count
        # For now, use 1 as initial count
        initial_count = 1
        max_count = 5
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        sigArts = SigVnfArtsGenerator()
        initial_count = sigArts.get_initial_sc_value()
        current_count, max_count = self.sigDriver.get_scale_status(aspectId="sc_Aspect")
        log('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        while current_count < max_count:
            self.sigDriver.scale(scaleType='SCALE_OUT', step=1, aspectId="sc_Aspect")
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status(aspectId="sc_Aspect")
        while current_count > initial_count:
            self.sigDriver.scale(scaleType='SCALE_IN', step=1, aspectId="sc_Aspect")
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status(aspectId="sc_Aspect")

    def sigvnf_ScaleOutToMax(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        initial_count = 1
        max_count = 5
        self.setup_sigDriver()
        sigArts = SigVnfArtsGenerator()
        initial_count = sigArts.get_initial_sc_value()
        current_count, max_count = self.sigDriver.get_scale_status(aspectId="sc_Aspect")
        log('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        if current_count == max_count:
            log('current_count == max_count. No Scale-Out needed.')
            return
        while current_count < max_count:
            self.sigDriver.scale(scaleType='SCALE_OUT', step=1, aspectId="sc_Aspect")
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status(aspectId="sc_Aspect")

    def sigvnf_ScaleInToInit(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        initial_count = 1
        max_count = 5
        self.setup_sigDriver()
        sigArts = SigVnfArtsGenerator()
        initial_count = sigArts.get_initial_sc_value()
        current_count, max_count = self.sigDriver.get_scale_status(aspectId="sc_Aspect")
        log('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        if current_count == initial_count:
            log('current_count == initial_count. No Scale-In needed.')
            return
        while current_count > initial_count:
            self.sigDriver.scale(scaleType='SCALE_IN', step=1, aspectId="sc_Aspect")
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status(aspectId="sc_Aspect")

    def sigvnf_Heal_Single(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        log('vnfcInstanceId_List: ' + str(vnfcInstanceId_List))
        # Single VM heal, for all VMs, one VM each time
        for vnfcid in vnfcInstanceId_List:
            log('Start to Heal: ' + vnfcid)
            self.sigDriver.heal(vnfcInstanceId=[vnfcid], cause='Sig Plane VM Heal: ' + vnfcid)
            self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    def sigvnf_Heal_Single_Random(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        log('vnfcInstanceId_List: ' + str(vnfcInstanceId_List))
        # Single VM heal, randomly select 1 VM
        num = random.randint(0, len(vnfcInstanceId_List) - 1)
        vnfcid = vnfcInstanceId_List[num]
        log('Start to Heal: ' + vnfcid)
        self.sigDriver.heal(vnfcInstanceId=[vnfcid], cause='Sig Plane VM Heal: ' + vnfcid)
        self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    # Randomly heal single VM, mutiple times
    def sigvnf_Heal_Single_Random_MTs(self, times=5):
        log('In Func: ' + sys._getframe().f_code.co_name)
        log('times: ' + str(times))
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        log('vnfcInstanceId_List: ' + str(vnfcInstanceId_List))
        # Single VM heal, randomly select 1 VM, times times
        step = 0
        while step < times:
            log('the step: ' + str(step + 1))
            num = random.randint(0, len(vnfcInstanceId_List) - 1)
            vnfcid = vnfcInstanceId_List[num]
            log('Start to Heal: ' + vnfcid)
            self.sigDriver.heal(vnfcInstanceId=[vnfcid], cause='Sig Plane VM Heal: ' + vnfcid)
            self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)
            step = step + 1

    def sigvnf_Heal_Multiple(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        if not vnfcInstanceId_List:
            log('Failed to get vnfcInstanceId_List, Error out.')
            exit(1)
        log('vnfcInstanceId_List:' + str(vnfcInstanceId_List))
        maxHealListLength = int(len(vnfcInstanceId_List) / 2)
        start = 2
        end = maxHealListLength + 1
        healList = []
        for healNum in range(start, end):
            while len(healList) < healNum:
                num = random.randint(0, len(vnfcInstanceId_List) - 1)
                duplicated = False
                vnfc = vnfcInstanceId_List[num]
                if vnfc not in healList:
                    for vnfc_tmp in healList:
                        if vnfc[0:10] == vnfc_tmp[0:10]:
                            # duplicated vnfc, not insert in healList
                            duplicated = True
                            break
                    if not duplicated:
                        healList.append(vnfc)
            log('Multiple Heal: healNum:' + str(healNum) + ';healList:' + str(healList))
            log('Start to Heal: ' + str(healList))
            self.sigDriver.heal(vnfcInstanceId=healList, cause='Sig Plane Multiple VM Heal: ' + str(healList))
            self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    def sigvnf_Backup(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        # setup pubkey before backup
        # self.sigDriver.prep_bkserver_pubkey()
        self.sigDriver.rm_existing_raw_zip(sshtype='passwd')
        for ap in [
            self.sigDriver.additinalParams_Backup_Local,
            # Temporarily disable pubkey way backup, only use credentials way
            # self.sigDriver.additinalParams_Backup_Remote_1,
            # self.sigDriver.additinalParams_Backup_Remote_2,
            # self.sigDriver.additinalParams_Backup_Remote_12,
            self.sigDriver.additinalParams_Backup_Remote_Creds_1,
            self.sigDriver.additinalParams_Backup_Remote_Creds_2,
            self.sigDriver.additinalParams_Backup_Remote_Creds_12
        ]:
            self.sigDriver.custom_backup(ap)
            self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
            self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Local(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Local)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote1(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        # self.sigDriver.prep_bkserver_pubkey()
        self.sigDriver.rm_existing_raw_zip(sshtype='passwd')
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_1)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote2(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        # self.sigDriver.prep_bkserver_pubkey()
        self.sigDriver.rm_existing_raw_zip(sshtype='passwd')
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_2)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote12(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        # self.sigDriver.prep_bkserver_pubkey()
        self.sigDriver.rm_existing_raw_zip(sshtype='passwd')
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_12)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote_Creds1(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.rm_existing_raw_zip(sshtype='passwd')
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_Creds_1)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote_Creds2(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.rm_existing_raw_zip(sshtype='passwd')
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_Creds_2)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote_Creds12(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.rm_existing_raw_zip(sshtype='passwd')
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_Creds_12)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_ConnectSbcMediaVnf(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_connect_SBC_Media_VNF(self.sigDriver.additinalParams_Connection)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:connect_SBC_Media_VNF', timeout=20)

    def sigvnf_DisconnectSbcMediaVnf(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_disconnect_SBC_Media_VNF(self.sigDriver.additinalParams_Disconnection)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:disconnect_SBC_Media_VNF', timeout=20)

    # The pubkey needs to be setup before DB restore as this is the only way
    def sigvnf_DBRestore(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_dbrestore(self.sigDriver.additinalParams_DBRestore)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:DB_Restore', timeout=60)

    # For UpgradePrecheck, Upgrade1Apply, Upgrade2Activate, Upgrade3Commit,
    # assume the VNF package has been changed by ChangePackageVersion,
    # so use sig_vnfdId_SUToLoad as vnfdId
    # For ChangePackageVersion, need to pass in the vnfdId of the TO load
    def sigvnf_ChangePackageVersion(self, vnfdId_SU):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.change_package_version(vnfdId_SU)
        # Change Package version consists of: UPGRADE then MODIFY_INFO
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)
        time.sleep(60)

    def sigvnf_UpgradePrecheck(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_precheck()
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_precheck', timeout=20)

    def sigvnf_Upgrade1Apply(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_1_apply(self.sigDriver.additinalParams_Upgrade1Apply)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_1_apply', timeout=90)

    def sigvnf_Upgrade2Activate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_2_activate()
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_2_activate', timeout=30)

    def sigvnf_Upgrade3Commit(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_3_commit()
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_3_commit', timeout=90)

    def sigvnf_UpgradeArchive(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        for ap in [self.sigDriver.additinalParams_UpgradeArchive,
                   self.sigDriver.additinalParams_UpgradeArchive_Creds
                   ]:
            self.sigDriver.custom_upgrade_archive(ap)
            self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_archive', timeout=20)
            self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='cssu')

    def sigvnf_UpgradeArchive_Pubkey(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_archive(self.sigDriver.additinalParams_UpgradeArchive)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_archive', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='cssu')

    def sigvnf_UpgradeArchive_Creds(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_archive(self.sigDriver.additinalParams_UpgradeArchive_Creds)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_archive', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='cssu')

    def sigvnf_CSSUInstantiate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        # For CreateInstantiate, cann't use setup_sigDriver as vnf is not created yet
        # use SigVnfLcmTestDriver directly
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.cssu_create_instantiate()
        self.sigDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)

    def sigvnf_DRInstantiate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        # For CreateInstantiate, cann't use setup_sigDriver as vnf is not created yet
        # use SigVnfLcmTestDriver directly
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.dr_create_instantiate()
        self.sigDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)

    # rel is current release
    # For instantiation, vnf is not created yet, so not able to use setup_sigDriver()
    # need to call sigvnf_GenArtsInstantiation() directly in TS
    # def sigvnf_GenArts4Instantiation(self, rel='R20.0'):
    #     log('In Func: ' + sys._getframe().f_code.co_name)
    #     self.setup_sigDriver()
    #     sigArts = SigVnfArtsGenerator(type='instantiation', rel=rel)
    #     sigArts.prep_arts()

    # rel is current release
    # 6 server_type for DR: httpserver1, httpserver2, httpserver12
    #               bkupserver1, bkupserver2, bkupserver12
    def sigvnf_GenArts4DR(self, rel='', server_type='httpserver12'):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        current_count, max_count = self.sigDriver.get_scale_status()
        sigArts = SigVnfArtsGenerator(type='dr', rel=rel, server_type=server_type)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    # rel is TO load release
    def sigvnf_GenArts4SU(self, rel=''):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        current_count, max_count = self.sigDriver.get_scale_status()
        sigArts = SigVnfArtsGenerator(type='su', rel=rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    # rel is TO load release
    def sigvnf_GenArts4CSSU(self, rel='', server_type='httpserver1'):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        current_count, max_count = self.sigDriver.get_scale_status()
        sigArts = SigVnfArtsGenerator(type='cssu', rel=rel, server_type=server_type)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    # this is test for artifact generation for su, dr, cssu
    def sigvnf_GenArtsTests(self, rel=''):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        log('Generate DIF for Instantiation')
        sigArts = SigVnfArtsGenerator(type='instantiation', rel=rel)
        sigArts.prep_arts()
        current_count, max_count = self.sigDriver.get_scale_status()
        log('Generate DIF for DR')
        sigArts = SigVnfArtsGenerator(type='dr', rel=rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()
        log('Generate DIF for SU')
        sigArts = SigVnfArtsGenerator(type='su', rel=sig_toload_rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()
        log('Generate DIF for CSSU')
        sigArts = SigVnfArtsGenerator(type='cssu', rel=sig_toload_rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    '''
    Media Plane
    '''
    def medvnf_CreateInstantiate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.medDriver = MediaVnfLcmTestDriver(MediaVnf(vnfdId=self.vnfdId))
        self.medDriver.create_instantiate()
        self.medDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)
        self.medDriver.prep_bkserver_pubkey()

    def medvnf_Terminate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.terminate()
        self.medDriver.check_opstatus(operation='TERMINATE', operationName='', timeout=10)

    def medvnf_Delete(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.delete()

    def medvnf_Backup(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        # setup pubkey before backup
        self.medDriver.prep_bkserver_pubkey()
        # remove existing backup file on backup server
        self.medDriver.rm_backup_zip()
        for ap in [
            self.medDriver.additinalParams_Backup_Local,
            # Temporarily disable pubkey way backup, only use credentials way
            # self.medDriver.additinalParams_Backup_Remote_1,
            # self.medDriver.additinalParams_Backup_Remote_2,
            # self.medDriver.additinalParams_Backup_Remote_12,
            self.medDriver.additinalParams_Backup_Remote_Creds_1,
            self.medDriver.additinalParams_Backup_Remote_Creds_2,
            self.medDriver.additinalParams_Backup_Remote_Creds_12
        ]:
            self.medDriver.custom_backup(ap)
            self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_Backup_Local(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.rm_backup_zip()
        self.medDriver.custom_backup(self.medDriver.additinalParams_Backup_Local)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_Backup_Remote1(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.rm_backup_zip()
        self.medDriver.custom_backup(self.medDriver.additinalParams_Backup_Remote_1)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_Backup_Remote2(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.rm_backup_zip()
        self.medDriver.custom_backup(self.medDriver.additinalParams_Backup_Remote_2)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_Backup_Remote12(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.rm_backup_zip()
        self.medDriver.custom_backup(self.medDriver.additinalParams_Backup_Remote_12)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_Backup_Remote_Creds1(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.rm_backup_zip()
        self.medDriver.custom_backup(self.medDriver.additinalParams_Backup_Remote_Creds_1)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_Backup_Remote_Creds2(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.rm_backup_zip()
        self.medDriver.custom_backup(self.medDriver.additinalParams_Backup_Remote_Creds_2)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_Backup_Remote_Creds12(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.rm_backup_zip()
        self.medDriver.custom_backup(self.medDriver.additinalParams_Backup_Remote_Creds_12)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)

    def medvnf_ScaleOut(self, aspectId="pimAspect"):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.scale(scaleType='SCALE_OUT', step=1, aspectId=aspectId)
        self.medDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)

    def medvnf_ScaleIn(self, aspectId="pimAspect"):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.scale(scaleType='SCALE_IN', step=1, aspectId=aspectId)
        self.medDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)

    def medvnf_ScaleOutToMax(self, aspectId="pimAspect"):
        log('In Func: ' + sys._getframe().f_code.co_name)
        initial_count = 1
        max_count = 5
        self.setup_medDriver()
        current_count, max_count = self.medDriver.get_scale_status(aspectId=aspectId)
        log('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        if current_count == max_count:
            log('current_count == max_count. No Scale-Out needed.')
            return
        while current_count < max_count:
            self.medDriver.scale(scaleType='SCALE_OUT', step=1, aspectId=aspectId)
            self.medDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.medDriver.get_scale_status(aspectId=aspectId)

    def medvnf_ScaleInToInit(self, aspectId="pimAspect"):
        log('In Func: ' + sys._getframe().f_code.co_name)
        initial_count = 1
        max_count = 5
        self.setup_medDriver()
        current_count, max_count = self.medDriver.get_scale_status(aspectId=aspectId)
        log('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        if current_count == initial_count:
            log('current_count == initial_count. No Scale-In needed.')
            return
        while current_count > initial_count:
            self.medDriver.scale(scaleType='SCALE_IN', step=1, aspectId=aspectId)
            self.medDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.medDriver.get_scale_status(aspectId=aspectId)

    def medvnf_Heal_Single(self, vmType='ALL'):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.gen_stby_vm_list()
        if vmType not in ['ALL', 'PIM', 'MCM', 'SCM']:
            log('vmType not supported, Error out.')
            exit(1)
        if vmType == 'ALL':
            vmList = self.medDriver.stbyALL_list
        if vmType == 'PIM':
            vmList = self.medDriver.stbyPIM_list
        if vmType == 'MCM':
            vmList = self.medDriver.stbyMCM_list
        if vmType == 'SCM':
            vmList = self.medDriver.stbySCM_list
        log('vmList: ' + str(vmList))
        # Single VM heal, for all VMs, one VM each time
        for vmName in vmList:
            log('Start to Heal: ' + vmName)
            self.medDriver.heal(vmHealList=vmName, cause='Meda Plane VM Heal: ' + vmName)
            self.medDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    def medvnf_Heal_Single_Random(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        vmList = self.medDriver.stbyALL_list
        log('vmList: ' + str(vmList))
        # Single VM heal, randomly select 1 VM
        num = random.randint(0, len(vmList) - 1)
        vmName = vmList[num]
        log('Start to Heal: ' + vmName)
        self.medDriver.heal(vmHealList=vmName, cause='Meda Plane VM Heal: ' + vmName)
        self.medDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    def medvnf_Heal_Multiple(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.gen_stby_vm_list()
        vmList = self.medDriver.stbyALL_list
        if not vmList:
            log('Failed to get vmList, Error out.')
            exit(1)
        log('vmList:' + str(vmList))
        start = 1
        end = len(vmList) + 1
        healList = []
        for healNum in range(start, end):
            while len(healList) < healNum:
                num = random.randint(0, len(vmList) - 1)
                duplicated = False
                vmname = vmList[num]
                if vmname not in healList:
                    healList.append(vmname)
            vmHealList = ','.join(healList)
            log('Multiple Heal: healNum:' + str(healNum) + ';healList:' + vmHealList)
            log('Start to Heal: ' + vmHealList)
            self.medDriver.heal(vmHealList=vmHealList, cause='Media Plane Multiple VMs Heal: ' + vmHealList)
            self.medDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    def medvnf_Restore(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.custom_restore(self.medDriver.additinalParams_Restore)
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:Restore', timeout=60)

    def medvnf_DRInstantiate(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        # For CreateInstantiate, cann't use setup_medDriver as vnf is not created yet
        # use medVnfLcmTestDriver directly
        self.medDriver = MediaVnfLcmTestDriver(MediaVnf(vnfdId=self.vnfdId))
        self.medDriver.dr_create_instantiate()
        self.medDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)

    def medvnf_PostDR(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.custom_post_dr()
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:Post_Disaster_Recovery', timeout=60)

    def medvnf_DR(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.medDriver = MediaVnfLcmTestDriver(MediaVnf(vnfdId=self.vnfdId))
        self.medDriver.dr_create_instantiate()
        self.medDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)
        self.medDriver.custom_post_dr()
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:Post_Disaster_Recovery', timeout=60)

    # rel is TO load release
    def medvnf_GenArts4SU(self):
        pass

    def medvnf_ChangePackageVersion(self, vnfdId_SU):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.change_package_version(vnfdId_SU)
        # Change Package version consists of: UPGRADE then MODIFY_INFO
        self.medDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)
        time.sleep(60)

    def medvnf_ISSU(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.custom_issu()
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:issu', timeout=90)

    def medvnf_NSSU(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.custom_nssu()
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:nssu', timeout=90)

    def medvnf_Rollback(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.custom_rollback()
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:rollback', timeout=90)

    def medvnf_Backout(self):
        log('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_medDriver()
        self.medDriver.custom_backout()
        self.medDriver.check_opstatus(operation='OTHER', operationName='custom:backout', timeout=90)

    '''
    Following are test suites out of the box
    '''
    '''
    Sig VNF
    '''
    def sigvnf_tests_alpha(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_ScaleOut()
        self.sigvnf_ScaleIn()
        self.sigvnf_ScaleOut()
        # self.sigvnf_Heal_Single()
        self.sigvnf_Heal_Single_Random_MTs()
        self.sigvnf_Backup()
        self.sigvnf_DBRestore()

    def sigvnf_tests_beta(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_Scale_OutIn()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_Heal_Single()
        self.sigvnf_Heal_Multiple()
        self.sigvnf_Backup()
        self.sigvnf_DBRestore()
        self.sigvnf_Terminate()
        self.sigvnf_Delete()

    def sigvnf_tests_gamma(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_ScaleInToInit()
        self.sigvnf_Scale_OutIn()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_Heal_Single()
        self.sigvnf_Heal_Multiple()
        self.sigvnf_Backup()
        self.sigvnf_DBRestore()
        self.sigvnf_GenArts4DR(rel=sig_default_rel)
        self.sigvnf_tests_td()
        sigvnf_UploadVnfpkg(swVersion=sigVersion, type='dr')
        setup_vnfdIds()
        self.refresh_sigDriver(sig_vnfdId)
        self.sigvnf_tests_dr()

    def sigvnf_tests_cim(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        # self.sigvnf_ConnectSbcMediaVnf()
        # self.sigvnf_Backup_Remote_Creds1()

    def sigvnf_tests_cims(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        # self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_ScaleOutToMax()

    def sigvnf_tests_cimcb(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Backup_Remote_Creds1()

    def sigvnf_tests_cimcsb(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Backup_Remote_Creds1()
        # self.sigvnf_ScaleOutToMax()
        self.sigvnf_ScaleOut()
        self.sigvnf_Backup_Remote_Creds1()

    def sigvnf_tests_modify_da(self):
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()

    def sigvnf_tests_modify_ea(self):
        self.sigvnf_Modify_EnableAutoBackup()
        self.sigvnf_Modify_EnableAutoScale()

    def sigvnf_tests_td(self):
        self.sigvnf_Terminate()
        self.sigvnf_Delete()

    def sigvnf_tests_mcdb(self):
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Backup()

    def sigvnf_tests_bkup(self):
        self.sigvnf_Backup()

    def sigvnf_tests_bkup_rt(self):
        self.sigvnf_Backup_Remote1()
        self.sigvnf_Backup_Remote2()
        self.sigvnf_Backup_Remote12()

    def sigvnf_tests_bkup_rt_creds(self):
        self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_Backup_Remote_Creds2()
        self.sigvnf_Backup_Remote_Creds12()

    def sigvnf_tests_br(self):
        # self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_Backup_Remote_Creds12()
        self.sigvnf_DBRestore()

    def sigvnf_tests_heal(self):
        self.sigvnf_Heal_Single()
        self.sigvnf_Heal_Multiple()

    def sigvnf_tests_heals(self):
        self.sigvnf_Heal_Single()

    def sigvnf_tests_healsr(self):
        self.sigvnf_Heal_Single_Random()

    def sigvnf_tests_healsr_mt(self):
        self.sigvnf_Heal_Single_Random_MTs(times=5)

    def sigvnf_tests_healm(self):
        self.sigvnf_Heal_Multiple()

    def sigvnf_tests_scale(self):
        self.sigvnf_Scale_OutIn()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_ScaleInToInit()

    def sigvnf_tests_scale_outin(self):
        self.sigvnf_Scale_OutIn()

    def sigvnf_tests_scale2max(self):
        self.sigvnf_ScaleOutToMax()

    def sigvnf_tests_scale2init(self):
        self.sigvnf_ScaleInToInit()

    def sigvnf_tests_chgpkvern(self, vnfdId_SU):
        self.sigvnf_ChangePackageVersion(vnfdId_SU)

    def sigvnf_tests_su_orignal(self):
        self.sigvnf_UpgradePrecheck()
        self.sigvnf_Upgrade1Apply()
        self.sigvnf_Upgrade2Activate()
        self.sigvnf_Upgrade3Commit()

    def sigvnf_tests_ua(self):
        self.sigvnf_UpgradeArchive()

    def sigvnf_tests_ua_td(self):
        self.sigvnf_UpgradeArchive()
        # self.sigvnf_UpgradeArchive_Creds()
        self.sigvnf_Terminate()
        self.sigvnf_Delete()

    def sigvnf_tests_cssu_instantiate(self):
        self.sigvnf_CSSUInstantiate()

    def sigvnf_tests_dr_instantiate(self):
        self.sigvnf_DRInstantiate()

    def sigvnf_tests_dr(self, server_type='httpserver12'):
        self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_GenArts4DR(rel=sig_default_rel, server_type=server_type)
        self.sigvnf_tests_td()
        sigvnf_UploadVnfpkg(swVersion=sigVersion, type='dr')
        setup_vnfdIds()
        self.refresh_sigDriver(sig_vnfdId)
        self.sigvnf_DRInstantiate()
        self.sigvnf_tests_modify_da()

    def sigvnf_tests_cssu(self, server_type='httpserver1'):
        # self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_GenArts4CSSU(rel=sig_toload_rel, server_type=server_type)
        self.sigvnf_tests_ua_td()
        sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='cssu')
        setup_vnfdIds()
        self.refresh_sigDriver(sig_vnfdId_SU)
        self.sigvnf_CSSUInstantiate()
        self.sigvnf_tests_modify_da()

    def sigvnf_tests_su(self):
        self.sigvnf_GenArts4SU(rel=sig_toload_rel)
        sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='su')
        setup_vnfdIds()
        self.sigvnf_tests_chgpkvern(sig_vnfdId_SU)
        self.refresh_sigDriver(sig_vnfdId_SU)
        self.sigvnf_UpgradePrecheck()
        self.sigvnf_Upgrade1Apply()
        self.sigvnf_Upgrade2Activate()
        self.sigvnf_Upgrade3Commit()
        self.sigvnf_tests_modify_da()

    '''
    Media VNF
    '''
    def medvnf_tests_alpha(self):
        self.medvnf_tests_cib()
        self.medvnf_tests_bkup()
        self.medvnf_tests_scale_out_pim()
        self.medvnf_tests_scale_in_pim()
        self.medvnf_tests_scale_out_mcm()
        self.medvnf_tests_scale_in_mcm()
        self.medvnf_tests_heals_pim()
        self.medvnf_tests_heals_mcm()
        self.medvnf_tests_heals_scm()
        self.medvnf_tests_br()

    def medvnf_tests_beta(self):
        self.medvnf_tests_cib()
        self.medvnf_tests_bkup()
        self.medvnf_tests_scale2max_pim()
        self.medvnf_tests_scale2max_mcm()
        self.medvnf_tests_scale2min_pim()
        self.medvnf_tests_scale2min_mcm()
        self.medvnf_tests_heals_all()
        self.medvnf_tests_healm()
        self.medvnf_tests_br()

    def medvnf_tests_gamma(self):
        self.medvnf_tests_cib()
        self.medvnf_tests_bkup()
        self.medvnf_tests_scale2max_pim()
        self.medvnf_tests_scale2max_mcm()
        self.medvnf_tests_scale2min_pim()
        self.medvnf_tests_scale2min_mcm()
        self.medvnf_tests_scale2max_pim()
        self.medvnf_tests_scale2max_mcm()
        self.medvnf_tests_heals_all()
        self.medvnf_tests_healm()
        self.medvnf_tests_br()
        self.medvnf_tests_dr()
        self.medvnf_tests_td()

    def medvnf_tests_ci(self):
        self.medvnf_CreateInstantiate()

    def medvnf_tests_cib(self):
        self.medvnf_CreateInstantiate()
        self.medvnf_Backup_Remote_Creds1()

    def medvnf_tests_td(self):
        self.medvnf_Terminate()
        self.medvnf_Delete()

    def medvnf_tests_bkup(self):
        self.medvnf_Backup()

    def medvnf_tests_scale_out_pim(self):
        self.medvnf_ScaleOut(aspectId='pimAspect')

    def medvnf_tests_scale_in_pim(self):
        self.medvnf_ScaleIn(aspectId='pimAspect')

    def medvnf_tests_scale_out_mcm(self):
        self.medvnf_ScaleOut(aspectId='mcmAspect')

    def medvnf_tests_scale_in_mcm(self):
        self.medvnf_ScaleIn(aspectId='mcmAspect')

    def medvnf_tests_scale2max_pim(self):
        self.medvnf_ScaleOutToMax(aspectId='pimAspect')

    def medvnf_tests_scale2max_mcm(self):
        self.medvnf_ScaleOutToMax(aspectId='mcmAspect')

    def medvnf_tests_scale2min_pim(self):
        self.medvnf_ScaleInToInit(aspectId='pimAspect')

    def medvnf_tests_scale2min_mcm(self):
        self.medvnf_ScaleInToInit(aspectId='mcmAspect')

    def medvnf_tests_scaleOutIn_pim(self):
        self.medvnf_ScaleOutToMax(aspectId='pimAspect')
        self.medvnf_ScaleInToInit(aspectId='pimAspect')

    def medvnf_tests_scaleOutIn_mcm(self):
        self.medvnf_ScaleOutToMax(aspectId='mcmAspect')
        self.medvnf_ScaleInToInit(aspectId='mcmAspect')

    def medvnf_tests_scaleOutIn_pim_mcm(self):
        self.medvnf_ScaleOutToMax(aspectId='pimAspect')
        self.medvnf_ScaleOutToMax(aspectId='mcmAspect')
        self.medvnf_ScaleInToInit(aspectId='pimAspect')
        self.medvnf_ScaleInToInit(aspectId='mcmAspect')

    def medvnf_tests_heals_pim(self):
        self.medvnf_Heal_Single(vmType='PIM')

    def medvnf_tests_heals_mcm(self):
        self.medvnf_Heal_Single(vmType='MCM')

    def medvnf_tests_heals_scm(self):
        self.medvnf_Heal_Single(vmType='SCM')

    def medvnf_tests_heals_all(self):
        self.medvnf_Heal_Single(vmType='ALL')

    def medvnf_tests_healm(self):
        self.medvnf_Heal_Multiple()

    def medvnf_tests_restore(self):
        self.medvnf_Restore()

    def medvnf_tests_br(self):
        self.medvnf_Backup()
        self.medvnf_Restore()

    def medvnf_tests_dr(self):
        self.medvnf_Backup_Remote_Creds12()
        self.medvnf_tests_td()
        setup_vnfdIds()
        self.refresh_medDriver(media_vnfdId)
        self.medvnf_DR()

    def medvnf_tests_post_dr(self):
        # self.medvnf_DRInstantiate()
        self.medvnf_PostDR()

    def medvnf_tests_issu(self):
        # self.medvnf_GenArts4SU()
        self.medvnf_Backup_Remote_Creds12()
        mediavnf_UploadVnfpkg(swVersion=mediaVersion_SU, type='su')
        setup_vnfdIds()
        self.medvnf_ChangePackageVersion(media_vnfdId_SU)
        self.refresh_medDriver(media_vnfdId_SU)
        self.medvnf_ISSU()

    def medvnf_tests_nssu(self):
        # self.medvnf_GenArts4SU()
        self.medvnf_Backup_Remote_Creds12()
        mediavnf_UploadVnfpkg(swVersion=mediaVersion_SU, type='su')
        setup_vnfdIds()
        self.medvnf_ChangePackageVersion(media_vnfdId_SU)
        self.refresh_medDriver(media_vnfdId_SU)
        self.medvnf_NSSU()

    def medvnf_tests_rollback(self):
        setup_vnfdIds()
        self.medvnf_ChangePackageVersion(media_vnfdId)
        self.refresh_medDriver(media_vnfdId)
        self.medvnf_Rollback()

    def medvnf_tests_backout(self):
        setup_vnfdIds()
        self.medvnf_ChangePackageVersion(media_vnfdId)
        self.refresh_medDriver(media_vnfdId)
        self.medvnf_Backout()


'''
Upper caller Test Suite
'''
'''
SBC Test Suite
'''
def TS_sbcvnf_tests_alpha():
    TS_medvnf_tests_alpha()
    TS_sigvnf_tests_alpha()

def TS_sbcvnf_tests_beta():
    TS_medvnf_tests_beta()
    TS_sigvnf_tests_beta()

def TS_sbcvnf_tests_gamma():
    TS_medvnf_tests_gamma()
    TS_sigvnf_tests_gamma()

def TS_sbcvnf_tests_ts01():
    """
    This test suite focuses on backup, restore and DR.
    This test suite covers following scenarios:
    1. signaling plane and media plane instantiation
    2. signaling plane connection to media plane
    3. media plane scale-out of PIM to max
    4. media plane scale-out of MCM to max
    5. signaling plane scale-out of SC to max
    6. signaling plane backup
    7. media plane backup
    8. signaling plane DB restore
    9. media plane restore
    10. signaling plane DR
    11. media plane DR
    """
    TS_medvnf_tests_ci()
    TS_sigvnf_tests_cimcb()
    TS_medvnf_tests_scale_out_all()
    TS_sigvnf_tests_scale2max()
    TS_sigvnf_tests_br()
    TS_medvnf_tests_br()
    TS_sigvnf_tests_dr_http1()
    TS_medvnf_tests_dr()

def TS_sbcvnf_tests_ts02():
    """
    This test suite focuses on scale, healing.
    This test suite covers following scenarios:
    1. signaling plane and media plane instantiation
    2. signaling plane connection to media plane
    3. media plane scale-out of PIM to max
    4. media plane scale-out of MCM to max
    5. media plane scale-in of PIM to min
    6. media plane scale-in of MCM to min
    7. signaling plane scale-out of SC to max
    8. signaling plane scale-in of SC to min
    9. media plane scale-out of PIM to max
    10. media plane scale-out of MCM to max
    11. signaling plane scale-out of SC to max
    12. media plane single VM healing
    13. media plane multiple VM healing
    14. signaling plane single VM healing
    15. signaling plane multiple VM healing
    16. signaling plane terminate and delete
    17. media plane terminate and delete
    """
    TS_medvnf_tests_ci()
    TS_sigvnf_tests_cimcb()
    TS_medvnf_tests_scale_out_all()
    TS_medvnf_tests_scale_in_all()
    TS_sigvnf_tests_scale2max()
    TS_sigvnf_tests_scale2init()
    TS_medvnf_tests_scale_out_all()
    TS_sigvnf_tests_scale2max()
    TS_medvnf_tests_heals_all()
    TS_medvnf_tests_healm()
    TS_sigvnf_tests_heal_s()
    TS_sigvnf_tests_heal_m()
    TS_sigvnf_tests_td()
    TS_medvnf_tests_td()

def TS_sbcvnf_tests_ts03():
    """
    This test suite focuses on SU.
    This test suite covers following scenarios:
    1. signaling plane and media plane instantiation
    2. signaling plane connection to media plane
    3. media plane scale-out of PIM to max
    4. media plane scale-out of MCM to max
    5. signaling plane scale-out of SC to max
    6. media plane ISSU
    7. media plane rollback
    8. media plane NSSU
    9. media plane rollback
    10. signaling plane SU
    11. signaling plane terminate and delete
    12. media plane terminate and delete
    """
    TS_medvnf_tests_ci()
    TS_sigvnf_tests_cimcb()
    TS_medvnf_tests_scale_out_all()
    TS_sigvnf_tests_scale2max()
    TS_medvnf_tests_issu_rollback()
    TS_medvnf_tests_nssu_rollback()
    TS_sigvnf_tests_su()
    TS_sigvnf_tests_td()
    TS_medvnf_tests_td()

def TS_sbcvnf_tests_ts04():
    """
    This test suite focuses on SU and CSSU.
    This test suite covers following scenarios:
    1. signaling plane and media plane instantiation
    2. signaling plane connection to media plane
    3. media plane scale-out of PIM to max
    4. media plane scale-out of MCM to max
    5. signaling plane scale-out of SC to max
    6. media plane ISSU
    7. media plane rollback
    8. media plane NSSU
    9. media plane rollback
    10. signaling plane CSSU
    11. signaling plane terminate and delete
    12. media plane terminate and delete
    """
    TS_medvnf_tests_ci()
    TS_sigvnf_tests_cimcb()
    TS_medvnf_tests_scale_out_all()
    TS_sigvnf_tests_scale2max()
    TS_medvnf_tests_issu_rollback()
    TS_medvnf_tests_nssu_rollback()
    TS_sigvnf_tests_cssu_creds_1()
    TS_sigvnf_tests_td()
    TS_medvnf_tests_td()

def TS_sbcvnf_tests_ts05():
    """
    This test suite focuses on SU and LCM operations after SU.
    This test suite covers following scenarios:
    1. signaling plane and media plane instantiation
    2. signaling plane connection to media plane
    3. media plane scale-out of PIM to max
    4. media plane scale-out of MCM to max
    5. signaling plane scale-out of SC to max
    6. media plane ISSU
    7. signaling plane SU
    8. media plane scale-in of PIM to min
    9. media plane scale-in of MCM to min
    10. media plane scale-out of PIM to max
    11. media plane scale-out of MCM to max
    12. signaling plane scale-in of SC to min
    13. signaling plane scale-out of SC to max
    14. media plane backup
    15. media plane restore
    16. signaling plane backup
    17. signaling plane restore
    18. media plane multiple VM healing
    19. Sigaling plane multiple VM healing
    20. signaling plane terminate and delete
    21. media plane terminate and delete
    """
    TS_medvnf_tests_ci()
    TS_sigvnf_tests_cimcb()
    TS_medvnf_tests_scale_out_all()
    TS_sigvnf_tests_scale2max()
    TS_medvnf_tests_issu()
    TS_sigvnf_tests_su()
    TS_medvnf_tests_scale_in_all_toload()
    TS_medvnf_tests_scale_out_all_toload()
    TS_sigvnf_tests_scale2min_toload()
    TS_sigvnf_tests_scale2max_toload()
    TS_medvnf_tests_br_toload()
    TS_sigvnf_tests_br_toload()
    TS_medvnf_tests_healm_toload()
    TS_sigvnf_tests_heal_m_toload()
    TS_sigvnf_tests_td_toload()
    TS_medvnf_tests_td_toload()

'''
Sig Test Suite
'''
def TS_sigvnf_tests_alpha():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_alpha()

def TS_sigvnf_tests_beta():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_beta()

def TS_sigvnf_tests_gamma():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_gamma()

def TS_sigvnf_tests_arts():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_GenArtsTests()

def TS_sigvnf_tests_td():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_td()

def TS_sigvnf_tests_td_toload():
    lcmDriver = LcmTestDriver(sigVersion_SU)
    lcmDriver.sigvnf_tests_td()

def TS_sigvnf_tests_cim():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cim()

def TS_sigvnf_tests_cimcb():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cimcb()

def TS_sigvnf_tests_cimcsb():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cimcsb()

def TS_sigvnf_tests_mcdb():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_mcdb()

def TS_sigvnf_tests_bkup():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_bkup()

def TS_sigvnf_tests_bkup_rt():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_bkup_rt()

def TS_sigvnf_tests_bkup_rt_creds():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_bkup_rt_creds()

def TS_sigvnf_tests_br():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_br()

def TS_sigvnf_tests_br_toload():
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_br()

def TS_sigvnf_tests_heal():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_heal()

def TS_sigvnf_tests_heal_m():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_healm()

def TS_sigvnf_tests_heal_m_toload():
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_healm()

def TS_sigvnf_tests_heal_s():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_heals()

def TS_sigvnf_tests_heal_s_r():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_healsr()

def TS_sigvnf_tests_heal_s_r_mt():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_healsr_mt()

def TS_sigvnf_tests_scale():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale()

def TS_sigvnf_tests_scale_outin():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale_outin()

def TS_sigvnf_tests_scale_toload():
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_scale()

def TS_sigvnf_tests_scale2max():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale2max()

def TS_sigvnf_tests_scale2max_toload():
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_scale2max()
    
def TS_sigvnf_tests_scale2init():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale2init()

def TS_sigvnf_tests_scale2init_toload():
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_scale2init()

def TS_sigvnf_tests_sh():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale2max()
    lcmDriver.sigvnf_tests_heal()

# def TS_sigvnf_tests_su():
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_GenArts4SU(rel=toload_rel)
#     sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='su')
#     setup_vnfdIds()
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_tests_chgpkvern(sig_vnfdId_SU)
#     lcmDriver = LcmTestDriver(sig_vnfdId_SU)
#     lcmDriver.sigvnf_tests_su()

# def TS_sigvnf_tests_cssu():
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_GenArts4CSSU(rel=toload_rel)
#     lcmDriver.sigvnf_tests_ua_td()
#     sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='cssu')
#     setup_vnfdIds()
#     lcmDriver = LcmTestDriver(sig_vnfdId_SU)
#     lcmDriver.sigvnf_tests_cssu()

# backup -> gen arts for DR -> terminate, delete
# -> upload new vnf pkg -> DR
# The original version for dr
# def TS_sigvnf_tests_dr():
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_Backup_Remote_Creds1()
#     lcmDriver.sigvnf_GenArts4DR(rel=default_rel, server_type='httpserver12')
#     lcmDriver.sigvnf_tests_td()
#     sigvnf_UploadVnfpkg(swVersion=sigVersion, type='dr')
#     setup_vnfdIds()
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_tests_dr()

def TS_sigvnf_tests_dr():
    for server_type in sig_server_type_list:
        lcmDriver = LcmTestDriver(sig_vnfdId)
        lcmDriver.sigvnf_tests_dr(server_type=server_type)

def TS_sigvnf_tests_dr_http1():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr(server_type='httpserver1')

def TS_sigvnf_tests_dr_http2():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr(server_type='httpserver2')

def TS_sigvnf_tests_dr_http12():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr(server_type='httpserver12')

def TS_sigvnf_tests_dr_creds_1():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr(server_type='bkupserver1')

def TS_sigvnf_tests_dr_creds_2():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr(server_type='bkupserver2')

def TS_sigvnf_tests_dr_creds_12():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr(server_type='bkupserver12')

# Not completed yet
def TS_sigvnf_tests_cssu():
    for server_type in ['httpserver1', 'bkupserver1']:
        lcmDriver = LcmTestDriver(sig_vnfdId)
        lcmDriver.sigvnf_tests_cssu(server_type=server_type)
        setup_vnfdIds()
        lcmDriver.refresh_sigDriver(sig_vnfdId_SU)
        lcmDriver.sigvnf_tests_td()
        sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
        setup_vnfdIds()
        lcmDriver = LcmTestDriver(sig_vnfdId)
        lcmDriver.sigvnf_tests_cim()

# Create -> Instantiation -> Modify Auto Operations -> Connection2Media -> Backup
# -> UpgradeArchive -> Terminate -> Delete -> CSSU Instantiation
# def TS_sigvnf_tests_cim_cssu_http():
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_tests_cim()
#     lcmDriver.sigvnf_GenArts4CSSU(rel=toload_rel, server_type='httpserver1')
#     lcmDriver.sigvnf_tests_ua_td()
#     sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='cssu')
#     lcmDriver = LcmTestDriver(sig_vnfdId_SU)
#     lcmDriver.sigvnf_tests_cssu()

def TS_sigvnf_tests_cim_cssu_http():
    server_type = 'httpserver1'
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cim()
    lcmDriver.sigvnf_tests_cssu(server_type=server_type)

def TS_sigvnf_tests_cim_cssu_creds():
    server_type = 'bkupserver1'
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cim()
    lcmDriver.sigvnf_tests_cssu(server_type=server_type)

def TS_sigvnf_tests_cssu_http1():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cssu(server_type='httpserver1')

def TS_sigvnf_tests_cssu_creds_1():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cssu(server_type='bkupserver1')

def TS_sigvnf_tests_su():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_su()

def TS_sigvnf_tests_cim_su():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cim()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_su()

def TS_sigvnf_tests_cims_su():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cims()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_su()

'''
Media Test Suite
'''
def TS_medvnf_tests_alpha():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_alpha()

def TS_medvnf_tests_beta():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_beta()

def TS_medvnf_tests_gamma():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_gamma()

def TS_medvnf_tests_getvnf():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.setup_medDriver()
    data = lcmDriver.medDriver.sbcvnf.get_vnf()
    log('Media Plane vnf data: ' + str(data))

def TS_medvnf_tests_td():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_td()

def TS_medvnf_tests_td_toload():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_td()

def TS_medvnf_tests_ci():
    # medvnf_GenArtsInstantiation()
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_ci()

def TS_medvnf_tests_cib():
    # medvnf_GenArtsInstantiation()
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_cib()

def TS_medvnf_tests_bkup():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_bkup()

def TS_medvnf_tests_scale_out_pim():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale_out_pim()

def TS_medvnf_tests_scale_in_pim():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale_in_pim()

def TS_medvnf_tests_scale_out_mcm():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale_out_mcm()

def TS_medvnf_tests_scale_in_mcm():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale_in_mcm()

def TS_medvnf_tests_scale_out_in_pim():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale_out_pim()
    lcmDriver.medvnf_tests_scale_in_pim()

def TS_medvnf_tests_scale_out_in_mcm():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale_out_mcm()
    lcmDriver.medvnf_tests_scale_in_mcm()

def TS_medvnf_tests_scale_pim():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2min_pim()

def TS_medvnf_tests_scale_mcm():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale2max_mcm()
    lcmDriver.medvnf_tests_scale2min_mcm()

def TS_medvnf_tests_scale():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2max_mcm()
    lcmDriver.medvnf_tests_scale2min_pim()
    lcmDriver.medvnf_tests_scale2min_mcm()

def TS_medvnf_tests_scale_out_all():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2max_mcm()

def TS_medvnf_tests_scale_out_all_toload():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2max_mcm()

def TS_medvnf_tests_scale_in_all():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scale2min_pim()
    lcmDriver.medvnf_tests_scale2min_mcm()

def TS_medvnf_tests_scale_in_all_toload():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_scale2min_pim()
    lcmDriver.medvnf_tests_scale2min_mcm()

def TS_medvnf_tests_scale_out_in_all():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_scaleOutIn_pim_mcm()

def TS_medvnf_tests_heals_pim():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_heals_pim()

def TS_medvnf_tests_heals_mcm():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_heals_mcm()

def TS_medvnf_tests_heals_scm():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_heals_scm()

def TS_medvnf_tests_heals_all():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_heals_all()

def TS_medvnf_tests_healm():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_healm()

def TS_medvnf_tests_healm_toload():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_healm()

def TS_medvnf_tests_restore():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_restore()

def TS_medvnf_tests_br():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_br()

def TS_medvnf_tests_br_toload():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_br()

def TS_medvnf_tests_dr():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_dr()

def TS_medvnf_tests_dr_toload():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_dr()

# def TS_medvnf_tests_post_dr():
#     lcmDriver = LcmTestDriver(media_vnfdId)
#     lcmDriver.medvnf_tests_post_dr()

def TS_medvnf_tests_issu():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_issu()

def TS_medvnf_tests_nssu():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_nssu()

def TS_medvnf_tests_rollback():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_rollback()

def TS_medvnf_tests_backout():
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_backout()

def TS_medvnf_tests_issu_rollback():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_issu()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_rollback()

def TS_medvnf_tests_nssu_rollback():
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_nssu()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_rollback()

def TS_medvnf_tests_cib_issu_rollback():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_cib()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_issu()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_rollback()

def TS_medvnf_tests_cibsb_issu_rollback():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_cib()
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2max_mcm()
    lcmDriver.medvnf_tests_bkup()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_issu()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_rollback()

def TS_medvnf_tests_cib_nssu_rollback():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_cib()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_nssu()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_rollback()

def TS_medvnf_tests_cibsb_nssu_rollback():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_cib()
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2max_mcm()
    lcmDriver.medvnf_tests_bkup()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_nssu()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_rollback()

def TS_medvnf_tests_cib_issu_scale_healm_br():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_cib()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_issu()
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2max_mcm()
    lcmDriver.medvnf_tests_healm()
    lcmDriver.medvnf_tests_br()

def TS_medvnf_tests_cib_nssu_scale_healm_br():
    mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_cib()
    lcmDriver = LcmTestDriver(media_vnfdId)
    lcmDriver.medvnf_tests_nssu()
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(media_vnfdId_SU)
    lcmDriver.medvnf_tests_scale2max_pim()
    lcmDriver.medvnf_tests_scale2max_mcm()
    lcmDriver.medvnf_tests_healm()
    lcmDriver.medvnf_tests_br()

'''
Setup Env Variables
'''
def setup_env():
    # Disable following warning:
    # D:\Program Files (x86)\Python37-32\lib\site-packages\urllib3\connectionpool.py:1004:
    # InsecureRequestWarning: Unverified HTTPS request is being made.
    # Adding certificate verification is strongly advised.
    # See: https://urllib3.readthedocs.io/en/latest/advanced-usage.html#ssl-warnings
    urllib3.disable_warnings()

'''
Setup Logging
'''
def setup_logging():
    logger.setLevel(logging.INFO)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    fh = logging.FileHandler(log_file)
    fh.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s %(levelname)s:%(name)s:%(message)s')
    ch.setFormatter(formatter)
    fh.setFormatter(formatter)
    logger.addHandler(ch)
    logger.addHandler(fh)

'''
Setup VnfdIds
Here is to use software version and vnfdId to identify vnf
TODO: use vnf name to identify vnf
'''
def setup_vnfdIds():
    # First to find vnfdId for both signaling plane and media plane
    global sig_vnfdId
    global media_vnfdId
    global sig_vnfdId_SU
    global media_vnfdId_SU

    vnfpkgs = VnfPkgs()
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()

    # Determine vnfdId via svnfSoftwareVersion
    # For sig plane, example of vnfSoftwareVersion:
    #'vnfSoftwareVersion': 'sbclcm01~37.28.06'
    # 'vnfSoftwareVersion': 'sbclcm01~37.28.06.0020'
    if vnfpkgs.vnfpkgs_list:
        for vp in vnfpkgs.vnfpkgs_list:
            if vp.vnfProductName == 'SBC':
                if vp.vnfSoftwareVersion.endswith(sigVersion):
                    # This is sig plane vnfp. Normal case for sig plane
                    sig_vnfdId = vp.vnfdId
                    log('Sig Plane vnfdId: ' + sig_vnfdId)
                    log('Sig Plane vnfSoftwareVersion: ' + vp.vnfSoftwareVersion)
                elif vp.vnfSoftwareVersion.endswith(sigVersion_SU):
                    # This is sig plane vnfp. SU case for sig plane
                    sig_vnfdId_SU = vp.vnfdId
                    log('Sig Plane vnfdId_SU: ' + sig_vnfdId_SU)
                    log('Sig Plane vnfSoftwareVersion for SUTOLoad: ' + vp.vnfSoftwareVersion)
            elif vp.vnfProductName == 'SBC-media':
                if vp.vnfSoftwareVersion.endswith(mediaVersion):
                    # This is media plane vnfp. Normal case for media plane
                    media_vnfdId = vp.vnfdId
                    log('Media Plane vnfdId: ' + media_vnfdId)
                    log('Media Plane vnfSoftwareVersion: ' + vp.vnfSoftwareVersion)
                elif vp.vnfSoftwareVersion.endswith(mediaVersion_SU):
                    # This is media plane vnfp. SU case for media plane
                    media_vnfdId_SU = vp.vnfdId
                    log('Media Plane vnfdId_SU: ' + media_vnfdId_SU)
                    log('Media Plane vnfSoftwareVersion for SUTOLoad: ' + vp.vnfSoftwareVersion)
            else:
                log('Currently supported vnfProductName: SBC or SBC-media.')
                exit(1)

'''
Run various Test Suites
'''
def run_ts():
    '''
    This is the upper caller and control for various test suilts
    Sig plane and media plane have different test suilts
    '''

    '''
    Following are for sig VNF tests
    '''
    # sigvnf_GenArts()
    # sigvnf_GenArtsInstantiation()
    # sigvnf_GenArtsDR()
    # sigvnf_GenArtsSU()
    # sigvnf_GenArtsCSSU()

    # TS_sigvnf_tests_td()
    # TS_sigvnf_tests_cim()
    # TS_sigvnf_tests_cimcb()
    # TS_sigvnf_tests_mcdb()
    # TS_sigvnf_tests_bkup_rt_creds()
    # TS_sigvnf_tests_arts()
    # TS_sigvnf_tests_dr()
    # TS_sigvnf_tests_cssu()
    # TS_sigvnf_tests_td()
    # TS_sigvnf_tests_td_toload()
    # TS_sigvnf_tests_su()

    # TS_sigvnf_tests_scale2max()
    # TS_sigvnf_tests_scale2init()
    # TS_sigvnf_tests_scale_outin()

    # TS_sigvnf_tests_alpha()

    # TS_sigvnf_tests_beta()

    # TS_sigvnf_tests_cimcsb()

    # TS_sigvnf_tests_td()
    # TS_sigvnf_tests_cimcsb()

    # TS_sigvnf_tests_heal_m()
    # TS_sigvnf_tests_heal_s()

    # TS_sigvnf_tests_heal_s_r()
    # TS_sigvnf_tests_heal_s_r_mt()
    #
    # TS_sigvnf_tests_bkup()
    TS_sigvnf_tests_br()

    # TS_sigvnf_tests_dr()
    # TS_sigvnf_tests_dr_http12()
    # TS_sigvnf_tests_dr_creds_1()
    # TS_sigvnf_tests_dr_creds_2()
    # TS_sigvnf_tests_dr_creds_12()

    # TS_sigvnf_tests_su()
    # TS_sigvnf_tests_cim_su()

    # TS_sigvnf_tests_cssu_http1()
    # TS_sigvnf_tests_cssu_creds_1()
    # TS_sigvnf_tests_cim_cssu_http()
    # TS_sigvnf_tests_cim_cssu_creds()

    '''
    Following are for media VNF tests
    '''
    # mediavnf_UploadVnfpkg(swVersion=mediaVersion, type='instantiation')
    # mediavnf_DeleteVnfpkg(swVersion=mediaVersion)
    # medvnf = MediaVnf(vnfdId = media_vnfdId, apiVersion='3')
    # medvnf.create_vnf()
    # medvnf.modify_extension_vnf()
    # medvnf.instantiate_vnf()

    # TS_medvnf_tests_getvnf()

    # TS_medvnf_tests_td()
    # TS_medvnf_tests_ci()
    # TS_medvnf_tests_cib()

    # TS_medvnf_tests_bkup()

    # TS_medvnf_tests_scale_out_pim()
    # TS_medvnf_tests_scale_out_in_pim()
    # TS_medvnf_tests_scale_out_in_mcm()
    # TS_medvnf_tests_scale_out_all()
    # TS_medvnf_tests_scale_in_all()
    # TS_medvnf_tests_scale()
    # TS_medvnf_tests_restore()

    # TS_medvnf_tests_bkup()

    # TS_medvnf_tests_br()

    # TS_medvnf_tests_dr()

    # TS_medvnf_tests_heals_scm()
    # TS_medvnf_tests_heals_mcm()
    # TS_medvnf_tests_heals_pim()
    # TS_medvnf_tests_heals_all()
    # TS_medvnf_tests_healm()

    # TS_medvnf_tests_issu()
    # TS_medvnf_tests_rollback()
    # TS_medvnf_tests_backout()
    # TS_medvnf_tests_issu_rollback()
    # TS_medvnf_tests_nssu_rollback()
    # TS_medvnf_tests_cib_issu_rollback()


'''
Main
 Some functions:
 setup_env() : disable InsecureRequestWarning by urllib3
 setup_logging() : setup logger. use info as default
 setup_vnfdIds() : read all vnf pkgs. determine vnfdId of base load and TO load via vnfSoftwareVersion
 sigvnf_GenArts() : generate artifacts directly based on DIFs that have already been ready (manually prepared)
                    this is different with the methods sigvnf_GenArts4XX in class LcmTestDriver
                    the LcmTestDriver->sigvnf_GenArts4XX will read sc values from vnf then modify DIFs then gen arts
'''

if __name__ == '__main__':

    setup_env()

    setup_logging()

    setup_globals()

    setup_vnfdIds()

    run_ts()
