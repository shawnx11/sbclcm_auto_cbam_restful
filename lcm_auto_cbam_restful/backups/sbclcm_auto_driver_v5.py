# -*- coding:UTF-8 -*-

import requests
from requests.adapters import HTTPAdapter
import urllib3
import json
import time
import os
import sys
import subprocess
import random
import paramiko
import logging
import shutil
import zipfile
import openpyxl
from openpyxl import load_workbook

########################################################################################################################
# Globals
# Following is for cbam 10.75.44.20, CBAM19.5, on PL node-011
cbam_url = 'https://10.75.44.20'
client_id = 'cbam_rest'
client_secret = 'ed5683fb-2af7-45b5-be63-78b4e4c37bf5'
gui_client_id = 'lcm'
gui_client_passwd = '-Assured11'

# Following is for cbam 100.69.127.147, CBAM19.5, on PL node-043
# cbam_url = 'https://100.69.127.147'
# client_id = 'cbam_rest'
# client_secret = '26fe9a4a-5836-42a1-9ef5-fb0404675d60'
# gui_client_id = 'lcm'
# gui_client_passwd = '-Assured11'

working_dir = r'D:\Programs\JetBrains\PycharmProjects\py37projects\lcm_auto_cbam_restful'
sig_data_dir = working_dir + r'\data\sig-plane'
media_data_dir = working_dir + r'\data\media-plane'
sig_arts_dir = working_dir + r'\data\sig-plane-arts'
media_arts_dir = working_dir + r'\data\media-plane-arts'
log_file = working_dir + r'\sbclcm_auto.log'

global logger
logger = logging.getLogger()

vnflcm_base_path = cbam_url + '/vnflcm/v1'
vnfpkgm_base_path = cbam_url + '/vnfpkgm/v1'

operationState_list = ['STARTING', 'FAILED', 'ROLLED_BACK', 'PROCESSING', 'COMPLETED']
operation_list = ['MODIFY_INFO', 'INSTANTIATE', 'TERMINATE', 'SCALE', 'OTHER']

sig_vnfdId = ''
media_vnfdId = ''
sig_vnfdId_SU = ''
media_vnfdId_SU = ''
# # R20.0
# sigVersion = '37.28.06'
# sigVersion_SU = '37.28.06.0020'
# mediaVersion = 'an100052'
# mediaVersion_SU = 'an100053'
# R20.2
sigVersion = '37.34.00.6200'
sigVersion_SU = '37.34.06'
mediaVersion = 'ap100007'
mediaVersion_SU = 'ap100008'

########################################################################################################################
# Util Functions
def log():
    pass

def print_response(response):
    print('response: ', response)
    print('response.headers: ', response.headers)
    print('response.text: ', response.text)

def print_response_details(response):
    print('response: ', response)
    if response.headers:
        print('resposne.headers:')
        for i, j in response.headers.items():
            print(i, j)
    if response.request.headers:
        print('response.request.headers:')
        for i, j in response.request.headers.items():
            print(i, j)
    if response.text:
        print('resposne.text:\n', response.text)
        # for i, j in response.text.json():
        #     print(i, j)

def logger_response(response):
    logger.info(response)
    logger.info('response.headers: ' + response.headers)
    logger.info('response.text: ' + response.text)

def dump_response_data(response, funcname):
    logger.info('Now at function:' + funcname)
    logger.info(response.text)

def ssh_command(cmd, ip, login, pass_key, type):
    logger.info('In Func: ' + sys._getframe().f_code.co_name)
    logger.info('cmd: ' + cmd)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    if type == 'passwd':
        ssh.connect(hostname=ip, port=22, username=login, password=pass_key)
    elif type == 'pubkey':
        private_key = paramiko.RSAKey.from_private_key_file(pass_key)
        ssh.connect(hostname=ip, port=22, username=login, pkey=private_key)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    res, err = stdout.read(), stderr.read()
    result = res if res else err
    ssh.close()
    return result.decode()

def ssh_scp_put(ip, login, passwd, local_file, remote_file):
    logger.info('In Func: ' + sys._getframe().f_code.co_name)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.put(local_file, remote_file)
    ssh.close()

def ssh_scp_get(ip, login, passwd, remote_file, local_file):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.get(remote_file, local_file)
    ssh.close()

def get_token():
    # logger.info('In Func:' + sys._getframe().f_code.co_name)
    response = None
    headers = {'Accept': '*/*', 'Content-Type': 'application/x-www-form-urlencoded'}
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret
    }
    s = requests.Session()
    s.mount('http://', HTTPAdapter(max_retries=10))
    s.mount('https://', HTTPAdapter(max_retries=10))
    address = cbam_url + '/auth/realms/cbam/protocol/openid-connect/token'
    try:
        response = s.post(address, data=data, headers=headers, timeout=240, verify=False)
    except requests.exceptions.RequestException as e:
        print(e)
    # check response code
    if response:
        if response.status_code != 200:
            logger.info('Func get_token failed. Response status_code: ' + str(response.status_code))
            logger.info('Response text: ' + response.text)
            exit(1)
        return 'bearer ' + response.json()["access_token"]
    else:
        logger.info('Func get_token failed. Response: None.')
        return None

requests_types = ['GET', 'POST', 'DELETE', 'PATCH']

def send_request(type='GET', url='', data=None, expect_code=200, timeout=180, retries=10, func=''):
    logger.info('In Func:' + sys._getframe().f_code.co_name)
    response = None
    # check params
    if data is None:
        data = {}
    if type not in requests_types:
        logger.info('Unsupported request type.')
        exit(1)
    logger.info('type: ' + type)
    logger.info('url: ' + url)
    logger.info('data: ' + str(data))
    logger.info('expect_code: ' + str(expect_code))
    # form request then send the request, get response
    token_bear = get_token()
    s = requests.Session()
    s.mount('http://', HTTPAdapter(max_retries=retries))
    s.mount('https://', HTTPAdapter(max_retries=retries))
    try:
        if type == 'GET':
            headers = {"Authorization": token_bear}
            response = s.get(url, headers=headers, timeout=timeout, verify=False)
        if type == 'POST':
            headers = {"Authorization": token_bear, "Content-Type": "application/json"}
            response = s.post(url, headers=headers, json=data, timeout=timeout, verify=False)
        if type == 'DELETE':
            headers = {"Authorization": token_bear}
            response = s.delete(url, headers=headers, timeout=timeout, verify=False)
        if type == 'PATCH':
            headers = {"Authorization": token_bear, "Content-Type": "application/json"}
            response = s.patch(url, headers=headers, json=data, timeout=timeout, verify=False)
    except requests.exceptions.RequestException as e:
        print(e)
    # check response code
    if response:
        if response.status_code != expect_code:
            logger.info('Response failed at: ' + func)
            logger.info('Response status_code: ' + str(response.status_code))
            logger.info('Response text: ' + response.text)
            exit(1)
        return response
    else:
        logger.info('Func send_request failed. Response: None.')
        return None

# This is to get vnf list
def get_vnfs():
    token_bear = get_token()
    headers = {'Authorization': token_bear}
    address = cbam_url + '/vnflcm/v1' + '/vnf_instances'
    response = requests.get(address, headers=headers, verify=False)
    data = json.loads(response.text)
    for vnf in data:
        print('\n Now is: ', vnf['name'])
        for i, j in vnf.items():
            print(i, ':', j)

def get_vnf_packages():
    token_bear = get_token()
    headers = {'Authorization': token_bear}
    address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages'
    response = requests.get(address, headers=headers, verify=False)
    dump_response_data(response, 'get_vnf_packages')

class SetUps(object):
    pass

class CleanUps(object):
    pass

########################################################################################################################
# VNF Packages
vnfpkg_name_baseload    = r'\Nokia_sig_SBC-VNF_Package.zip'
vnfpkg_name_drbaseload  = r'\Nokia_sig_SBC-VNF_Package-DR.zip'
vnfpkg_name_sutoload    = r'\Nokia_sig_SBC-VNF_Package-SUToLoad.zip'
vnfpkg_name_cssutoload  = r'\Nokia_sig_SBC-VNF_Package-CSSUToLoad.zip'
vnfpkg_supported_type   = ['instantiation', 'su', 'dr', 'cssu']

class VnfPkgs(object):
    def __init__(self):
        self.vnfpkg_baseload    = sig_data_dir + vnfpkg_name_baseload
        self.vnfpkg_drbaseload  = sig_data_dir + vnfpkg_name_drbaseload
        self.vnfpkg_sutoload    = sig_data_dir + vnfpkg_name_sutoload
        self.vnfpkg_cssutoload  = sig_data_dir + vnfpkg_name_cssutoload
        self.vnfpkgs_list = []

    def get_vnfpkgs(self):
        self.vnfpkgs_list = []
        token_bear = get_token()
        headers = {'Authorization': token_bear}
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages'
        response = requests.get(address, headers=headers, verify=False)
        data = json.loads(response.text)
        if len(data) != 0:
            for vp in data:
                vnfpkg = VnfPkg()
                vnfpkg.id = vp['id']
                vnfpkg.vnfdId = vp['vnfdId']
                vnfpkg.vnfProvider = vp['vnfProvider']
                vnfpkg.vnfProductName = vp['vnfProductName']
                vnfpkg.vnfSoftwareVersion = vp['vnfSoftwareVersion']
                vnfpkg.vnfdVersion = vp['vnfdVersion']
                vnfpkg.onboardingState = vp['onboardingState']
                vnfpkg.operationalState = vp['operationalState']
                vnfpkg.userDefinedData = vp['userDefinedData']
                vnfpkg._links = vp['_links']
                self.vnfpkgs_list.append(vnfpkg)

    def dump_vnfpkgs(self):
        if len(self.vnfpkgs_list) != 0:
            for vp in self.vnfpkgs_list:
                vp.dump_vnfpkg()

    def print_vnfpkgs(self):
        if len(self.vnfpkgs_list) != 0:
            for vp in self.vnfpkgs_list:
                vp.print_vnfpkg()

    # Supported type: ['instantiation', 'su', 'dr', 'cssu']
    def upload_vnfpkg(self, version = sigVersion, type='instantiation'):
        vnfpkg = ''
        id = ''
        if type not in vnfpkg_supported_type:
            logger.info('vnfpkg type not supported. Need to be one of ' +
                        str(vnfpkg_supported_type))
        if version == sigVersion:
            if type == 'instantiation':
                vnfpkg = self.vnfpkg_baseload
            elif type == 'dr':
                vnfpkg = self.vnfpkg_drbaseload
        elif version == sigVersion_SU:
            if type == 'su':
                vnfpkg = self.vnfpkg_sutoload
            elif type == 'cssu':
                vnfpkg = self.vnfpkg_cssutoload
        logger.info('vnfpkg: ' + vnfpkg)
        # To-do: add check if the vnfpkg has already been uploaded
        id = self.get_vnfpkg_by_swversion(version)
        if id:
            # delete the vnf pkg
            self.delete_vnfpkg_by_id(id)
        # Now upload the vnd pkg
        token_bear = get_token()
        # First POST vnf_packages to generate one new vnf pkg
        headers = {'Authorization': token_bear}
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages'
        response = requests.post(address, headers=headers, verify=False)
        if response.status_code != 201:
            logger.info('vnfpkg creation failed.')
            exit(1)
        data = json.loads(response.text)
        id = data['id']
        logger.info('New vnfpkg id: ' + id)
        logger.info('New vnfpkg info: ' + response.text)
        #Then PUT content to content of new vnf pkg
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages/' + id + '/package_content'
        headers = {'Authorization': token_bear, 'Content-Type': 'application/octet-stream'}
        response = requests.put(address, data=open(vnfpkg, 'rb'), headers=headers, verify=False)
        if response.status_code != 202:
            logger.info('vnfpkg content upload failed.')
            # Here need to delete the newly created vnf pkg
            headers = {'Authorization': token_bear}
            address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages/' + id
            requests.delete(address, headers=headers, verify=False)
            exit(1)
        logger.info('VNF PKG uploaded successfully.')
        time.sleep(10)

    def get_vnfpkg_by_swversion(self, swversion):
        self.get_vnfpkgs()
        for vp in self.vnfpkgs_list:
            if vp.vnfSoftwareVersion.endswith(swversion):
                logger.info('id in get_vnfpkg_by_swversion: ' + vp.id)
                return vp.id

    def delete_vnfpkg_by_swversion(self, swversion):
        id = self.get_vnfpkg_by_swversion(swversion)
        self.delete_vnfpkg_by_id(id)

    def delete_vnfpkg_by_id(self, id):
        token_bear = get_token()
        headers = {'Authorization': token_bear}
        address = cbam_url + '/vnfpkgm/v1' + '/vnf_packages/' + id
        response = requests.get(address, headers=headers, verify=False)
        # data = json.loads(response.text)
        logger.info('The vnfpkg to be deleted: ' + response.text)
        requests.delete(address, headers=headers, verify=False)

# VNF Package
class VnfPkg(object):
    def __init__(self):
        self.id = ''
        self.vnfdId = ''
        self.vnfProvider = ''
        self.vnfProductName = ''
        self.vnfSoftwareVersion = ''
        self.vnfdVersion = ''
        self.onboardingState = ''
        self.operationalState = ''
        self.userDefinedData = ''
        self._links = ''

    def get_id(self):
        return self.id

    def get_vnfdId(self):
        return self.vnfdId

    def get_vnfProductName(self):
        return self.vnfProductName

    def get_vnfSoftwareVersion(self):
        return self.vnfSoftwareVersion

    def get_vnfdVersion(self):
        return self.vnfdVersion

    def dump_vnfpkg(self):
        logger.info('id: ' + self.id)
        logger.info('vnfdId: ' + self.vnfdId)
        logger.info('vnfProvider: ' + self.vnfProvider)
        logger.info('vnfProductName: ' + self.vnfProductName)
        logger.info('vnfSoftwareVersion: ' + self.vnfSoftwareVersion)
        logger.info('vnfdVersion: ' + self.vnfdVersion)
        logger.info('onboardingState: ' + self.onboardingState)
        logger.info('operationalState: ' + self.operationalState)
        logger.info('userDefinedData: ' + str(self.userDefinedData))
        logger.info('_links: ' + str(self._links))

    def print_vnfpkg(self):
        print('id: ', self.id)
        print('vnfdId: ', self.vnfdId)
        print('vnfProvider: ', self.vnfProvider)
        print('vnfProductName: ', self.vnfProductName)
        print('vnfSoftwareVersion: ', self.vnfSoftwareVersion)
        print('vnfdVersion: ', self.vnfdVersion)
        print('onboardingState: ', self.onboardingState)
        print('operationalState: ', self.operationalState)
        print('userDefinedData: ', self.userDefinedData)
        print('_links: ', self._links)
        print('\n')

########################################################################################################################
# Sig VNF Package Operations
def sigvnf_UploadVnfpkg(swVersion, type):
    vnfpkgs = VnfPkgs()
    # vnfpkgs.get_vnfpkgs()
    vnfpkgs.upload_vnfpkg(version=swVersion, type=type)
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()

def sigvnf_DeleteVnfpkg(swVersion):
    vnfpkgs = VnfPkgs()
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()
    vnfpkgs.delete_vnfpkg_by_swversion(swVersion)

########################################################################################################################
# Signaling plane VNF
sigvnf_instantiate_file = sig_data_dir + r'\LCM_instantiate_params.json'
sigvnf_cssu_instantiate_file = sig_data_dir + r'\cssu_LCM_instantiate_params.json'
sigvnf_dr_instantiate_file = sig_data_dir + r'\dr_LCM_instantiate_params.json'

class SigVnf(object):
    def __init__(self, vnfdId, name='sbclcm-sig-plane',
                 description = 'SBC LCM Sig-Plane VNF for Auto Test via CBAM REST APIs'):
        self.name = name
        self.description = description
        self.vnfdId = vnfdId
        self.id = ''
        self.opStatus = ''
        self.operationState = ''
        self.scaleLevel = 1
        self.maxScaleLevel = 5
        self.vnfcInstanceId_List = []
        self.instantiate_file = sigvnf_instantiate_file
        self.cssu_instantiate_file = sigvnf_cssu_instantiate_file
        self.dr_instantiate_file = sigvnf_dr_instantiate_file
        self.vnflcm_base_path = cbam_url + '/vnflcm/v1'
        self.vnflcm_instances = self.vnflcm_base_path + '/vnf_instances/'
        self.vnfpkgm_base_path = cbam_url + '/vnfpkgm/v1'
        self.vnflcm_create = vnflcm_base_path + '/vnf_instances'

    def set_vnflcm_apis(self):
        self.operationState             = vnflcm_base_path + '/vnf_lcm_op_occs'
        self.vnflcm_create              = vnflcm_base_path + '/vnf_instances'
        self.vnflcm_instantiate         = self.vnflcm_instances + self.id + '/instantiate'
        self.vnflcm_terminate           = self.vnflcm_instances + self.id + '/terminate'
        self.vnflcm_delete              = self.vnflcm_instances + self.id
        self.vnflcm_get                 = self.vnflcm_instances + self.id
        self.vnflcm_modify              = self.vnflcm_instances + self.id
        self.vnflcm_scale               = self.vnflcm_instances + self.id + '/scale'
        self.vnflcm_heal                = self.vnflcm_instances + self.id + '/heal'
        self.vnflcm_upgrade             = self.vnflcm_instances + self.id + '/upgrade'
        self.vnflcm_custom_backup       = self.vnflcm_instances + self.id + '/custom/backup'
        self.vnflcm_custom_connect      = self.vnflcm_instances + self.id + '/custom/connect_SBC_Media_VNF'
        self.vnflcm_custom_disconnect   = self.vnflcm_instances + self.id + '/custom/disconnect_SBC_Media_VNF'
        self.vnflcm_custom_dbrestore    = self.vnflcm_instances + self.id + '/custom/DB_Restore'
        self.vnflcm_custom_upgrade_precheck     = self.vnflcm_instances + self.id + '/custom/upgrade_precheck'
        self.vnflcm_custom_upgrade_1_apply      = self.vnflcm_instances + self.id + '/custom/upgrade_1_apply'
        self.vnflcm_custom_upgrade_2_activate   = self.vnflcm_instances + self.id + '/custom/upgrade_2_activate'
        self.vnflcm_custom_upgrade_3_commit     = self.vnflcm_instances + self.id + '/custom/upgrade_3_commit'
        self.vnflcm_custom_upgrade_archive      = self.vnflcm_instances + self.id + '/custom/upgrade_archive'

    def write_status(self):
        pass

    def get_vnfdId(self):
        return self.vnfdId

    def get_id(self):
        return self.id

    def set_id(self, id):
        self.id = id

    def create_vnf(self):
        data = {
            "vnfdId": self.vnfdId,
            "vnfInstanceName": self.name,
            "vnfInstanceDescription": self.description
        }
        response = send_request(type='POST', url=self.vnflcm_create, data=data,
                                expect_code=201, func=sys._getframe().f_code.co_name)
        # data = json.loads(response.text)
        dump_response_data(response, 'create_vnf')
        # get vnf ID
        data = json.loads(response.text)
        self.id = data['id']
        logger.info('create_vnf:id: '+ self.get_id())
        # print('create_vnf:id: ', self.get_id())
        # Now it's time to set rest api interfaces since we have got id
        self.set_vnflcm_apis()

    #MODIFY use PATCH
    def modify_auto_backup_vnf(self, periodInSeconds=86220, enabled=False):
        data = {
            "vnfConfigurableProperties": {
                "operation_triggers": {
                    "auto_backup": {
                        "periodInSeconds": periodInSeconds,
                        "enabled": enabled
                    }
                }
            }
        }
        response = send_request(type='PATCH', url=self.vnflcm_modify, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def modify_auto_scale_vnf(self, periodInSeconds=600, enabled=False):
        data = {
            "vnfConfigurableProperties": {
                "operation_triggers": {
                    "auto_scale": {
                        "periodInSeconds": periodInSeconds,
                        "enabled": enabled
                    }
                }
            }
        }
        response = send_request(type='PATCH', url=self.vnflcm_modify, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def instantiate_vnf(self):
        data = json.load(open(self.instantiate_file, "rb"))
        response = send_request(type='POST', url=self.vnflcm_instantiate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def cssu_instantiate_vnf(self):
        data = json.load(open(self.cssu_instantiate_file, "rb"))
        response = send_request(type='POST', url=self.vnflcm_instantiate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def dr_instantiate_vnf(self):
        data = json.load(open(self.dr_instantiate_file, "rb"))
        response = send_request(type='POST', url=self.vnflcm_instantiate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def terminate_vnf(self):
        data = {
            "terminationType": "FORCEFUL"
        }
        response = send_request(type='POST', url=self.vnflcm_terminate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def delete_vnf(self):
        response = send_request(type='DELETE', url=self.vnflcm_delete, data=None,
                                expect_code=204, func=sys._getframe().f_code.co_name)


    def upgrade_vnf(self, vnfdId):
        data = {
            "vnfdId": vnfdId,
            "apiVersion": "4.0"
        }
        response = send_request(type='POST', url=self.vnflcm_upgrade, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def get_vnf(self):
        response = send_request(type='GET', url=self.vnflcm_get, data=None,
                                expect_code=200, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        return data

    def get_vnfcInstanceId_List(self):
        data = self.get_vnf()
        for vnfcinfo in data['instantiatedVnfInfo']['vnfcResourceInfo']:
            self.vnfcInstanceId_List.append(vnfcinfo['id'])
        logger.info('vnfcInstanceId_List: '+ str(self.vnfcInstanceId_List))

    def backup_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_backup, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def connect_sbc_media_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_connect, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def disconnect_sbc_media_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_disconnect, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def dbrestore_vnf(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_dbrestore, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def heal_vnf(self, vnfcInstanceId=None, cause='Sig Plane VM Heal: OAM.NOKIA-LCP-VMA'):
        if vnfcInstanceId is None:
            vnfcInstanceId = ['OAM.NOKIA-LCP-VMA']
        data = {
            "cause": cause,
            "vnfcInstanceId": vnfcInstanceId,
            "additionalParams": {"monitorRetries": 90, "monitorDelay": 20}
        }
        response = send_request(type='POST', url=self.vnflcm_heal, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def scale_vnf(self, scaleType='SCALE_OUT', step=1, aspectId="sc_Aspect"):
        data = {
            "type": scaleType,
            "aspectId": aspectId,
            "numberOfSteps": step
        }
        response = send_request(type='POST', url=self.vnflcm_scale, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def get_scaleStatus(self):
        data = self.get_vnf()
        scaleStatus = data['instantiatedVnfInfo']['scaleStatus']
        logger.info('scaleStatus: ' + str(scaleStatus))
        for ss in scaleStatus:
            if ss['aspectId'] == 'sc_Aspect':
                self.scaleLevel=ss['scaleLevel']
                self.maxScaleLevel=ss['maxScaleLevel']
                logger.info('scaleLevel: ' + str(self.scaleLevel))
                logger.info('maxScaleLevel: ' + str(self.maxScaleLevel))

    def upgrade_precheck(self):
        data = {}
        # Need to carry empty data in requests.post, otherwise will fail
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_precheck, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_1_apply(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_1_apply, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_2_activate(self):
        data = {}
        # Need to carry empty data in requests.post, otherwise will fail
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_2_activate, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_3_commit(self):
        data = {}
        # Need to carry empty data in requests.post, otherwise will fail
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_3_commit, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def upgrade_archive(self, additionalParam):
        data = {
            "additionalParams": additionalParam
        }
        response = send_request(type='POST', url=self.vnflcm_custom_upgrade_archive, data=data,
                                expect_code=202, func=sys._getframe().f_code.co_name)

    def get_latestOperation(self):
        response = send_request(type='GET', url=self.operationState, data=None,
                                expect_code=200, timeout=10, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        return data[0]

    # def get_latestOperationState(self):
    #     response = send_request(type='GET', url=self.operationState, data=None,
    #                             expect_code=200, timeout=10, func=sys._getframe().f_code.co_name)
    #     data = json.loads(response.text)
    #     return data[0]['operationState']

    def get_latestOperationState(self):
        response = send_request(type='GET', url=self.operationState, data=None,
                                expect_code=200, timeout=10, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        for item in data:
            if item['vnfInstanceId'] == self.id:
                return item['operationState']
        return None

    # operation: 'MODIFY_INFO', 'INSTANTIATE', 'OTHER',
    def get_operationState(self, operation, operationName):
        self.opStatus = ''
        response = send_request(type='GET', url=self.operationState, data=None,
                                expect_code=200, timeout=60, func=sys._getframe().f_code.co_name)
        data = json.loads(response.text)
        for item in data:
            # Need to find a way to use operationName
            if (item['vnfInstanceId'] == self.id) and (item['operation'] == operation):
                self.opStatus = item['operationState']
                break
        return self.opStatus

# Media plane VNF
class MediaVnf(object):
    def __init__(self, vnfdId):
        self.vnfdId = vnfdId
        pass

########################################################################################################################
# Media plane VNF LCM Tests
class MediaVnfLcmTestDriver(object):
    def __init__(self, mediavnf):
        self.mediavnf = mediavnf
        pass

backup_server_creds = '-----BEGIN RSA PRIVATE KEY-----\nMIIEowIBAAKCAQEAvX/4YeeQcBqR2sjW18T8WbiAVdnVIs4ZNNlTEG+Ps6OFeQFm\nUBwXvtqbetitEhAAU2OT4ivuh+B20KM/WHM6r57URA1qgNK8Sk9tLUiZRvMDkDR8\nWpAEQSw0fvkO91J93q9Siu3h2uhiKtB1ARESb5DcayttXg0fr9U5hmT/mD5MJ/nC\npfE5ByuHzmWSJo9Vya+YM0UBZnja38vhfc8mAeJ0pxQVxkoL4KzYF2JJAqIzH9tw\nkJLfZbNlxQtOKRCSRWtdcCHckfS0/BISf9f0jdG9h15q3vvKc0j+dXCJny5jfgCc\nHl2+e4RYK3evvJPWQNSZ/+iMZqn1SyligpbKXQIDAQABAoIBABtfsAach73Z6LXd\nC0Px/a4MO+Wq6OH1Oajrt9cI9o4xkedP73KlDD0SoSEWybFxRErHeKZUSEmygBdV\nbaIeSxzxaaJG+dqQFoj5fkDrWtDn69zZ6BjA8wxjEVZCLgpGDU6srtTI1jZkGUIs\nCKrVx378QwrsJAlRBgHFYGDsmAtqvLetbMm272K/hz2e4jxMf08LcDx1fUgDO+hD\naMjzVmbIYNbFftLQD/yG+DgtvP7f728yoG8gNXC0shPqetHtfnwCbJ3H1ju9svGW\nDFYzrBwZZZMY0nTAWjY/DXSS1ORrq7LXSIPKf/3kobfMU67ugeC6NrEsncdIKvIE\n/yVKTrUCgYEA4iMKFeRdIWjzijg3Iie6k09vduTH+nsELDKjzHqcyPmEBisidFQA\nFMjJO8+jh+KKusqYEDp1l4qHLbtN1hatKZV6PwfcmPITA+uczZa9Z3YAKsI1GAOi\n0GUmjmoElOGc1I+k55fWAt48hWRoPd4vrlGaAw8otIdiSXHdpwdb3rsCgYEA1oZa\nH7R4j757Ny47NHIPsXsfZwyD2TTU+oOIDOPvBWN6GuU70N5ZW21529mwLKbr5e2H\n9ofMhSJQkXlg6e6F5wmj7Q0n+mGNoPeNoq7bRB1gHelODs4vAnnVNz2vCGt0uqSd\n+UoemidWKm8RFNAG8yS9uoQrk7sZpA5MYHrJBccCgYBELqxrzV8HI83Kbwiwk6n9\noIXLI0/ohg7MBLi+fnmnXxQfiAHrcShVG/UQw5pa7kNF7q/KtNWfy3TWpRLi6hNr\n5lXli0lIFDUHiZLNqhWRjFKgkc3QX8hHbTgi2HRpL11J+cWOzokIdFlrHssPXF6k\nAJafNYLga7GG034xTla04QKBgAK5Meu1HtK0WFwa+iVwTUKzjXKBdisLwKhtgwym\n2CH5YVN2FYxRRlEi0qk32kS22cfRfChlEPOfu+Yc5F4T6R9FwA8CW7+R/XpNqj6m\neaIjvVSj4ZnOhEpDwbEx10cEFjdIX7kKd9j9JtrjDhR1j6EGlmIHy4XUmj66771J\n0cOBAoGBAJyXMkKKbsSGDGYoyzazVjLYzIqmC2tOnYkHHl3heEZN/LkphPuLdbcj\n2pk2wlqskZ2LQxnylJDgIIA6rJBoLvBj5Xo+9EN7uidHHv+8JGBy4FPcDNl2O5RR\nLbiaAVSQNQNm+hku9e0XH0+YFrCP+0Q8D9DGYzhupslAzJEyoz3R\n-----END RSA PRIVATE KEY-----'

# Signaling plane VNF LCM Tests
class SigVnfLcmTestDriver(object):
    def __init__(self, sigvnf):
        # Need to determine which is sig plane vnf
        # Need to get all vnfs and decide
        self.getvnfs = cbam_url + '/vnflcm/v1' + '/vnf_instances/'
        self.status_file = working_dir + '.sig_status'
        self.sigvnf = sigvnf
        self.sig_id = ''
        self.sig_name = 'sbclcm03'
        self.sig_oama_ip = '10.75.44.23'
        self.sig_oam_login = 'root'
        self.sig_oam_passwd = 'newsys'
        self.fixed_scm_ip = '10.75.44.10,10.75.44.11'
        self.restore_media_plane = 'ALL'
        # centos is for pubkey login
        self.backup_server_login_pubkey = 'centos'
        self.local_private_key = r'C:\Users\shawnx\.ssh\id_rsa'
        # root is for passwd login
        self.backup_server_login_passwd = 'root'
        self.backup_server_passwd_passwd = 'newsys'
        self.backup_server_ip = '10.75.44.7'
        self.backup_server_dir = '/var/www/html/sbclcm-auto/'
        self.cssu_zip = 'cssu_archive.zip'
        self.backup_zip = 'backup.zip'
        self.backup_server1 = 'centos@10.75.44.7:/var/www/html/sbclcm-auto/'
        self.backup_server2 = 'centos@10.75.44.7:/var/www/html/sbclcm-auto/'
        self.backup_server_creds1 = backup_server_creds
        self.backup_server_creds2 = self.backup_server_creds1
        self.su_deft_url = 'http://10.75.44.7/deft_R37.28.XX_R37.28.XX.zip'
        self.su_deft_key = '13582'
        self.su_to_image = 'nokia-SBC_sig-RHEL7-R37.28.06.0020.x86_64.qcow2'
        self.su_to_version = 'R37.28.06.0020'
        # Following are additionalParams for various operations
        self.additinalParams_Backup_Local = {
            'backupServer1': '',
            'backup_server_credentials1': '',
            'backupServer2': '',
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_1 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': '',
            'backupServer2': '',
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_2 = {
            'backupServer1': '',
            'backup_server_credentials1': '',
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_12 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': '',
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_Creds_1 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': self.backup_server_creds1,
            'backupServer2': '',
            'backup_server_credentials2': ''
        }
        self.additinalParams_Backup_Remote_Creds_2 = {
            'backupServer1': '',
            'backup_server_credentials1': '',
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': self.backup_server_creds2
        }
        self.additinalParams_Backup_Remote_Creds_12 = {
            'backupServer1': self.backup_server1,
            'backup_server_credentials1': self.backup_server_creds1,
            'backupServer2': self.backup_server2,
            'backup_server_credentials2': self.backup_server_creds2
        }
        self.additinalParams_Connection = {
            'FixedScmIpAddress': self.fixed_scm_ip,
            'LcmUser': 'cloud-user',
            'ApplUser': 'diag',
            'ApplUserPw': '-assured'
        }
        self.additinalParams_Disconnection = {
            'FixedScmIpAddress': self.fixed_scm_ip,
            'LcmUser': 'cloud-user'
        }
        self.additinalParams_DBRestore = {
            'backupInfo': self.backup_server1 + self.backup_zip,
            'restore_media_plane': self.restore_media_plane
        }
        self.additinalParams_Upgrade1Apply = {
            'upgradeImageName': self.su_to_image,
            'upgradeVersion': self.su_to_version,
            'upgradeDeftKey': self.su_deft_key,
            'upgradeDeftUrl': self.su_deft_url
        }
        self.additinalParams_UpgradeArchive = {
            'upgradeDeftKey': self.su_deft_key,
            'upgradeDeftUrl': self.su_deft_url,
            'upgradeVersion': self.su_to_version,
            'upgradeServer': self.backup_server1,
            'upgradeServerCredentials': ''
        }
        self.additinalParams_UpgradeArchive_Creds = {
            'upgradeDeftKey': self.su_deft_key,
            'upgradeDeftUrl': self.su_deft_url,
            'upgradeVersion': self.su_to_version,
            'upgradeServer': self.backup_server1,
            'upgradeServerCredentials': self.backup_server_creds1
        }

    def set_vnf(self):
        # This is to set vnf info for existing VNF
        pass

    def restore_vnf(self):
        # This is to restore info of sig VNF
        self.sig_id = self.get_vnf_id()
        self.sigvnf.set_id(self.sig_id)
        self.sigvnf.set_vnflcm_apis()

    def get_vnf_id(self):
        token_bear = get_token()
        headers = {'Authorization': token_bear}
        response = requests.get(self.getvnfs, headers=headers, verify=False)
        if response.status_code == 200:
            data = json.loads(response.text)
            for vnf in data:
                # Fow now, use vnfProductName to determine if this is sig plane vnf
                if vnf['vnfProductName'] == 'SBC':
                    # print('get_vnfs: sig vnf id: ', vnf['id'])
                    logger.info('get_vnf_id: sig vnf id: ' + vnf['id'])
                    return vnf['id']
        else:
            logger.info('Failed to get_vnf_id')
            exit(1)

    def get_vnf_info(self):
        # For now, assume there is only one sig plane VNF
        pass

    def create_instantiate(self):
        self.sigvnf.create_vnf()
        self.sigvnf.instantiate_vnf()
        self.sigvnf.get_vnf()

    def terminate(self):
        self.sigvnf.terminate_vnf()

    def delete(self):
        self.sigvnf.delete_vnf()

    def change_package_version(self, vnfdId):
        self.sigvnf.upgrade_vnf(vnfdId)

    def get_vnfcInstanceId_List_4Heal(self):
        self.sigvnf.get_vnfcInstanceId_List()
        return self.sigvnf.vnfcInstanceId_List

    def heal(self, vnfcInstanceId=None, cause='Sig Plane VM Heal: OAM.NOKIA-LCP-VMA'):
        if vnfcInstanceId is None:
            vnfcInstanceId = ['OAM.NOKIA-LCP-VMA']
        self.sigvnf.heal_vnf(vnfcInstanceId, cause)

    def get_scale_status(self):
        self.sigvnf.get_scaleStatus()
        return self.sigvnf.scaleLevel, self.sigvnf.maxScaleLevel

    def scale(self, scaleType='SCALE_OUT', step=1, aspectId="sc_Aspect"):
        self.sigvnf.get_scaleStatus()
        scaleLevel = self.sigvnf.scaleLevel
        maxScaleLevel = self.sigvnf.maxScaleLevel
        if scaleType == 'SCALE_OUT':
            if scaleLevel == maxScaleLevel:
                logger.info('Sig Plane maxScaleLevel has been reached, no action.')
                return
            elif scaleLevel < maxScaleLevel:
                self.sigvnf.scale_vnf(scaleType='SCALE_OUT')
        elif scaleType == 'SCALE_IN':
            if scaleLevel == 1:
                logger.info('Sig Plane min SC number has been reached, no action.')
                return
            elif scaleLevel <= maxScaleLevel:
                self.sigvnf.scale_vnf(scaleType='SCALE_IN')

    def modify_auto_scale(self, periodInSeconds=600, enabled=False):
        self.sigvnf.modify_auto_scale_vnf(periodInSeconds, enabled)

    def modify_auto_backup(self, periodInSeconds=86220, enabled=False):
        self.sigvnf.modify_auto_backup_vnf(periodInSeconds, enabled)

    def custom_backup(self, additionalParam):
        self.sigvnf.backup_vnf(additionalParam)

    def custom_connect_SBC_Media_VNF(self, additionalParam):
        self.sigvnf.connect_sbc_media_vnf(additionalParam)

    def custom_disconnect_SBC_Media_VNF(self, additionalParam):
        self.sigvnf.disconnect_sbc_media_vnf(additionalParam)

    def custom_dbrestore(self, additionalParam):
        self.sigvnf.dbrestore_vnf(additionalParam)

    def custom_upgrade_precheck(self):
        self.sigvnf.upgrade_precheck()

    def custom_upgrade_1_apply(self, additionalParam):
        self.sigvnf.upgrade_1_apply(additionalParam)

    def custom_upgrade_2_activate(self):
        self.sigvnf.upgrade_2_activate()

    def custom_upgrade_3_commit(self):
        self.sigvnf.upgrade_3_commit()

    def custom_upgrade_archive(self, additionalParam):
        self.sigvnf.upgrade_archive(additionalParam)

    def cssu_create_instantiate(self):
        self.sigvnf.create_vnf()
        self.sigvnf.cssu_instantiate_vnf()
        self.sigvnf.get_vnf()

    def dr_create_instantiate(self):
        self.sigvnf.create_vnf()
        self.sigvnf.dr_instantiate_vnf()
        self.sigvnf.get_vnf()

    # Need to setup pubkey authentication after instantiation
    def prep_bkserver_pubkey(self):
        pubkey = ''
        # Delete the pubkey in backup server authorized_keys
        # Example: sed - i '/sbclcm03/d' /home/centos/.ssh/authorized_keys
        sshtype = 'passwd'
        cmd = "sed -i " + "'/" + self.sig_name + "/d' " + "/home/centos/.ssh/authorized_keys"
        ssh_command(cmd, self.backup_server_ip, self.backup_server_login_passwd,
                             self.backup_server_passwd_passwd, sshtype)
        # Get oam ssh pubkey of root
        cmd = 'cat /root/.ssh/id_rsa.pub'
        pubkey = ssh_command(cmd, self.sig_oama_ip, self.sig_oam_login, self.sig_oam_passwd, sshtype)
        os.chdir(working_dir)
        with open('tmp_key', 'w') as f:
            f.write(pubkey)
        # Ship the oam ssh pubkey to /tmp on backup server
        ssh_scp_put(self.backup_server_ip, self.backup_server_login_passwd,
                    self.backup_server_passwd_passwd, 'tmp_key', '/tmp/tmp_key')
        # echo the pubkey to authorized_keys
        cmd = 'cat /tmp/tmp_key >> /home/centos/.ssh/authorized_keys'
        result = ssh_command(cmd, self.backup_server_ip, self.backup_server_login_passwd,
                             self.backup_server_passwd_passwd, sshtype)

    # This is combined version to rename and chmod zip afer backup or cssu
    def prep_backup_cssu_zip(self, sshtype='passwd', ziptype='backup'):
        match_str = ''
        newzip = ''
        ip = self.backup_server_ip
        if ziptype == 'backup':
            match_str = '_LCP'
            newzip = self.backup_zip
        elif ziptype == 'cssu':
            match_str = 'su_archive_R'
            newzip = self.cssu_zip
        else:
            logger.info('Only type backup and cssu supprted for ziptype. Error out.')
            exit(1)
        if sshtype == 'passwd':
            login = self.backup_server_login_passwd
            passwd_key = self.backup_server_passwd_passwd
        elif sshtype == 'pubkey':
            login = self.backup_server_login_pubkey
            passwd_key = self.local_private_key
        else:
            logger.info('Only type passwd and pubkey supprted for sshtype. Error out.')
            exit(1)
        cmd = 'ls ' + self.backup_server_dir
        result = ssh_command(cmd, ip, login, passwd_key, sshtype)
        if match_str in result:
            rlist = result.split()
            for r in rlist:
                if match_str in r:
                    cmd = 'mv ' + self.backup_server_dir + r + ' ' + self.backup_server_dir + newzip
                    logger.info(cmd)
                    ssh_command(cmd, ip, login, passwd_key, sshtype)
                    cmd = 'chmod 777 ' + self.backup_server_dir + newzip
                    logger.info(cmd)
                    ssh_command(cmd, ip, login, passwd_key, sshtype)

    def check_latestOpState(self, timeout=30):
        status = self.sigvnf.get_latestOperationState()
        if status is not None:
            return status

    def wait_processing(self, timeout=30):
        logger.info('wait_processing')
        wait = 0
        while wait < timeout:
            status = self.check_latestOpState()
            if status in ['PROCESSING', 'STARTING']:
                time.sleep(60)
                wait = wait + 1
            else:
                break

    def check_opstatus(self, operation='INSTANTIATE', operationName='', timeout=30):
        wait = 0
        while wait < timeout:
            status = self.sigvnf.get_operationState(operation, operationName)
            if status == 'PROCESSING':
                logger.info('Sig Plane VNF ' + operation + ' ' + operationName + ' PROCESSING')
                logger.info('Wait for 60 secs...')
            elif status == 'FAILED':
                logger.info('Sig Plane VNF ' + operation + ' ' + operationName + ' FAILED')
                # log the detailed error info
                data = self.sigvnf.get_latestOperation()
                logger.info(str(data))
                exit(1)
            elif status == 'ROLLED_BACK ':
                logger.info('Sig Plane VNF ' + operation + ' ' + operationName + ' ROLLED_BACK')
                exit(2)
            elif status == 'COMPLETED':
                logger.info('Sig Plane VNF ' + operation + ' ' + operationName + ' COMPLETED')
                return 0
            else:
                logger.info('Sig Plane VNF ' + operation + ' ' + operationName + ' status: ' + status)
            # Check status every 60 secs
            time.sleep(60)
            wait = wait + 1


########################################################################################################################
# Signaling plane VNF artifact generator
# type : ['instantiation', 'su', 'dr', 'cssu']
# rel : load release to artifacts to be generated, to be passed to yact tool
# server_type: httpserver1, httpserver2, httpserver12, bkupserver1, bkupserver2, bkupserver12
default_rel = 'R20.2'
toload_rel = 'R20.2'
server_type_list =['httpserver1', 'httpserver2', 'httpserver12', 'bkupserver1', 'bkupserver2', 'bkupserver12']
class SigVnfArtsGenerator(object):
    # server_type is required for DR, CSSU
    def __init__(self, type='instantiation', rel=default_rel, server_type=None):
        self.type = type
        self.rel = rel
        self.server_type = server_type
        # self.torel = torel
        self.arts_type = ['instantiation', 'su', 'dr', 'cssu']
        # self.rel = 'R20.0'
        # self.torel = 'R20.0'
        self.yact_server_ip = '135.252.41.216'
        self.yact_user = 'yact-user'
        self.yact_passwd = '123456'
        self.yact_dir = '/home/yact-user/shawnx/20.2/'
        # self.sig_dif_name = 'SBC-signaling_R20.2-sbclcm03.xlsm'
        # self.sig_dr_dif_name = 'SBC-signaling_R20.2-sbclcm03-dr.xlsm'
        # self.sig_su_dif_name = 'SBC-signaling_R20.2-sbclcm03-su.xlsm'
        # self.sig_cssu_dif_name = 'SBC-signaling_R20.2-sbclcm03-cssu.xlsm'
        self.sig_dif_name = 'SBC-signaling_R20.2.xlsm'
        self.sig_dr_dif_name = 'SBC-signaling_R20.2-dr.xlsm'
        self.sig_su_dif_name = 'SBC-signaling_R20.2-su.xlsm'
        self.sig_cssu_dif_name = 'SBC-signaling_R20.2-cssu.xlsm'
        self.sig_dif_dict = {
            'base': self.sig_dif_name,
            'cssu': self.sig_cssu_dif_name,
            'dr': self.sig_dr_dif_name,
            'su': self.sig_su_dif_name
        }
        self.sig_ne_name = 'sbclcm03'
        self.sig_zip = self.sig_ne_name + '.zip'
        self.sig_dif_file = sig_arts_dir + '\\' + self.sig_dif_name
        self.sig_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package.zip'
        self.sig_dr_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-DR.zip'
        self.sig_su_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-SUToLoad.zip'
        self.sig_cssu_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-CSSUToLoad.zip'
        # orig_instantiation_json_file = 'orig_LCM_instantiate_params.json'
        self.instantiation_json_file = 'LCM_instantiate_params.json'
        self.dr_instantiation_json_file = 'dr_LCM_instantiate_params.json'
        self.cssu_instantiation_json_file = 'cssu_LCM_instantiate_params.json'
        self.os_passwd = 'a321sbc'
        self.backup_server_ip = '10.75.44.7'
        self.backup_server_login = 'root'
        self.backup_server_passwd = 'newsys'
        self.backup_server_dir = '/var/www/html/sbclcm-auto/'
        self.backup_file_name = 'backup.zip'
        self.ssh_type = 'passwd'
        self.backup_file = self.backup_server_dir + self.backup_file_name
        self.local_backup_dir = 'backup'
        self.ap_instantiation = {
            "backup_file1": "",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            # "bulk_conf_url": "http://10.75.44.7/sbclcm03/bulkconf_artifacts.zip",
            "bulk_conf_url": "http://10.75.44.7/sbclcm-auto/bulkconf_artifacts.zip",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation = {
            "backup_file1": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_file2": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_http1 = {
            "backup_file1": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_http2 = {
            "backup_file1": "",
            "backup_file2": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_http12 = {
            "backup_file1": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_file2": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_creds1 = {
            "backup_file1": "centos@10.75.44.7:/var/www/html/sbclcm-auto/backup.zip",
            "backup_file2": "",
            "backup_server_credentials1": backup_server_creds,
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_creds2 = {
            "backup_file1": "",
            "backup_file2": "centos@10.75.44.7:/var/www/html/sbclcm-auto/backup.zip",
            "backup_server_credentials1": "",
            "backup_server_credentials2": backup_server_creds,
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation_creds12 = {
            "backup_file1": "centos@10.75.44.7:/var/www/html/sbclcm-auto/backup.zip",
            "backup_file2": "centos@10.75.44.7:/var/www/html/sbclcm-auto/backup.zip",
            "backup_server_credentials1": backup_server_creds,
            "backup_server_credentials2": backup_server_creds,
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_cssu_instantiation = {
            "backup_file1": "",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "http://10.75.44.7/sbclcm-auto/cssu_archive.zip",
            "upgrade_server_credentials": ""
        }

    # Supported type: ['instantiation', 'su', 'dr', 'cssu']
    # rel is needed to handle SU case, SU TO load may be different release
    # def gen_arts_sig(type='instantiation', rel='R20.0'):
    def gen_arts_sig(self):
        logger.info('In Func:' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)

        if self.type not in self.arts_type:
            logger.info('Only ' + str(self.arts_type) + ' Supported.')
            exit(1)

        ts = int(time.time())
        logger.info('Current time is: ' + str(ts))

        sig_yact_dir = self.yact_dir + str(ts) + '/'
        cmd = 'mkdir -p ' + sig_yact_dir
        ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)

        if self.type == 'instantiation':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_dif_name
        elif self.type == 'dr':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_dr_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_dr_dif_name
        elif self.type == 'su':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_su_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_su_dif_name
        elif self.type == 'cssu':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_cssu_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_cssu_dif_name

        logger.info('sig_dif_file: ' + sig_dif_file)
        logger.info('sig_rmt_dif_file: ' + sig_rmt_dif_file)
        ssh_scp_put(self.yact_server_ip, self.yact_user, self.yact_passwd, sig_dif_file, sig_rmt_dif_file)

        yact_cmd = '/home/yact/YACT/yact.sh gen-by-dif ' + sig_rmt_dif_file + ' SBC-signaling ' + self.rel
        logger.info('yact_cmd: ' + yact_cmd)
        cmd = 'cd ' + sig_yact_dir + ';' + yact_cmd
        result = ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)
        logger.info('result of yact_cmd: '+ result)

        # Handle SU case
        if self.type == 'su':
            pkg_tools_dir = sig_yact_dir + self.sig_ne_name + '/package/pkg_tools/'
            logger.info('su pkg_tools_dir: ' + pkg_tools_dir)

            # ship from load vnf pkg to pkg_tools_dir and rename it to vnfpkg-from-load.zip
            vnfpkg_from_load = sig_data_dir + '\\' + self.sig_vnfpkg_name
            rmt_vnfpkg_from_load = pkg_tools_dir + 'vnfpkg-from-load.zip'
            ssh_scp_put(self.yact_server_ip, self.yact_user, self.yact_passwd, vnfpkg_from_load, rmt_vnfpkg_from_load)

            # run vnf_upgrade_pkg_gen
            cmd = 'cd ' + pkg_tools_dir + ';' + './vnf_upgrade_pkg_gen -n ../Nokia_sig_SBC-VNF_Package.zip ' \
                                                '-c ./vnfpkg-from-load.zip ' + self.sig_ne_name
            logger.info('vnf_upgrade_pkg_gen cmd: ' + cmd)
            ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)

            # now ship the upgrade vnf pkg to local
            # example: sbclcm03_Nokia_sig_SBC_upgrade-VNF_Package.zip
            vnfpkg_upgrade = pkg_tools_dir + self.sig_ne_name + '_Nokia_sig_SBC_upgrade-VNF_Package.zip'
            logger.info('upgrade package: ' + vnfpkg_upgrade)
            vnfpkg_upgrade_local = sig_data_dir + '\\' + self.sig_su_vnfpkg_name
            ssh_scp_get(self.yact_server_ip, self.yact_user, self.yact_passwd, vnfpkg_upgrade, vnfpkg_upgrade_local)
        # handle other cases
        else:
            # zip remote sig_art_files
            cmd = 'cd ' + sig_yact_dir + ';' + 'zip -r ' + self.sig_ne_name + '.zip ' + self.sig_ne_name
            result = ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)
            logger.info('result of zip sig: ' + result)

            # remove sig zip if exists
            if os.path.exists(self.sig_zip):
                os.remove(self.sig_zip)

            # remote the unziped dir if exists
            if os.path.exists(self.sig_ne_name):
                shutil.rmtree(self.sig_ne_name)

            rmt_zip_file = sig_yact_dir + self.sig_zip
            ssh_scp_get(self.yact_server_ip, self.yact_user, self.yact_passwd, rmt_zip_file, self.sig_zip)

            logger.info('unzip '+ self.sig_zip)
            zip_file = zipfile.ZipFile(self.sig_zip)
            # os.mkdir(sig_ne_name)
            for names in zip_file.namelist():
                zip_file.extract(names, '.')
            zip_file.close()

    def ship_bulkconf_zip(self):
        # scp the bulkconf_artifacts.zip to remote backup server
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            logger.info('artifacts dir not exist. Exit.')
            exit(1)

        # ship the bulkconf_artifacts.zip
        local_bulkconf_file = sig_arts_dir + '\\' + self.sig_ne_name + r'\bulk_netconf\scale\bulkconf_artifacts.zip'
        rmt_bulkconf_file = self.backup_server_dir + 'bulkconf_artifacts.zip'
        logger.info('local_bulkconf_file: ' + local_bulkconf_file)
        logger.info('rmt_bulkconf_file: ' + rmt_bulkconf_file)
        logger.info('ship bulkconf_artifacts.zip to remote backup server.')
        ssh_scp_put(self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                    local_bulkconf_file, rmt_bulkconf_file)

        # chmod
        cmd = 'cd ' + self.backup_server_dir + '; chmod 777 bulkconf_artifacts.zip'
        result = ssh_command(cmd, self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                             self.ssh_type)
        logger.info('result of chmod bulkconf_artifacts.zip: ' + result)

    # def cp_vnf_pkg(type='instantiation'):
    def cp_vnf_pkg(self):
        # cp the Nokia_sig_SBC-VNF_Package.zip to data\sig-plane
        # to correct name
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            logger.info('artifacts dir not exist. Exit.')
            exit(1)

        srcfile = self.sig_ne_name + r'\package\Nokia_sig_SBC-VNF_Package.zip'

        if self.type == 'instantiation':
            dstfile = sig_data_dir + '\\' + self.sig_vnfpkg_name
        elif self.type == 'dr':
            dstfile = sig_data_dir + '\\' + self.sig_dr_vnfpkg_name
        elif self.type == 'su':
            dstfile = sig_data_dir + '\\' + self.sig_su_vnfpkg_name
        elif self.type == 'cssu':
            dstfile = sig_data_dir + '\\' + self.sig_cssu_vnfpkg_name

        logger.info('Copy ' + srcfile + ' to ' + dstfile)
        shutil.copyfile(srcfile, dstfile)

    # json load, dump
    # with open(file_name, 'r') as f:
    #     data = json.load(f)
    # with open('output.json', 'w') as f:
    #     json.dump(data, f)
    # def prep_instantiation_json(type='instantiation'):
    def prep_instantiation_json(self):
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            logger.info('artifacts dir not exist. Exit.')
            exit(1)

        json_file = self.sig_ne_name + r'\package\cbam_json' + '\\' + self.instantiation_json_file

        # dr and cssu need to populate backup server by server_type
        if self.type in ['dr', 'cssu']:
            if self.server_type not in server_type_list:
                logger.info('server_type not populated for ' + self.type)
                exit(1)

        # server_type_list:['httpserver1', 'httpserver2', 'httpserver12', 'bkupserver1', 'bkupserver2', 'bkupserver12']
        if self.type == 'dr':
            if self.server_type == 'httpserver1':
                self.ap_dr_instantiation = self.ap_dr_instantiation_http1
            if self.server_type == 'httpserver2':
                self.ap_dr_instantiation = self.ap_dr_instantiation_http2
            if self.server_type == 'httpserver12':
                self.ap_dr_instantiation = self.ap_dr_instantiation_http12
            if self.server_type == 'bkupserver1':
                self.ap_dr_instantiation = self.ap_dr_instantiation_creds1
            if self.server_type == 'bkupserver2':
                self.ap_dr_instantiation = self.ap_dr_instantiation_creds2
            if self.server_type == 'bkupserver12':
                self.ap_dr_instantiation = self.ap_dr_instantiation_creds12
            logger.info('server_type : ' + self.server_type)
            logger.info('ap_dr_instantiation : ' + str(self.ap_dr_instantiation))

        # In case of dr, the instantiation json needs to be extracted from backup.zip
        # In case of su, instantiation json is not needed
        if self.type == 'dr':
            if os.path.exists(self.backup_file_name):
                os.remove(self.backup_file_name)

            if os.path.exists(self.local_backup_dir):
                shutil.rmtree(self.local_backup_dir)

            ssh_scp_get(self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                        self.backup_file, self.backup_file_name)

            logger.info('unzip ' + self.backup_file_name)
            os.mkdir(self.local_backup_dir)
            zip_file = zipfile.ZipFile(self.backup_file_name)
            for names in zip_file.namelist():
                zip_file.extract(names, self.local_backup_dir)
            zip_file.close()

            json_file = self.local_backup_dir + '\\' + self.instantiation_json_file

        # load json data
        try:
            with open(json_file, 'r') as jFile:
                data = json.load(jFile)
        except Exception as exc:
            print('ERROR: Failed to load data from {0}.\n[{1}]'.format(
                json_file, str(exc)))

        # change data
        data['vimConnectionInfo'][0]['accessInfo']['password'] = self.os_passwd

        if self.type == 'instantiation':
            data['additionalParams'] = self.ap_instantiation
            output_json = sig_data_dir + '\\' + self.instantiation_json_file
        elif self.type == 'dr':
            data['additionalParams'] = self.ap_dr_instantiation
            output_json = sig_data_dir + '\\' + self.dr_instantiation_json_file
        elif self.type == 'cssu':
            data['additionalParams'] = self.ap_cssu_instantiation
            output_json = sig_data_dir + '\\' + self.cssu_instantiation_json_file

        # dump modified data
        try:
            with open(output_json, 'w') as jFile:
                json.dump(data, jFile, indent=2)
        except Exception as exc:
            print('ERROR: Failed to write data to {0}.\n[{1}]'.format(
                output_json, str(exc)))

        logger.info('instantiation json for ' + self.type + ' created completed.')

    # get the sc count of provided workbook
    # return: the row of sc vm_count, the vm_count value
    def get_sc_count(self, workbook):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        wb = load_workbook(workbook, keep_vba=True)
        sh = wb['DP']
        colB = sh['B']

        row_list = []
        for i in colB:
            if i.value == 'group_tag':
                row_list.append(i.row)

        row_sc = None
        for r in row_list:
            if sh.cell(row=r, column=26).value == 'SC':
                row_sc = r
        if row_sc is not None:
            logger.info('row_sc: ' + str(row_sc))

        row_vm_count = None
        sc_count = None
        for r in range(row_sc, row_sc + 8):
            if sh.cell(row=r, column=2).value == 'vm_count':
                row_vm_count = r
        if row_vm_count is not None:
            logger.info('row_vm_count: ' + str(row_vm_count))

        sc_count = sh.cell(row=row_vm_count, column=26).value
        logger.info('current SC vaule of vm_count: ' + str(sc_count))

        return (row_vm_count, sc_count)

    # set the sc count of provided workbook
    def set_sc_count(self, workbook, sc_count=5):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        row_vm_count, initial_sc_count = self.get_sc_count(workbook)

        wb = load_workbook(workbook, keep_vba=True)
        sh = wb['DP']

        # set vm_count value to updated value
        sh.cell(row=row_vm_count, column=26).value = sc_count

        # wb.save(filename='output.xlsm')
        wb.save(filename=workbook)

        logger.info('Successfully set sc count to: ' + str(sc_count))

    # get initial sc value
    def get_initial_sc_value(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_dif_name):
            logger.info('Original DIF does not exist. Exit.')
            exit(1)

        workbook = self.sig_dif_name

        row_vm_count, initial_sc_count = self.get_sc_count(workbook)

        logger.info('initial_sc_count: ' + str(initial_sc_count))

        return int(initial_sc_count)

    # update DIF using updated sc value
    def gen_dif_updated_sc_value(self, sc_count=5):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)

        if self.type == 'su':
            workbook = self.sig_su_dif_name
        elif self.type == 'cssu':
            workbook = self.sig_cssu_dif_name
        elif self.type == 'dr':
            workbook = self.sig_dr_dif_name

        if not os.path.exists(workbook):
            logger.info('workbook does not exist. Exit.')
            exit(1)

        self.set_sc_count(workbook, sc_count)

    # prep arts for instantiation, dr,
    def prep_arts(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        # su doesn't need json file, only vnf pkg needed
        if self.type == 'su':
            self.gen_arts_sig()
        else:
            self.gen_arts_sig()
            self.prep_instantiation_json()
            self.cp_vnf_pkg()
        # needs to ship bulkconf artifacts if instantiation
        if self.type == 'instantiation':
            self.ship_bulkconf_zip()

def sigvnf_GenArtsInstantiation():
    sigArts = SigVnfArtsGenerator(type='instantiation', rel=default_rel)
    sigArts.prep_arts()

def sigvnf_GenArtsDR():
    sigArts = SigVnfArtsGenerator(type='dr', rel=default_rel, server_type='httpserver12')
    sigArts.prep_arts()

def sigvnf_GenArtsSU():
    sigArts = SigVnfArtsGenerator(type='su', rel=default_rel)
    sigArts.prep_arts()

def sigvnf_GenArtsCSSU():
    sigArts = SigVnfArtsGenerator(type='cssu', rel=default_rel)
    sigArts.prep_arts()

def sigvnf_GenArts():
    sigvnf_GenArtsInstantiation()
    sigvnf_GenArtsDR()
    sigvnf_GenArtsSU()
    sigvnf_GenArtsCSSU()

########################################################################################################################
# This is the upper caller
# It use vnfdId to identify which vnf to operation on
# It's stateless by its best
class LcmTestDriver(object):
    def __init__(self, vnfdId):
        self.vnfdId = vnfdId

    def setup_sigDriver(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.restore_vnf()
        self.sigDriver.wait_processing()

    # In case of vnfdId changed (for example, DR)
    # recreate sigDriver with updated vnfdId
    def refresh_sigDriver(self, vnfdId):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.vnfdId = vnfdId
        # self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=vnfdId))

    def setup_mediaDriver(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.mediaDriver = MediaVnfLcmTestDriver(MediaVnf(vnfdId=self.vnfdId))

    def sigvnf_CreateInstantiate(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        # For CreateInstantiate, cann't use setup_sigDriver as vnf is not created yet
        # use SigVnfLcmTestDriver directly
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.create_instantiate()
        self.sigDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)
        self.sigDriver.prep_bkserver_pubkey()

    def sigvnf_Terminate(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.terminate()
        self.sigDriver.check_opstatus(operation='TERMINATE', operationName='', timeout=10)

    def sigvnf_Delete(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.delete()

    def sigvnf_Modify_DisableAutoScale(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_scale(enabled=False)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_Modify_EnableAutoScale(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_scale(enabled=True)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_Modify_DisableAutoBackup(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_backup(enabled=False)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_Modify_EnableAutoBackup(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.modify_auto_backup(enabled=True)
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)

    def sigvnf_ScaleOut(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.scale(scaleType='SCALE_OUT')
        self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)

    def sigvnf_ScaleIn(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.scale(scaleType='SCALE_IN')
        self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)

    def sigvnf_Scale_OutIn(self):
        # First to scale out SC to max number: 5
        # then scale in SC to initial count: depends on initial deployment
        # The system initial count at: /storage/auto_scale/initial_sys_config.json
        # To-do: add func to parser the file and get initial SC count
        # For now, use 1 as initial count
        initial_count = 1
        max_count = 5
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        sigArts = SigVnfArtsGenerator()
        initial_count = sigArts.get_initial_sc_value()
        current_count, max_count = self.sigDriver.get_scale_status()
        logger.info('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        while current_count < max_count:
            self.sigDriver.scale(scaleType='SCALE_OUT')
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status()
        while current_count > initial_count:
            self.sigDriver.scale(scaleType='SCALE_IN')
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status()

    def sigvnf_ScaleOutToMax(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        initial_count = 1
        max_count = 5
        self.setup_sigDriver()
        sigArts = SigVnfArtsGenerator()
        initial_count = sigArts.get_initial_sc_value()
        current_count, max_count = self.sigDriver.get_scale_status()
        logger.info('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        if current_count == max_count:
            logger.info('current_count == max_count. No Scale-Out needed.')
            return
        while current_count < max_count:
            self.sigDriver.scale(scaleType='SCALE_OUT')
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status()

    def sigvnf_ScaleInToInit(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        initial_count = 1
        max_count = 5
        self.setup_sigDriver()
        sigArts = SigVnfArtsGenerator()
        initial_count = sigArts.get_initial_sc_value()
        current_count, max_count = self.sigDriver.get_scale_status()
        logger.info('initial_count, current_count, max_count: ' +
                    str(initial_count) + ' ' + str(current_count) + ' ' + str(max_count))
        if current_count == initial_count:
            logger.info('current_count == initial_count. No Scale-In needed.')
            return
        while current_count > initial_count:
            self.sigDriver.scale(scaleType='SCALE_IN')
            self.sigDriver.check_opstatus(operation='SCALE', operationName='', timeout=60)
            current_count, max_count = self.sigDriver.get_scale_status()

    def sigvnf_Heal_Single(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        logger.info('vnfcInstanceId_List: ' + str(vnfcInstanceId_List))
        # Single VM heal, for all VMs, one VM each time
        for vnfcid in vnfcInstanceId_List:
            logger.info('Start to Heal: ' + vnfcid)
            self.sigDriver.heal(vnfcInstanceId=[vnfcid], cause='Sig Plane VM Heal: ' + vnfcid)
            self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    def sigvnf_Heal_Single_Random(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        logger.info('vnfcInstanceId_List: ' + str(vnfcInstanceId_List))
        # Single VM heal, randomly select 1 VM
        num = random.randint(0, len(vnfcInstanceId_List) - 1)
        vnfcid = vnfcInstanceId_List[num]
        logger.info('Start to Heal: ' + vnfcid)
        self.sigDriver.heal(vnfcInstanceId=[vnfcid], cause='Sig Plane VM Heal: ' + vnfcid)
        self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    # Randomly heal single VM, mutiple times
    def sigvnf_Heal_Single_Random_MTs(self, times=5):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        logger.info('times: ' + str(times))
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        logger.info('vnfcInstanceId_List: ' + str(vnfcInstanceId_List))
        # Single VM heal, randomly select 1 VM, times times
        step = 0
        while step < times:
            logger.info('the step: ' + str(step+1))
            num = random.randint(0, len(vnfcInstanceId_List) - 1)
            vnfcid = vnfcInstanceId_List[num]
            logger.info('Start to Heal: ' + vnfcid)
            self.sigDriver.heal(vnfcInstanceId=[vnfcid], cause='Sig Plane VM Heal: ' + vnfcid)
            self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)
            step = step + 1

    def sigvnf_Heal_Multiple(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        vnfcInstanceId_List = self.sigDriver.get_vnfcInstanceId_List_4Heal()
        if not vnfcInstanceId_List:
            print('Failed to get vnfcInstanceId_List, Error out.')
            exit(1)
        logger.info('vnfcInstanceId_List:' + str(vnfcInstanceId_List))
        maxHealListLength = int(len(vnfcInstanceId_List) / 2)
        start = 2
        end = maxHealListLength + 1
        healList = []
        for healNum in range(start, end):
            while len(healList) < healNum:
                num = random.randint(0, len(vnfcInstanceId_List) - 1)
                duplicated = False
                vnfc = vnfcInstanceId_List[num]
                if vnfc not in healList:
                    for vnfc_tmp in healList:
                        if vnfc[0:10] == vnfc_tmp[0:10]:
                            # duplicated vnfc, not insert in healList
                            duplicated = True
                            break
                    if not duplicated:
                        healList.append(vnfc)
            logger.info('Multiple Heal: healNum:' + str(healNum) + ';healList:' + str(healList))
            logger.info('Start to Heal: ' + str(healList))
            self.sigDriver.heal(vnfcInstanceId=healList, cause='Sig Plane Multiple VM Heal: ' + str(healList))
            self.sigDriver.check_opstatus(operation='HEAL', operationName='', timeout=60)

    def sigvnf_Backup(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        # setup pubkey before backup
        self.sigDriver.prep_bkserver_pubkey()
        for ap in [
            self.sigDriver.additinalParams_Backup_Local,
            self.sigDriver.additinalParams_Backup_Remote_1,
            self.sigDriver.additinalParams_Backup_Remote_2,
            self.sigDriver.additinalParams_Backup_Remote_12,
            self.sigDriver.additinalParams_Backup_Remote_Creds_1,
            self.sigDriver.additinalParams_Backup_Remote_Creds_2,
            self.sigDriver.additinalParams_Backup_Remote_Creds_12
        ]:
            self.sigDriver.custom_backup(ap)
            self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
            self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Local(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Local)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote1(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_1)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote2(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_2)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote12(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_12)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote_Creds1(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_Creds_1)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote_Creds2(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_Creds_2)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_Backup_Remote_Creds12(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_backup(self.sigDriver.additinalParams_Backup_Remote_Creds_12)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:backup', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='backup')

    def sigvnf_ConnectSbcMediaVnf(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_connect_SBC_Media_VNF(self.sigDriver.additinalParams_Connection)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:connect_SBC_Media_VNF', timeout=20)

    def sigvnf_DisconnectSbcMediaVnf(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_disconnect_SBC_Media_VNF(self.sigDriver.additinalParams_Disconnection)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:disconnect_SBC_Media_VNF', timeout=20)

    # The pubkey needs to be setup before DB restore as this is the only way
    def sigvnf_DBRestore(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_dbrestore(self.sigDriver.additinalParams_DBRestore)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:DB_Restore', timeout=60)

    # For UpgradePrecheck, Upgrade1Apply, Upgrade2Activate, Upgrade3Commit,
    # assume the VNF package has been changed by ChangePackageVersion,
    # so use sig_vnfdId_SUToLoad as vnfdId
    # For ChangePackageVersion, need to pass in the vnfdId of the TO load
    def sigvnf_ChangePackageVersion(self, vnfdId_SU):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.change_package_version(vnfdId_SU)
        # Change Package version consists of: UPGRADE then MODIFY_INFO
        self.sigDriver.check_opstatus(operation='MODIFY_INFO', operationName='', timeout=10)
        time.sleep(60)

    def sigvnf_UpgradePrecheck(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_precheck()
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_precheck', timeout=20)

    def sigvnf_Upgrade1Apply(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_1_apply(self.sigDriver.additinalParams_Upgrade1Apply)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_1_apply', timeout=90)

    def sigvnf_Upgrade2Activate(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_2_activate()
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_2_activate', timeout=30)

    def sigvnf_Upgrade3Commit(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_3_commit()
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_3_commit', timeout=90)

    def sigvnf_UpgradeArchive(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        for ap in [self.sigDriver.additinalParams_UpgradeArchive,
                   self.sigDriver.additinalParams_UpgradeArchive_Creds
                   ]:
            self.sigDriver.custom_upgrade_archive(ap)
            self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_archive', timeout=20)
            self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='cssu')

    def sigvnf_UpgradeArchive_Pubkey(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_archive(self.sigDriver.additinalParams_UpgradeArchive)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_archive', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='cssu')

    def sigvnf_UpgradeArchive_Creds(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        self.sigDriver.custom_upgrade_archive(self.sigDriver.additinalParams_UpgradeArchive_Creds)
        self.sigDriver.check_opstatus(operation='OTHER', operationName='custom:upgrade_archive', timeout=20)
        self.sigDriver.prep_backup_cssu_zip(sshtype='passwd', ziptype='cssu')

    def sigvnf_CSSUInstantiate(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        # For CreateInstantiate, cann't use setup_sigDriver as vnf is not created yet
        # use SigVnfLcmTestDriver directly
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.cssu_create_instantiate()
        self.sigDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)

    def sigvnf_DRInstantiate(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        # For CreateInstantiate, cann't use setup_sigDriver as vnf is not created yet
        # use SigVnfLcmTestDriver directly
        self.sigDriver = SigVnfLcmTestDriver(SigVnf(vnfdId=self.vnfdId))
        self.sigDriver.dr_create_instantiate()
        self.sigDriver.check_opstatus(operation='INSTANTIATE', operationName='', timeout=90)

    # rel is current release
    # For instantiation, vnf is not created yet, so not able to use setup_sigDriver()
    # need to call sigvnf_GenArtsInstantiation() directly in TS
    # def sigvnf_GenArts4Instantiation(self, rel='R20.0'):
    #     logger.info('In Func: ' + sys._getframe().f_code.co_name)
    #     self.setup_sigDriver()
    #     sigArts = SigVnfArtsGenerator(type='instantiation', rel=rel)
    #     sigArts.prep_arts()

    # rel is current release
    # 6 server_type for DR: httpserver1, httpserver2, httpserver12
    #               bkupserver1, bkupserver2, bkupserver12
    def sigvnf_GenArts4DR(self, rel=default_rel, server_type='httpserver12'):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        current_count, max_count = self.sigDriver.get_scale_status()
        sigArts = SigVnfArtsGenerator(type='dr', rel=rel, server_type=server_type)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    # rel is TO load release
    def sigvnf_GenArts4SU(self, rel=toload_rel):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        current_count, max_count = self.sigDriver.get_scale_status()
        sigArts = SigVnfArtsGenerator(type='su', rel=rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    # rel is TO load release
    def sigvnf_GenArts4CSSU(self, rel=toload_rel):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        current_count, max_count = self.sigDriver.get_scale_status()
        sigArts = SigVnfArtsGenerator(type='cssu', rel=rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    # this is test for artifact generation for su, dr, cssu
    def sigvnf_GenArtsTests(self, rel=default_rel):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        self.setup_sigDriver()
        logger.info('Generate DIF for Instantiation')
        sigArts = SigVnfArtsGenerator(type='instantiation', rel=rel)
        sigArts.prep_arts()
        current_count, max_count = self.sigDriver.get_scale_status()
        logger.info('Generate DIF for DR')
        sigArts = SigVnfArtsGenerator(type='dr', rel=rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()
        logger.info('Generate DIF for SU')
        sigArts = SigVnfArtsGenerator(type='su', rel=toload_rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()
        logger.info('Generate DIF for CSSU')
        sigArts = SigVnfArtsGenerator(type='cssu', rel=toload_rel)
        sigArts.gen_dif_updated_sc_value(sc_count=current_count)
        sigArts.prep_arts()

    ##########################################
    # Following are test suites out of the box
    ##########################################
    def sigvnf_tests_alpha(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_ScaleOut()
        self.sigvnf_ScaleIn()
        self.sigvnf_ScaleOut()
        self.sigvnf_Heal_Single()
        self.sigvnf_Backup()
        self.sigvnf_DBRestore()

    def sigvnf_tests_beta(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Scale_OutIn()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_Heal_Single()
        self.sigvnf_Heal_Multiple()
        self.sigvnf_Backup()
        self.sigvnf_DBRestore()
        self.sigvnf_Terminate()
        self.sigvnf_Delete()

    def sigvnf_tests_gamma(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_ScaleInToInit()
        self.sigvnf_Scale_OutIn()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_Heal_Single()
        self.sigvnf_Heal_Multiple()
        self.sigvnf_Backup()
        self.sigvnf_DBRestore()
        self.sigvnf_GenArts4DR(rel=default_rel)
        self.sigvnf_tests_td()
        sigvnf_UploadVnfpkg(swVersion=sigVersion, type='dr')
        setup_vnfdIds()
        self.refresh_sigDriver(sig_vnfdId)
        self.sigvnf_tests_dr()

    def sigvnf_tests_cim(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Backup_Remote_Creds1()

    def sigvnf_tests_cimcb(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Backup_Remote_Creds1()

    def sigvnf_tests_cimcsb(self):
        self.sigvnf_CreateInstantiate()
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Backup_Remote_Creds1()
        # self.sigvnf_ScaleOutToMax()
        self.sigvnf_ScaleOut()
        self.sigvnf_Backup_Remote_Creds1()

    def sigvnf_tests_td(self):
        self.sigvnf_Terminate()
        self.sigvnf_Delete()

    def sigvnf_tests_mcdb(self):
        self.sigvnf_Modify_DisableAutoBackup()
        self.sigvnf_Modify_DisableAutoScale()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_DisconnectSbcMediaVnf()
        self.sigvnf_ConnectSbcMediaVnf()
        self.sigvnf_Backup()

    def sigvnf_tests_bkup(self):
        self.sigvnf_Backup()

    def sigvnf_tests_bkup_rt(self):
        self.sigvnf_Backup_Remote1()
        self.sigvnf_Backup_Remote2()
        self.sigvnf_Backup_Remote12()

    def sigvnf_tests_bkup_rt_creds(self):
        self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_Backup_Remote_Creds2()
        self.sigvnf_Backup_Remote_Creds12()

    def sigvnf_tests_br(self):
        self.sigvnf_Backup_Remote_Creds1()
        # self.sigvnf_Backup_Remote_Creds12()
        self.sigvnf_DBRestore()

    def sigvnf_tests_heal(self):
        self.sigvnf_Heal_Single()
        self.sigvnf_Heal_Multiple()

    def sigvnf_tests_heals(self):
        self.sigvnf_Heal_Single()

    def sigvnf_tests_healsr(self):
        self.sigvnf_Heal_Single_Random()

    def sigvnf_tests_healsr_mt(self):
        self.sigvnf_Heal_Single_Random_MTs(times=5)

    def sigvnf_tests_healm(self):
        self.sigvnf_Heal_Multiple()

    def sigvnf_tests_scale(self):
        self.sigvnf_Scale_OutIn()
        self.sigvnf_ScaleOutToMax()
        self.sigvnf_ScaleInToInit()

    def sigvnf_tests_scale_outin(self):
        self.sigvnf_Scale_OutIn()

    def sigvnf_tests_scale2max(self):
        self.sigvnf_ScaleOutToMax()

    def sigvnf_tests_scale2init(self):
        self.sigvnf_ScaleInToInit()

    def sigvnf_tests_chgpkvern(self, vnfdId_SU):
        self.sigvnf_ChangePackageVersion(vnfdId_SU)

    def sigvnf_tests_su(self):
        self.sigvnf_UpgradePrecheck()
        self.sigvnf_Upgrade1Apply()
        self.sigvnf_Upgrade2Activate()
        self.sigvnf_Upgrade3Commit()

    def sigvnf_tests_ua(self):
        self.sigvnf_UpgradeArchive()

    def sigvnf_tests_ua_td(self):
        self.sigvnf_UpgradeArchive()
        self.sigvnf_Terminate()
        self.sigvnf_Delete()

    def sigvnf_tests_cssu(self):
        self.sigvnf_CSSUInstantiate()

    def sigvnf_tests_dr(self):
        self.sigvnf_DRInstantiate()

    def sigvnf_tests_dr_v2(self, server_type='httpserver12'):
        self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_GenArts4DR(rel=default_rel, server_type=server_type)
        self.sigvnf_tests_td()
        sigvnf_UploadVnfpkg(swVersion=sigVersion, type='dr')
        setup_vnfdIds()
        self.refresh_sigDriver(sig_vnfdId)
        self.sigvnf_DRInstantiate()

    def sigvnf_tests_cssu_v2(self):
        self.sigvnf_Backup_Remote_Creds1()
        self.sigvnf_GenArts4CSSU(rel=toload_rel)
        self.sigvnf_tests_ua_td()
        sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='cssu')
        setup_vnfdIds()
        self.refresh_sigDriver(sig_vnfdId_SU)
        self.sigvnf_CSSUInstantiate()

    def sigvnf_tests_su_v2(self):
        self.sigvnf_GenArts4SU(rel=toload_rel)
        sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='su')
        setup_vnfdIds()
        self.sigvnf_tests_chgpkvern(sig_vnfdId_SU)
        self.refresh_sigDriver(sig_vnfdId_SU)
        self.sigvnf_UpgradePrecheck()
        self.sigvnf_Upgrade1Apply()
        self.sigvnf_Upgrade2Activate()
        self.sigvnf_Upgrade3Commit()


########################################################################################################################
# Upper caller Test Suite
########################################################################################################################
def TS_sigvnf_tests_alpha():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_alpha()

def TS_sigvnf_tests_beta():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_beta()

def TS_sigvnf_tests_gamma():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_gamma()

def TS_sigvnf_tests_arts():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_GenArtsTests()

def TS_sigvnf_tests_td():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_td()

def TS_sigvnf_tests_td_toload():
    lcmDriver = LcmTestDriver(sigVersion_SU)
    lcmDriver.sigvnf_tests_td()

def TS_sigvnf_tests_cim():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cim()

def TS_sigvnf_tests_cimcb():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cimcb()

def TS_sigvnf_tests_cimcsb():
    sigvnf_GenArtsInstantiation()
    sigvnf_UploadVnfpkg(swVersion=sigVersion, type='instantiation')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cimcsb()

def TS_sigvnf_tests_mcdb():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_mcdb()

def TS_sigvnf_tests_bkup():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_bkup()

def TS_sigvnf_tests_bkup_rt():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_bkup_rt()

def TS_sigvnf_tests_bkup_rt_creds():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_bkup_rt_creds()

def TS_sigvnf_tests_br():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_br()

def TS_sigvnf_tests_heal():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_heal()

def TS_sigvnf_tests_heal_m():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_healm()

def TS_sigvnf_tests_heal_s():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_heals()

def TS_sigvnf_tests_heal_s_r():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_healsr()

def TS_sigvnf_tests_heal_s_r_mt():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_healsr_mt()

# def TS_sigvnf_tests_scale(vnfdId):
#     lcmDriver = LcmTestDriver(vnfdId)
#     lcmDriver.sigvnf_tests_scale()

def TS_sigvnf_tests_scale():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale()

def TS_sigvnf_tests_scale_outin():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale_outin()

def TS_sigvnf_tests_scale_toload():
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_scale()

def TS_sigvnf_tests_scale2max():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale2max()

def TS_sigvnf_tests_scale2max_toload():
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_scale2max()
    
def TS_sigvnf_tests_scale2init():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_scale2init()

def TS_sigvnf_tests_sh():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    # lcmDriver.sigvnf_ScaleOutToMax()
    lcmDriver.sigvnf_ScaleOut()
    lcmDriver.sigvnf_ScaleOut()
    lcmDriver.sigvnf_ScaleOut()
    lcmDriver.sigvnf_ScaleOut()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_heal()

def TS_sigvnf_tests_su():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_GenArts4SU(rel=toload_rel)
    sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='su')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_chgpkvern(sig_vnfdId_SU)
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_su()

def TS_sigvnf_tests_cssu():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_GenArts4CSSU(rel=toload_rel)
    lcmDriver.sigvnf_tests_ua_td()
    sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='cssu')
    setup_vnfdIds()
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_cssu()

# backup -> gen arts for DR -> terminate, delete
# -> upload new vnf pkg -> DR
# The original version for dr
# def TS_sigvnf_tests_dr():
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_Backup_Remote_Creds1()
#     lcmDriver.sigvnf_GenArts4DR(rel=default_rel, server_type='httpserver12')
#     lcmDriver.sigvnf_tests_td()
#     sigvnf_UploadVnfpkg(swVersion=sigVersion, type='dr')
#     setup_vnfdIds()
#     lcmDriver = LcmTestDriver(sig_vnfdId)
#     lcmDriver.sigvnf_tests_dr()

# Create -> Instantiation -> Modify Auto Operations -> Connection2Media -> Backup
# -> UpgradeArchive -> Terminate -> Delete -> CSSU Instantiation
def TS_sigvnf_tests_cimcb_cssu():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cimcb()
    lcmDriver.sigvnf_GenArts4CSSU(rel=toload_rel)
    lcmDriver.sigvnf_tests_ua_td()
    sigvnf_UploadVnfpkg(swVersion=sigVersion_SU, type='cssu')
    lcmDriver = LcmTestDriver(sig_vnfdId_SU)
    lcmDriver.sigvnf_tests_cssu()

def TS_sigvnf_tests_dr():
    for server_type in server_type_list:
        lcmDriver = LcmTestDriver(sig_vnfdId)
        lcmDriver.sigvnf_tests_dr_v2(server_type=server_type)

def TS_sigvnf_tests_dr_http1():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr_v2(server_type='httpserver1')

def TS_sigvnf_tests_dr_http2():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr_v2(server_type='httpserver2')

def TS_sigvnf_tests_dr_http12():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr_v2(server_type='httpserver12')

def TS_sigvnf_tests_dr_creds_1():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr_v2(server_type='bkupserver1')

def TS_sigvnf_tests_dr_creds_2():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr_v2(server_type='bkupserver2')

def TS_sigvnf_tests_dr_creds_12():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_dr_v2(server_type='bkupserver12')

def TS_sigvnf_tests_cssu_v2():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_cssu_v2()

def TS_sigvnf_tests_su_v2():
    lcmDriver = LcmTestDriver(sig_vnfdId)
    lcmDriver.sigvnf_tests_su_v2()

########################################################################################################################
# Setup env variables
def setup_env():
    # Disable following warning:
    # D:\Program Files (x86)\Python37-32\lib\site-packages\urllib3\connectionpool.py:1004:
    # InsecureRequestWarning: Unverified HTTPS request is being made.
    # Adding certificate verification is strongly advised.
    # See: https://urllib3.readthedocs.io/en/latest/advanced-usage.html#ssl-warnings
    urllib3.disable_warnings()

def setup_logging():
    logger.setLevel(logging.INFO)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    fh = logging.FileHandler(log_file)
    fh.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s %(levelname)s:%(name)s:%(message)s')
    ch.setFormatter(formatter)
    fh.setFormatter(formatter)
    logger.addHandler(ch)
    logger.addHandler(fh)

# Here use software version and vnfdId to identify vnf
# To-do: use vnf name to identify vnf
def setup_vnfdIds():
    # First to find vnfdId for both signaling plane and media plane
    global sig_vnfdId
    global media_vnfdId
    global sig_vnfdId_SU
    global media_vnfdId_SU
    global sigVersion
    global sigVersion_SU
    global mediaVersion
    global mediaVersion_SU

    # sig_vnfdId = ''
    # media_vnfdId = ''
    # sig_vnfdId_SU = ''
    # media_vnfdId_SU = ''
    # sigVersion = '37.28.06'
    # sigVersion_SU = '37.28.06.0020'
    # mediaVersion = 'an100052'
    # mediaVersion_SU = 'an100053'

    vnfpkgs = VnfPkgs()
    vnfpkgs.get_vnfpkgs()
    vnfpkgs.dump_vnfpkgs()

    # Determine vnfdId via svnfSoftwareVersion
    # For sig plane, example of vnfSoftwareVersion:
    #'vnfSoftwareVersion': 'sbclcm01~37.28.06'
    # 'vnfSoftwareVersion': 'sbclcm01~37.28.06.0020'
    if vnfpkgs.vnfpkgs_list:
        for vp in vnfpkgs.vnfpkgs_list:
            if vp.vnfProductName == 'SBC':
                if vp.vnfSoftwareVersion.endswith(sigVersion):
                    # This is sig plane vnfp. Normal case for sig plane
                    sig_vnfdId = vp.vnfdId
                    logger.info('Sig Plane vnfdId: ' + sig_vnfdId)
                    logger.info('Sig Plane vnfSoftwareVersion: ' + vp.vnfSoftwareVersion)
                elif vp.vnfSoftwareVersion.endswith(sigVersion_SU):
                    # This is sig plane vnfp. SU case for sig plane
                    sig_vnfdId_SU = vp.vnfdId
                    logger.info('Sig Plane vnfdId_SU: ' + sig_vnfdId_SU)
                    logger.info('Sig Plane vnfSoftwareVersion for SUTOLoad: ' + vp.vnfSoftwareVersion)
            elif vp.vnfProductName == 'SBC-media':
                if vp.vnfSoftwareVersion.endswith(mediaVersion):
                    # This is media plane vnfp. Normal case for media plane
                    media_vnfdId = vp.vnfdId
                    logger.info('Media Plane vnfdId: ' + media_vnfdId)
                    logger.info('Media Plane vnfSoftwareVersion: ' + vp.vnfSoftwareVersion)
                elif vp.vnfSoftwareVersion.endswith(mediaVersion_SU):
                    # This is media plane vnfp. SU case for media plane
                    media_vnfdId_SU = vp.vnfdId
                    logger.info('Media Plane vnfdId_SU: ' + media_vnfdId_SU)
                    logger.info('Media Plane vnfSoftwareVersion for SUTOLoad: ' + vp.vnfSoftwareVersion)
            else:
                logger.info('Currently supported vnfProductName: SBC or SBC-media.')
                exit(1)


########################################################################################################################
# Main
# Some functions:
# setup_env() : disable InsecureRequestWarning by urllib3
# setup_logging() : setup logger. use info as default
# setup_vnfdIds() : read all vnf pkgs. determine vnfdId of base load and TO load via vnfSoftwareVersion
# sigvnf_GenArts() : generate artifacts directly based on DIFs that have already been ready (manually prepared)
#                    this is different with the methods sigvnf_GenArts4XX in class LcmTestDriver
#                    the LcmTestDriver->sigvnf_GenArts4XX will read sc values from vnf then modify DIFs then gen arts
#
########################################################################################################################
if __name__ == '__main__':

    setup_env()

    setup_logging()

    setup_vnfdIds()

    # sigvnf_GenArts()

    # DR then CSSU
    # TS_sigvnf_tests_dr()
    # TS_sigvnf_tests_cssu()

    TS_sigvnf_tests_td()
    # TS_sigvnf_tests_cim()
    TS_sigvnf_tests_cimcb()
    # TS_sigvnf_tests_bkup_rt_creds()
    # TS_sigvnf_tests_arts()
    # TS_sigvnf_tests_dr()
    # TS_sigvnf_tests_cssu()
    # TS_sigvnf_tests_td()
    # TS_sigvnf_tests_td_toload()
    # TS_sigvnf_tests_su()

    # TS_sigvnf_tests_scale2max()
    # TS_sigvnf_tests_scale2init()
    # TS_sigvnf_tests_scale_outin()

    # TS_sigvnf_tests_beta()

    # TS_sigvnf_tests_cimcsb()
    # TS_sigvnf_tests_dr_v2()

    # TS_sigvnf_tests_cssu_v2()

    # TS_sigvnf_tests_td()
    # TS_sigvnf_tests_cimcsb()
    # TS_sigvnf_tests_su_v2()

    # TS_sigvnf_tests_td()
    # TS_sigvnf_tests_cimcsb()

    # TS_sigvnf_tests_scale2max()
    # TS_sigvnf_tests_scale2init()

    # TS_sigvnf_tests_heal_m()
    # TS_sigvnf_tests_heal_s()

    # TS_sigvnf_tests_heal_s_r()
    # TS_sigvnf_tests_heal_s_r_mt()

    # TS_sigvnf_tests_br()
    # TS_sigvnf_tests_dr_v2()
    # TS_sigvnf_tests_scale2init()
    # TS_sigvnf_tests_scale2max()

    # TS_sigvnf_tests_dr()
    # TS_sigvnf_tests_dr_http12()
    # TS_sigvnf_tests_dr_creds_1()
    # TS_sigvnf_tests_dr_creds_2()
    # TS_sigvnf_tests_dr_creds_12()





