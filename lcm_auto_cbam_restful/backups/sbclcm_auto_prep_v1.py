# -*- coding:UTF-8 -*-

import requests
import json
import time
import paramiko
import os
import sys
import logging
import shutil
import zipfile
import openpyxl
from openpyxl import load_workbook

rel='R20.0'
yact_server_ip = '135.252.41.216'
yact_user = 'yact-user'
yact_passwd = '123456'
yact_dir = '/home/yact-user/shawnx/20.0/'

working_dir = r'D:\Programs\JetBrains\PycharmProjects\py37projects\lcm_auto_cbam_restful'
sig_data_dir = working_dir + r'\data\sig-plane'
sig_arts_dir = working_dir + r'\data\sig-plane-arts'
media_data_dir = working_dir + r'\data\media-plane'

arts_type = ['instantiation', 'su', 'dr', 'cssu']

sig_dif_name = 'SBC-signaling_R20.0-sbclcm03.xlsm'
sig_dr_dif_name = 'SBC-signaling_R20.0-sbclcm03-dr.xlsm'
sig_su_dif_name = 'SBC-signaling_R20.0-sbclcm03-su.xlsm'
sig_cssu_dif_name = 'SBC-signaling_R20.0-sbclcm03-cssu.xlsm'

sig_dif_dict = {
    'base': sig_dif_name,
    'cssu': sig_cssu_dif_name,
    'dr':   sig_dr_dif_name,
    'su':   sig_su_dif_name
}

sig_ne_name = 'sbclcm03'
sig_zip = sig_ne_name + '.zip'

sig_dif_file = sig_arts_dir + '\\' + sig_dif_name

sig_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package.zip'
sig_dr_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-DR.zip'
sig_su_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-SUToLoad.zip'
sig_cssu_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-CSSUToLoad.zip'

# orig_instantiation_json_file = 'orig_LCM_instantiate_params.json'
instantiation_json_file = 'LCM_instantiate_params.json'
dr_instantiation_json_file = 'dr_LCM_instantiate_params.json'
cssu_instantiation_json_file = 'cssu_LCM_instantiate_params.json'

os_passwd = 'a321sbc'

backup_server_ip = '10.75.44.7'
backup_server_login = 'root'
backup_server_passwd = 'newsys'
backup_server_dir = '/var/www/html/sbclcm-auto/'
backup_file_name = 'backup.zip'
backup_file = backup_server_dir + backup_file_name
local_backup_dir = 'backup'

ap_instantiation = {
    "backup_file1": "",
    "backup_file2": "",
    "backup_server_credentials1": "",
    "backup_server_credentials2": "",
    "bulk_conf_url": "http://10.75.44.7/sbclcm03/bulkconf_artifacts.zip",
    "skip_health_check": "No",
    "upgrade_file": "",
    "upgrade_server_credentials":""
  }

ap_dr_instantiation = {
    "backup_file1": "http://10.75.44.7/sbclcm-auto/backup.zip",
    "backup_file2": "http://10.75.44.7/sbclcm-auto/backup.zip",
    "backup_server_credentials1": "",
    "backup_server_credentials2": "",
    "bulk_conf_url": "",
    "skip_health_check": "No",
    "upgrade_file": "",
    "upgrade_server_credentials":""
  }

ap_cssu_instantiation = {
    "backup_file1": "",
    "backup_file2": "",
    "backup_server_credentials1": "",
    "backup_server_credentials2": "",
    "bulk_conf_url": "",
    "skip_health_check": "No",
    "upgrade_file": "http://10.75.44.7/sbclcm-auto/cssu_archive.zip",
    "upgrade_server_credentials":""
  }

def ssh_command(cmd, ip, login, pass_key, type):
    print('In Func: ' + sys._getframe().f_code.co_name)
    print('cmd: ' + cmd)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    if type == 'passwd':
        ssh.connect(hostname=ip, port=22, username=login, password=pass_key)
    elif type == 'pubkey':
        private_key = paramiko.RSAKey.from_private_key_file(pass_key)
        ssh.connect(hostname=ip, port=22, username=login, pkey=private_key)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    res, err = stdout.read(), stderr.read()
    result = res if res else err
    ssh.close()
    return result.decode()

def ssh_scp_put(ip, login, passwd, local_file, remote_file):
    print('In Func: ' + sys._getframe().f_code.co_name)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.put(local_file, remote_file)
    ssh.close()

def ssh_scp_get(ip, login, passwd, remote_file, local_file):
    print('In Func: ' + sys._getframe().f_code.co_name)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.get(remote_file, local_file)
    ssh.close()


# Supported type: ['instantiation', 'su', 'dr', 'cssu']
# rel is needed to handle SU case, SU TO load may be different release
def gen_arts_sig(type='instantiation', rel='R20.0'):
    print('type == ', type)
    os.chdir(sig_arts_dir)

    if type not in arts_type:
        print('Only ', str(arts_type), ' Supported.')
        exit(1)

    ts = int(time.time())
    print('Current time is: ', ts)
    sig_yact_dir = yact_dir + str(ts) + '/'

    ssh_type = 'passwd'
    cmd = 'mkdir -p ' + sig_yact_dir
    ssh_command(cmd, yact_server_ip, yact_user, yact_passwd, ssh_type)

    if type == 'instantiation':
        sig_dif_file = sig_arts_dir + '\\' + sig_dif_name
        sig_rmt_dif_file = sig_yact_dir + sig_dif_name
    elif type == 'dr':
        sig_dif_file = sig_arts_dir + '\\' + sig_dr_dif_name
        sig_rmt_dif_file = sig_yact_dir + sig_dr_dif_name
    elif type == 'su':
        sig_dif_file = sig_arts_dir + '\\' + sig_su_dif_name
        sig_rmt_dif_file = sig_yact_dir + sig_su_dif_name
    elif type == 'cssu':
        sig_dif_file = sig_arts_dir + '\\' + sig_cssu_dif_name
        sig_rmt_dif_file = sig_yact_dir + sig_cssu_dif_name

    print('sig_dif_file: ', sig_dif_file)
    print('sig_rmt_dif_file: ', sig_rmt_dif_file)
    ssh_scp_put(yact_server_ip, yact_user, yact_passwd, sig_dif_file, sig_rmt_dif_file)

    yact_cmd = '/home/yact/YACT/yact.sh gen-by-dif ' + sig_rmt_dif_file + ' SBC-signaling ' + rel
    print('yact_cmd: ', yact_cmd)
    cmd = 'cd ' + sig_yact_dir + ';' + yact_cmd
    result = ssh_command(cmd, yact_server_ip, yact_user, yact_passwd, ssh_type)
    print('result of yact_cmd: ', result)

    # Handle SU case
    if type == 'su':
        pkg_tools_dir = sig_yact_dir + sig_ne_name + '/package/pkg_tools/'
        print('su pkg_tools_dir: ', pkg_tools_dir)

        # ship from load vnf pkg to pkg_tools_dir and rename it to vnfpkg-from-load.zip
        vnfpkg_from_load = sig_data_dir + '\\' + sig_vnfpkg_name
        rmt_vnfpkg_from_load = pkg_tools_dir + 'vnfpkg-from-load.zip'
        ssh_scp_put(yact_server_ip, yact_user, yact_passwd, vnfpkg_from_load, rmt_vnfpkg_from_load)

        # run vnf_upgrade_pkg_gen
        cmd = 'cd ' + pkg_tools_dir + ';' + './vnf_upgrade_pkg_gen -n ../Nokia_sig_SBC-VNF_Package.zip ' \
                                            '-c ./vnfpkg-from-load.zip ' + sig_ne_name
        print('vnf_upgrade_pkg_gen cmd: ', cmd)
        ssh_command(cmd, yact_server_ip, yact_user, yact_passwd, ssh_type)

        # now ship the upgrade vnf pkg to local
        # example: sbclcm03_Nokia_sig_SBC_upgrade-VNF_Package.zip
        vnfpkg_upgrade = pkg_tools_dir + sig_ne_name + '_Nokia_sig_SBC_upgrade-VNF_Package.zip'
        print('upgrade package: ', vnfpkg_upgrade)
        vnfpkg_upgrade_local = sig_data_dir + '\\' + sig_su_vnfpkg_name
        ssh_scp_get(yact_server_ip, yact_user, yact_passwd, vnfpkg_upgrade, vnfpkg_upgrade_local)
    # handle other cases
    else:
        # zip remote sig_art_files
        cmd = 'cd ' + sig_yact_dir + ';' + 'zip -r ' + sig_ne_name + '.zip ' + sig_ne_name
        result = ssh_command(cmd, yact_server_ip, yact_user, yact_passwd, ssh_type)
        print('result of zip sig: ', result)

        if os.path.exists(sig_zip):
            os.remove(sig_zip)

        if os.path.exists(sig_ne_name):
            shutil.rmtree(sig_ne_name)

        rmt_zip_file = sig_yact_dir + sig_zip
        ssh_scp_get(yact_server_ip, yact_user, yact_passwd, rmt_zip_file, sig_zip)

        print('unzip ', sig_zip)
        zip_file = zipfile.ZipFile(sig_zip)
        # os.mkdir(sig_ne_name)
        for names in zip_file.namelist():
            zip_file.extract(names, '.')
        zip_file.close()


def cp_vnf_pkg(type='instantiation'):
    # cp the Nokia_sig_SBC-VNF_Package.zip to data\sig-plane
    # to correct name
    os.chdir(sig_arts_dir)
    if not os.path.exists(sig_ne_name):
        print('artifacts dir not exist. Exit.')
        exit(1)

    srcfile = sig_ne_name + r'\package\Nokia_sig_SBC-VNF_Package.zip'

    if type == 'instantiation':
        dstfile = sig_data_dir + '\\' + sig_vnfpkg_name
    elif type == 'dr':
        dstfile = sig_data_dir + '\\' + sig_dr_vnfpkg_name
    elif type == 'su':
        dstfile = sig_data_dir + '\\' + sig_su_vnfpkg_name
    elif type == 'cssu':
        dstfile = sig_data_dir + '\\' + sig_cssu_vnfpkg_name

    print('Copy ', srcfile, ' to ', dstfile)
    shutil.copyfile(srcfile, dstfile)

# def cp_json_files():
#     # copy instnatiation json
#     os.chdir(sig_arts_dir)
#     print('copy ', instantiation_json_file, ' to ', orig_instantiation_json_file)
#     srcfile = sig_ne_name + r'\package\cbam_json\LCM_instantiate_params.json'
#     dstfile = sig_data_dir + '\\' + orig_instantiation_json_file
#     shutil.copyfile(srcfile, dstfile)
#     # copy extension json
#     print('copy LCM_extensions.json')
#     srcfile = sig_ne_name + r'\package\cbam_json\LCM_extensions.json'
#     dstfile = sig_data_dir + '\\' + 'LCM_extensions.json'
#     shutil.copyfile(srcfile, dstfile)
#     # copy LCM_pkg_upgrade_api4.json
#     print('copy LCM_pkg_upgrade_api4.json')
#     srcfile = sig_ne_name + r'\package\cbam_json\LCM_pkg_upgrade_api4.json'
#     dstfile = sig_data_dir + '\\' + 'LCM_pkg_upgrade_api4.json'
#     shutil.copyfile(srcfile, dstfile)

def prep_instantiation_json(type='instantiation'):
    os.chdir(sig_arts_dir)
    if not os.path.exists(sig_ne_name):
        print('artifacts dir not exist. Exit.')
        exit(1)

    json_file = sig_ne_name + r'\package\cbam_json' + '\\' + instantiation_json_file

    # In case of dr, the instantiation json needs to be extracted from backup.zip
    # In case of su, instantiation json is not needed
    if type == 'dr':
        if os.path.exists(backup_file_name):
            os.remove(backup_file_name)

        if os.path.exists(local_backup_dir):
            shutil.rmtree(local_backup_dir)

        ssh_scp_get(backup_server_ip, backup_server_login, backup_server_passwd, backup_file, backup_file_name)

        print('unzip ', backup_file_name)
        os.mkdir(local_backup_dir)
        zip_file = zipfile.ZipFile(backup_file_name)
        for names in zip_file.namelist():
            zip_file.extract(names, local_backup_dir)
        zip_file.close()

        json_file = local_backup_dir + '\\' + instantiation_json_file

    try:
        with open(json_file, 'r') as jFile:
            data = json.load(jFile)
    except Exception as exc:
        print('ERROR: Failed to load data from {0}.\n[{1}]'.format(
            json_file, str(exc)))

    data['vimConnectionInfo'][0]['accessInfo']['password'] = os_passwd
    if type == 'instantiation':
        data['additionalParams'] = ap_instantiation
        output_json = sig_data_dir + '\\' + instantiation_json_file
    elif type == 'dr':
        data['additionalParams'] = ap_dr_instantiation
        output_json = sig_data_dir + '\\' + dr_instantiation_json_file
    elif type == 'cssu':
        data['additionalParams'] = ap_cssu_instantiation
        output_json = sig_data_dir + '\\' + cssu_instantiation_json_file

    try:
        with open(output_json, 'w') as jFile:
            json.dump(data, jFile, indent=2)
    except Exception as exc:
        print('ERROR: Failed to write data to {0}.\n[{1}]'.format(
            output_json, str(exc)))

    print('instantiation json for ', type, ' created completed.')


# json load, dump
# with open(file_name, 'r') as f:
#     data = json.load(f)
# with open('output.json', 'w') as f:
#     json.dump(data, f)
def prep_dr_instantiation_json():
    # First, get backup zip from backup server
    # The orig instantiation json is extracted from backup zip
    os.chdir(sig_arts_dir)

    if os.path.exists(backup_file_name):
        os.remove(backup_file_name)

    if os.path.exists(local_backup_dir):
        shutil.rmtree(local_backup_dir)

    ssh_scp_get(backup_server_ip, backup_server_login, backup_server_passwd, backup_file, backup_file_name)

    print('unzip ', backup_file_name)
    os.mkdir(local_backup_dir)
    zip_file = zipfile.ZipFile(backup_file_name)
    for names in zip_file.namelist():
        zip_file.extract(names, local_backup_dir)
    zip_file.close()

    json_file = local_backup_dir + '\\' + instantiation_json_file
    try:
        with open(json_file, 'r') as jFile:
            data = json.load(jFile)
    except Exception as exc:
        print('ERROR: Failed to load data from {0}.\n[{1}]'.format(
            json_file, str(exc)))

    data['additionalParams'] = ap_dr_instantiation
    data['vimConnectionInfo'][0]['accessInfo']['password'] = os_passwd

    try:
        with open(dr_instantiation_json_file, 'w') as jFile:
            json.dump(data, jFile, indent=2)
    except Exception as exc:
        print('ERROR: Failed to write data to {0}.\n[{1}]'.format(
            dr_instantiation_json_file, str(exc)))

    # copy the dr instantiation to data\sig-plane
    print('copy ', dr_instantiation_json_file)
    srcfile = dr_instantiation_json_file
    dstfile = sig_data_dir + '\\' + dr_instantiation_json_file
    shutil.copyfile(srcfile, dstfile)


def gen_dif_updated_sc_value(type='su', sc_count=5):
    os.chdir(sig_arts_dir)

    if type == 'su':
        workbook = sig_su_dif_name
    elif type == 'cssu':
        workbook = sig_cssu_dif_name
    elif type == 'dr':
        workbook = sig_dr_dif_name

    wb = load_workbook(workbook, keep_vba=True)
    sh = wb['DP']
    colB = sh['B']

    row_list = []
    for i in colB:
        if i.value == 'group_tag':
            row_list.append(i.row)

    row_sc = None
    for r in row_list:
        if sh.cell(row=r, column=26).value == 'SC':
            row_sc = r
    if row_sc is not None:
        print('row_sc: ', row_sc)

    row_vm_count = None
    for r in range(row_sc, row_sc + 8):
        if sh.cell(row=r, column=2).value == 'vm_count':
            row_vm_count = r
    if row_vm_count is not None:
        print('row_vm_count: ', row_vm_count)

    print('current SC vaule of vm_count: ', sh.cell(row=row_vm_count, column=26).value)

    sh.cell(row=row_vm_count, column=26).value = sc_count

    # wb.save(filename='output.xlsm')
    wb.save(filename=workbook)


def prep_arts_instantiation():
    # 1. generate arts from yact server
    # 2. generate instantiation json
    # 3. cp vnf pkg to data\sig-plane
    gen_arts_sig(type='instantiation')
    prep_instantiation_json(type='instantiation')
    cp_vnf_pkg(type='instantiation')

def prep_arts_dr():
    gen_arts_sig(type='dr')
    prep_instantiation_json(type='dr')
    cp_vnf_pkg(type='dr')

def prep_arts_cssu():
    gen_arts_sig(type='cssu', rel='R20.0')
    prep_instantiation_json(type='cssu')
    cp_vnf_pkg(type='cssu')

def prep_arts_su():
    # SU is more complicated
    # 1. generate ToLoad vnf pkg from yact server
    # 2. get FromLoad vnf pkg
    #   - via local instantiation vnf pkg (easier)
    #   - via fetching vnf pkg from cbam
    # 3. vnf pkg upgrade
    gen_arts_sig(type='su', rel='R20.0')


if __name__ == '__main__':

    # prep_arts_instantiation()
    # prep_arts_dr()
    # prep_arts_cssu()
    # prep_arts_su()

    gen_dif_updated_sc_value(type='su', sc_count=5)
    prep_arts_su()

