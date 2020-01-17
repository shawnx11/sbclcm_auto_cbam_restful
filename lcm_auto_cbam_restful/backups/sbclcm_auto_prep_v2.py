# -*- coding:UTF-8 -*-

import requests
import json
import time
import paramiko
import os
import sys
import logging
import shutil
import zipfile
import openpyxl
import logging
from openpyxl import load_workbook

rel='R20.0'
yact_server_ip = '135.252.41.216'
yact_user = 'yact-user'
yact_passwd = '123456'
yact_dir = '/home/yact-user/shawnx/20.0/'

working_dir = r'D:\Programs\JetBrains\PycharmProjects\py37projects\lcm_auto_cbam_restful'
sig_data_dir = working_dir + r'\data\sig-plane'
media_data_dir = working_dir + r'\data\media-plane'
sig_arts_dir = working_dir + r'\data\sig-plane-arts'
media_arts_dir = working_dir + r'\data\media-plane-arts'
log_file = working_dir + r'\sbclcm_auto_prep.log'

global logger
logger = logging.getLogger()

def ssh_command(cmd, ip, login, pass_key, type):
    print('In Func: ' + sys._getframe().f_code.co_name)
    print('cmd: ' + cmd)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    if type == 'passwd':
        ssh.connect(hostname=ip, port=22, username=login, password=pass_key)
    elif type == 'pubkey':
        private_key = paramiko.RSAKey.from_private_key_file(pass_key)
        ssh.connect(hostname=ip, port=22, username=login, pkey=private_key)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    res, err = stdout.read(), stderr.read()
    result = res if res else err
    ssh.close()
    return result.decode()

def ssh_scp_put(ip, login, passwd, local_file, remote_file):
    print('In Func: ' + sys._getframe().f_code.co_name)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.put(local_file, remote_file)
    ssh.close()

def ssh_scp_get(ip, login, passwd, remote_file, local_file):
    print('In Func: ' + sys._getframe().f_code.co_name)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=22, username=login, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh.get_transport())
    sftp = ssh.open_sftp()
    sftp.get(remote_file, local_file)
    ssh.close()

def setup_logging():
    logger.setLevel(logging.INFO)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    fh = logging.FileHandler(log_file)
    fh.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s %(levelname)s:%(name)s:%(message)s')
    ch.setFormatter(formatter)
    fh.setFormatter(formatter)
    logger.addHandler(ch)
    logger.addHandler(fh)

# Signaling plane VNF artifact generator
# type : ['instantiation', 'su', 'dr', 'cssu']
# rel : load release to artifacts to be generated, to be passed to yact tool
class SigVnfArtsGenerator(object):
    def __init__(self, type, rel):
        self.type = type
        self.rel = rel
        # self.torel = torel
        self.arts_type = ['instantiation', 'su', 'dr', 'cssu']
        # self.rel = 'R20.0'
        # self.torel = 'R20.0'
        self.yact_server_ip = '135.252.41.216'
        self.yact_user = 'yact-user'
        self.yact_passwd = '123456'
        self.yact_dir = '/home/yact-user/shawnx/20.0/'
        self.sig_dif_name = 'SBC-signaling_R20.0-sbclcm03.xlsm'
        self.sig_dr_dif_name = 'SBC-signaling_R20.0-sbclcm03-dr.xlsm'
        self.sig_su_dif_name = 'SBC-signaling_R20.0-sbclcm03-su.xlsm'
        self.sig_cssu_dif_name = 'SBC-signaling_R20.0-sbclcm03-cssu.xlsm'
        self.sig_dif_dict = {
            'base': self.sig_dif_name,
            'cssu': self.sig_cssu_dif_name,
            'dr': self.sig_dr_dif_name,
            'su': self.sig_su_dif_name
        }
        self.sig_ne_name = 'sbclcm03'
        self.sig_zip = self.sig_ne_name + '.zip'
        self.sig_dif_file = sig_arts_dir + '\\' + self.sig_dif_name
        self.sig_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package.zip'
        self.sig_dr_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-DR.zip'
        self.sig_su_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-SUToLoad.zip'
        self.sig_cssu_vnfpkg_name = 'Nokia_sig_SBC-VNF_Package-CSSUToLoad.zip'
        # orig_instantiation_json_file = 'orig_LCM_instantiate_params.json'
        self.instantiation_json_file = 'LCM_instantiate_params.json'
        self.dr_instantiation_json_file = 'dr_LCM_instantiate_params.json'
        self.cssu_instantiation_json_file = 'cssu_LCM_instantiate_params.json'
        self.os_passwd = 'a321sbc'
        self.backup_server_ip = '10.75.44.7'
        self.backup_server_login = 'root'
        self.backup_server_passwd = 'newsys'
        self.backup_server_dir = '/var/www/html/sbclcm-auto/'
        self.backup_file_name = 'backup.zip'
        self.ssh_type = 'passwd'
        self.backup_file = self.backup_server_dir + self.backup_file_name
        self.local_backup_dir = 'backup'
        self.ap_instantiation = {
            "backup_file1": "",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            # "bulk_conf_url": "http://10.75.44.7/sbclcm03/bulkconf_artifacts.zip",
            "bulk_conf_url": "http://10.75.44.7/sbclcm-auto/bulkconf_artifacts.zip",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_dr_instantiation = {
            "backup_file1": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_file2": "http://10.75.44.7/sbclcm-auto/backup.zip",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "",
            "upgrade_server_credentials": ""
        }
        self.ap_cssu_instantiation = {
            "backup_file1": "",
            "backup_file2": "",
            "backup_server_credentials1": "",
            "backup_server_credentials2": "",
            "bulk_conf_url": "",
            "skip_health_check": "No",
            "upgrade_file": "http://10.75.44.7/sbclcm-auto/cssu_archive.zip",
            "upgrade_server_credentials": ""
        }

    # Supported type: ['instantiation', 'su', 'dr', 'cssu']
    # rel is needed to handle SU case, SU TO load may be different release
    # def gen_arts_sig(type='instantiation', rel='R20.0'):
    def gen_arts_sig(self):
        logger.info('In Func:' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)

        if self.type not in self.arts_type:
            logger.info('Only ' + str(self.arts_type) + ' Supported.')
            exit(1)

        ts = int(time.time())
        logger.info('Current time is: ' + str(ts))

        sig_yact_dir = self.yact_dir + str(ts) + '/'
        cmd = 'mkdir -p ' + sig_yact_dir
        ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)

        if self.type == 'instantiation':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_dif_name
        elif self.type == 'dr':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_dr_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_dr_dif_name
        elif self.type == 'su':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_su_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_su_dif_name
        elif self.type == 'cssu':
            sig_dif_file = sig_arts_dir + '\\' + self.sig_cssu_dif_name
            sig_rmt_dif_file = sig_yact_dir + self.sig_cssu_dif_name

        logger.info('sig_dif_file: ' + sig_dif_file)
        logger.info('sig_rmt_dif_file: ' + sig_rmt_dif_file)
        ssh_scp_put(self.yact_server_ip, self.yact_user, self.yact_passwd, sig_dif_file, sig_rmt_dif_file)

        yact_cmd = '/home/yact/YACT/yact.sh gen-by-dif ' + sig_rmt_dif_file + ' SBC-signaling ' + self.rel
        logger.info('yact_cmd: ' + yact_cmd)
        cmd = 'cd ' + sig_yact_dir + ';' + yact_cmd
        result = ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)
        logger.info('result of yact_cmd: '+ result)

        # Handle SU case
        if self.type == 'su':
            pkg_tools_dir = sig_yact_dir + self.sig_ne_name + '/package/pkg_tools/'
            logger.info('su pkg_tools_dir: ' + pkg_tools_dir)

            # ship from load vnf pkg to pkg_tools_dir and rename it to vnfpkg-from-load.zip
            vnfpkg_from_load = sig_data_dir + '\\' + self.sig_vnfpkg_name
            rmt_vnfpkg_from_load = pkg_tools_dir + 'vnfpkg-from-load.zip'
            ssh_scp_put(self.yact_server_ip, self.yact_user, self.yact_passwd, vnfpkg_from_load, rmt_vnfpkg_from_load)

            # run vnf_upgrade_pkg_gen
            cmd = 'cd ' + pkg_tools_dir + ';' + './vnf_upgrade_pkg_gen -n ../Nokia_sig_SBC-VNF_Package.zip ' \
                                                '-c ./vnfpkg-from-load.zip ' + self.sig_ne_name
            logger.info('vnf_upgrade_pkg_gen cmd: ' + cmd)
            ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)

            # now ship the upgrade vnf pkg to local
            # example: sbclcm03_Nokia_sig_SBC_upgrade-VNF_Package.zip
            vnfpkg_upgrade = pkg_tools_dir + self.sig_ne_name + '_Nokia_sig_SBC_upgrade-VNF_Package.zip'
            logger.info('upgrade package: ' + vnfpkg_upgrade)
            vnfpkg_upgrade_local = sig_data_dir + '\\' + self.sig_su_vnfpkg_name
            ssh_scp_get(self.yact_server_ip, self.yact_user, self.yact_passwd, vnfpkg_upgrade, vnfpkg_upgrade_local)
        # handle other cases
        else:
            # zip remote sig_art_files
            cmd = 'cd ' + sig_yact_dir + ';' + 'zip -r ' + self.sig_ne_name + '.zip ' + self.sig_ne_name
            result = ssh_command(cmd, self.yact_server_ip, self.yact_user, self.yact_passwd, self.ssh_type)
            logger.info('result of zip sig: ' + result)

            # remove sig zip if exists
            if os.path.exists(self.sig_zip):
                os.remove(self.sig_zip)

            # remote the unziped dir if exists
            if os.path.exists(self.sig_ne_name):
                shutil.rmtree(self.sig_ne_name)

            rmt_zip_file = sig_yact_dir + self.sig_zip
            ssh_scp_get(self.yact_server_ip, self.yact_user, self.yact_passwd, rmt_zip_file, self.sig_zip)

            logger.info('unzip '+ self.sig_zip)
            zip_file = zipfile.ZipFile(self.sig_zip)
            # os.mkdir(sig_ne_name)
            for names in zip_file.namelist():
                zip_file.extract(names, '.')
            zip_file.close()

    def ship_bulkconf_zip(self):
        # scp the bulkconf_artifacts.zip to remote backup server
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            logger.info('artifacts dir not exist. Exit.')
            exit(1)

        # ship the bulkconf_artifacts.zip
        local_bulkconf_file = sig_arts_dir + '\\' + self.sig_ne_name + r'\bulk_netconf\scale\bulkconf_artifacts.zip'
        rmt_bulkconf_file = self.backup_server_dir + 'bulkconf_artifacts.zip'
        logger.info('local_bulkconf_file: ' + local_bulkconf_file)
        logger.info('rmt_bulkconf_file: ' + rmt_bulkconf_file)
        logger.info('ship bulkconf_artifacts.zip to remote backup server.')
        ssh_scp_put(self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                    local_bulkconf_file, rmt_bulkconf_file)

        # chmod
        cmd = 'cd ' + self.backup_server_dir + '; chmod 777 bulkconf_artifacts.zip'
        result = ssh_command(cmd, self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                             self.ssh_type)
        logger.info('result of chmod bulkconf_artifacts.zip: ' + result)

    # def cp_vnf_pkg(type='instantiation'):
    def cp_vnf_pkg(self):
        # cp the Nokia_sig_SBC-VNF_Package.zip to data\sig-plane
        # to correct name
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            logger.info('artifacts dir not exist. Exit.')
            exit(1)

        srcfile = self.sig_ne_name + r'\package\Nokia_sig_SBC-VNF_Package.zip'

        if self.type == 'instantiation':
            dstfile = sig_data_dir + '\\' + self.sig_vnfpkg_name
        elif self.type == 'dr':
            dstfile = sig_data_dir + '\\' + self.sig_dr_vnfpkg_name
        elif self.type == 'su':
            dstfile = sig_data_dir + '\\' + self.sig_su_vnfpkg_name
        elif self.type == 'cssu':
            dstfile = sig_data_dir + '\\' + self.sig_cssu_vnfpkg_name

        logger.info('Copy ' + srcfile + ' to ' + dstfile)
        shutil.copyfile(srcfile, dstfile)

    # json load, dump
    # with open(file_name, 'r') as f:
    #     data = json.load(f)
    # with open('output.json', 'w') as f:
    #     json.dump(data, f)
    # def prep_instantiation_json(type='instantiation'):
    def prep_instantiation_json(self):
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_ne_name):
            logger.info('artifacts dir not exist. Exit.')
            exit(1)

        json_file = self.sig_ne_name + r'\package\cbam_json' + '\\' + self.instantiation_json_file

        # In case of dr, the instantiation json needs to be extracted from backup.zip
        # In case of su, instantiation json is not needed
        if self.type == 'dr':
            if os.path.exists(self.backup_file_name):
                os.remove(self.backup_file_name)

            if os.path.exists(self.local_backup_dir):
                shutil.rmtree(self.local_backup_dir)

            ssh_scp_get(self.backup_server_ip, self.backup_server_login, self.backup_server_passwd,
                        self.backup_file, self.backup_file_name)

            logger.info('unzip ' + self.backup_file_name)
            os.mkdir(self.local_backup_dir)
            zip_file = zipfile.ZipFile(self.backup_file_name)
            for names in zip_file.namelist():
                zip_file.extract(names, self.local_backup_dir)
            zip_file.close()

            json_file = self.local_backup_dir + '\\' + self.instantiation_json_file

        # load json data
        try:
            with open(json_file, 'r') as jFile:
                data = json.load(jFile)
        except Exception as exc:
            print('ERROR: Failed to load data from {0}.\n[{1}]'.format(
                json_file, str(exc)))

        # change data
        data['vimConnectionInfo'][0]['accessInfo']['password'] = self.os_passwd

        if self.type == 'instantiation':
            data['additionalParams'] = self.ap_instantiation
            output_json = sig_data_dir + '\\' + self.instantiation_json_file
        elif self.type == 'dr':
            data['additionalParams'] = self.ap_dr_instantiation
            output_json = sig_data_dir + '\\' + self.dr_instantiation_json_file
        elif self.type == 'cssu':
            data['additionalParams'] = self.ap_cssu_instantiation
            output_json = sig_data_dir + '\\' + self.cssu_instantiation_json_file

        # dump modified data
        try:
            with open(output_json, 'w') as jFile:
                json.dump(data, jFile, indent=2)
        except Exception as exc:
            print('ERROR: Failed to write data to {0}.\n[{1}]'.format(
                output_json, str(exc)))

        logger.info('instantiation json for ' + self.type + ' created completed.')

    # get the sc count of provided workbook
    # return: the row of sc vm_count, the vm_count value
    def get_sc_count(self, workbook):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        wb = load_workbook(workbook, keep_vba=True)
        sh = wb['DP']
        colB = sh['B']

        row_list = []
        for i in colB:
            if i.value == 'group_tag':
                row_list.append(i.row)

        row_sc = None
        for r in row_list:
            if sh.cell(row=r, column=26).value == 'SC':
                row_sc = r
        if row_sc is not None:
            logger.info('row_sc: ' + str(row_sc))

        row_vm_count = None
        sc_count = None
        for r in range(row_sc, row_sc + 8):
            if sh.cell(row=r, column=2).value == 'vm_count':
                row_vm_count = r
        if row_vm_count is not None:
            logger.info('row_vm_count: ' + str(row_vm_count))

        sc_count = sh.cell(row=row_vm_count, column=26).value
        logger.info('get_sc_count SC vaule of vm_count: ' + str(sc_count))

        return (row_vm_count, sc_count)

    # set the sc count of provided workbook
    def set_sc_count(self, workbook, sc_count=5):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        row_vm_count, initial_sc_count = self.get_sc_count(workbook)

        wb = load_workbook(workbook, keep_vba=True)
        sh = wb['DP']

        # set vm_count value to updated value
        sh.cell(row=row_vm_count, column=26).value = sc_count

        # wb.save(filename='output.xlsm')
        wb.save(filename=workbook)

        logger.info('Successfully set sc count to: ' + str(sc_count))

    # get initial sc value
    def get_initial_sc_value(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        os.chdir(sig_arts_dir)
        if not os.path.exists(self.sig_dif_name):
            logger.info('Original DIF does not exist. Exit.')
            exit(1)

        workbook = self.sig_dif_name

        row_vm_count, initial_sc_count = self.get_sc_count(workbook)

        logger.info('initial_sc_count: ' + str(initial_sc_count))
        return initial_sc_count

    # update DIF using updated sc value
    def gen_dif_updated_sc_value(self, sc_count=5):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        logger.info('To set sc count to: ' + str(sc_count))
        os.chdir(sig_arts_dir)

        if self.type == 'su':
            workbook = self.sig_su_dif_name
        elif self.type == 'cssu':
            workbook = self.sig_cssu_dif_name
        elif self.type == 'dr':
            workbook = self.sig_dr_dif_name

        if not os.path.exists(workbook):
            logger.info('workbook does not exist. Exit.')
            exit(1)

        self.set_sc_count(workbook, sc_count)

    # prep arts for instantiation, dr,
    def prep_arts(self):
        logger.info('In Func: ' + sys._getframe().f_code.co_name)
        # su doesn't need json file, only vnf pkg needed
        if self.type == 'su':
            self.gen_arts_sig()
        else:
            self.gen_arts_sig()
            self.prep_instantiation_json()
            self.cp_vnf_pkg()
        # needs to ship bulkconf artifacts if instantiation
        if self.type == 'instantiation':
            self.ship_bulkconf_zip()

def sigvnf_GenArtsInstantiation():
    sigArts = SigVnfArtsGenerator(type='instantiation', rel='R20.0')
    sigArts.prep_arts()

def sigvnf_GenArtsDR():
    sigArts = SigVnfArtsGenerator(type='dr', rel='R20.0')
    sigArts.prep_arts()

def sigvnf_GenArtsSU():
    sigArts = SigVnfArtsGenerator(type='su', rel='R20.0')
    sigArts.prep_arts()

def sigvnf_GenArtsCSSU():
    sigArts = SigVnfArtsGenerator(type='cssu', rel='R20.0')
    sigArts.prep_arts()

def sigvnf_GenArts():
    sigvnf_GenArtsInstantiation()
    sigvnf_GenArtsDR()
    sigvnf_GenArtsSU()
    sigvnf_GenArtsCSSU()


if __name__ == '__main__':

    setup_logging()

    # sigvnf_GenArts()

    sigArts = SigVnfArtsGenerator(type='dr', rel='R20.0')
    sigArts.get_initial_sc_value()
    sigArts.gen_dif_updated_sc_value(sc_count=3)

    sigArts = SigVnfArtsGenerator(type='cssu', rel='R20.0')
    sigArts.gen_dif_updated_sc_value(sc_count=4)

    sigArts = SigVnfArtsGenerator(type='su', rel='R20.0')
    sigArts.gen_dif_updated_sc_value(sc_count=2)




